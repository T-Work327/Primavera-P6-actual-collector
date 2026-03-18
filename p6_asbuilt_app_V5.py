"""
Primavera P6 Asbuilt Data Collector — Streamlit Web App
=========================================================
Run with:
    streamlit run p6_asbuilt_app.py

Requirements:
    pip install streamlit openpyxl

──────────────────────────────────────────────────────────
USER MANAGEMENT
──────────────────────────────────────────────────────────
Edit the USERS dictionary below to add, remove, or change
passwords and roles. Passwords are stored as SHA-256 hashes.

To generate a hash for a new password, run in Python:
    import hashlib
    print(hashlib.sha256("yourpassword".encode()).hexdigest())

Roles:
  "readonly"   — View entries only
  "readwrite"  — View + Submit/Update + Import from Excel + Photo Log
  "admin"      — All of the above + Export to Excel
──────────────────────────────────────────────────────────
"""

import hashlib
import io
import json
import uuid
from datetime import date, datetime, time
from pathlib import Path

import streamlit as st

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    st.error("openpyxl is required.  Run:  pip install openpyxl")
    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
# USER DEFINITIONS  —  edit this section to manage users
# ══════════════════════════════════════════════════════════════════════════════
#
# To change a password:
#   1. Run in Python:  import hashlib; print(hashlib.sha256("newpassword".encode()).hexdigest())
#   2. Replace the hash string below
#
# To add a user:  copy any line and change username, hash, role, and display name
# To remove a user:  delete their entry

def _h(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

USERS = {
    #  username       password hash           role           display name
    "admin":    {"hash": _h("admin123"),    "role": "admin",     "name": "Administrator"},
    "engineer": {"hash": _h("engineer1"),   "role": "readwrite", "name": "Site Engineer"},
    "viewer":   {"hash": _h("viewer1"),     "role": "readonly",  "name": "Project Viewer"},
}

# ── Role Permission Matrix ─────────────────────────────────────────────────────
PERMISSIONS = {
    "readonly":  {"view"},
    "readwrite": {"view", "submit", "import", "photos"},
    "admin":     {"view", "submit", "import", "export", "photos"},
}

def has_permission(perm: str) -> bool:
    return perm in PERMISSIONS.get(st.session_state.get("role", ""), set())

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

DATA_FILE  = Path("p6_asbuilt_store.json")
PHOTO_DIR  = Path("p6_images")
PHOTO_FILE = Path("p6_photo_log.json")

USER_DATA = "DurationQtyType=QT_Day\nShowAsPercentage=0\nSmallScaleQtyType=QT_Hour\nDateFormat=dd/mm/yyyy\nCurrencyFormat=US Dollar"

STATUS_OPTIONS = ["Not Started", "In Progress", "Completed"]

STATUS_COLOUR = {
    "Not Started": "#6b7280",
    "In Progress":  "#d97706",
    "Completed":    "#16a34a",
}

ROLE_LABEL = {
    "readonly":  "Read Only",
    "readwrite": "Read / Write",
    "admin":     "Admin",
}

# P6 internal field key names (row 1 of TASK sheet)
P6_FIELD_KEYS = [
    "task_code", "task_name", "status_code", "act_start_date",
    "act_end_date", "complete_pct", "remain_drtn_hr_cnt",
    "complete_pct_type", "wbs_id", "user_field_910",
]

# Column definitions: (display header, column width, data key)
P6_COLUMNS = [
    ("Activity ID",          14, "activity_id"),
    ("Activity Name",        36, "activity_name"),
    ("Activity Status",      16, "activity_status"),
    ("Actual Start",         20, "actual_start"),
    ("Actual Finish",        20, "actual_finish"),
    ("Duration % Complete",  20, "pct_complete"),
    ("Remaining Duration",   20, "remaining_dur"),
    ("Percent Complete Type",20, "complete_pct_type"),
    ("WBS Code",             36, "wbs_id"),
    ("Comments",             50, "comments_export"),
]

DATE_KEYS = {"actual_start", "actual_finish"}

# ══════════════════════════════════════════════════════════════════════════════
# DATE HELPERS
# Dates stored in JSON as ISO strings: "YYYY-MM-DDTHH:MM:00"
# ══════════════════════════════════════════════════════════════════════════════

def dt_to_iso(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%dT%H:%M:00")

def iso_to_dt(value: str) -> datetime | None:
    if not value or str(value).strip() == "":
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:00", "%Y-%m-%dT%H:%M",
                "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d",
                "%d-%b-%y %H:%M", "%d-%b-%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt)
        except ValueError:
            continue
    return None

def display_dt(value: str) -> str:
    dt = iso_to_dt(value)
    return dt.strftime("%d/%m/%Y %H:%M") if dt else "—"

def normalise_imported_date(raw_val) -> str:
    if raw_val is None or str(raw_val).strip() == "":
        return ""
    if isinstance(raw_val, datetime):
        return dt_to_iso(raw_val)
    dt = iso_to_dt(str(raw_val))
    return dt_to_iso(dt) if dt else str(raw_val).strip()


# ══════════════════════════════════════════════════════════════════════════════
# COMMENT HELPERS
# Comments are stored on each entry as a list of dicts:
#   _comments: [{"text": "...", "by": "...", "at": "DD/MM/YYYY HH:MM"}, ...]
# Newest first.  Exported to P6 as a single ';'-separated string (no timestamps).
# ══════════════════════════════════════════════════════════════════════════════

def comments_to_export(comments: list[dict]) -> str:
    """Flatten comment list → '; '-joined string, newest first, no timestamps."""
    return "; ".join(c["text"] for c in comments if c.get("text", "").strip())

def import_string_to_comments(raw: str, imported_by: str) -> list[dict]:
    """
    Split a '; '-separated import string into individual comment records.
    Each segment gets the same import timestamp and is marked as imported.
    Order is preserved (P6 exports newest first, so we keep that).
    """
    if not raw or not raw.strip():
        return []
    segments = [s.strip() for s in raw.split(";") if s.strip()]
    ts = datetime.now().strftime("%d/%m/%Y %H:%M")
    return [{"text": seg, "by": f"{imported_by} (imported)", "at": ts}
            for seg in segments]

# ══════════════════════════════════════════════════════════════════════════════
# STORAGE
# ══════════════════════════════════════════════════════════════════════════════

def load_entries() -> list[dict]:
    if DATA_FILE.exists():
        try:
            return json.loads(DATA_FILE.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return []
    return []

def save_entries(entries: list[dict]) -> None:
    DATA_FILE.write_text(json.dumps(entries, ensure_ascii=False, indent=2), encoding="utf-8")

def upsert_entry(entries: list[dict], new: dict) -> tuple:
    idx = next(
        (i for i, e in enumerate(entries)
         if e.get("activity_id", "").upper() == new.get("activity_id", "").upper()),
        None,
    )
    if idx is not None:
        entries[idx] = new
        return entries, "updated"
    entries.append(new)
    return entries, "saved"

# ══════════════════════════════════════════════════════════════════════════════
# PHOTO STORAGE
# ══════════════════════════════════════════════════════════════════════════════

def ensure_photo_dir() -> None:
    PHOTO_DIR.mkdir(exist_ok=True)

def load_photos() -> list[dict]:
    if PHOTO_FILE.exists():
        try:
            return json.loads(PHOTO_FILE.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return []
    return []

def save_photos(photos: list[dict]) -> None:
    PHOTO_FILE.write_text(json.dumps(photos, ensure_ascii=False, indent=2), encoding="utf-8")

def add_photo(activity_id: str, activity_name: str, photo_date: date,
              comment: str, file_bytes: bytes, original_name: str,
              uploaded_by: str) -> dict:
    ensure_photo_dir()
    ext      = Path(original_name).suffix.lower() or ".jpg"
    filename = f"{uuid.uuid4().hex}{ext}"
    (PHOTO_DIR / filename).write_bytes(file_bytes)
    record = {
        "id":            uuid.uuid4().hex,
        "activity_id":   activity_id,
        "activity_name": activity_name,
        "photo_date":    photo_date.isoformat(),
        "comment":       comment.strip(),
        "filename":      filename,
        "uploaded_at":   datetime.now().strftime("%d/%m/%Y %H:%M"),
        "uploaded_by":   uploaded_by,
    }
    photos = load_photos()
    photos.append(record)
    save_photos(photos)
    return record

def delete_photo(photo_id: str) -> None:
    photos = load_photos()
    record = next((p for p in photos if p["id"] == photo_id), None)
    if record:
        img_path = PHOTO_DIR / record["filename"]
        if img_path.exists():
            img_path.unlink()
        photos = [p for p in photos if p["id"] != photo_id]
        save_photos(photos)

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════

# P6 internal status codes — must NOT contain spaces (0x20) or P6 rejects the import.
STATUS_TO_P6 = {
    "Not Started": "TK_NotStart",
    "In Progress":  "TK_Active",
    "Completed":    "TK_Complete",
}

THIN   = Side(style="thin", color="B0B8C8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

@st.cache_data
def build_excel(entries: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TASK"
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

    # Row 1 — P6 internal field keys
    for col_idx, (key, (_, width, _)) in enumerate(zip(P6_FIELD_KEYS, P6_COLUMNS), start=1):
        c = ws.cell(row=1, column=col_idx, value=key)
        c.font      = Font(name="Arial", italic=True, color="4472C4", size=9)
        c.fill      = PatternFill("solid", fgColor="D9E1F2")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Row 2 — Human-readable column headers
    for col_idx, (header, _, _) in enumerate(P6_COLUMNS, start=1):
        c = ws.cell(row=2, column=col_idx, value=header)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", fgColor="1C3557")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = BORDER

    # Data rows
    for row_idx, entry in enumerate(entries, start=3):
        fill = PatternFill("solid", fgColor="EEF2F8" if row_idx % 2 == 0 else "FFFFFF")
        for col_idx, (_, _, key) in enumerate(P6_COLUMNS, start=1):
            value = entry.get(key, "")
            if key in DATE_KEYS:
                dt_val = iso_to_dt(value)
                c = ws.cell(row=row_idx, column=col_idx, value=dt_val)
                if dt_val:
                    c.number_format = "DD/MM/YYYY HH:MM"
            elif key == "complete_pct_type":
                c = ws.cell(row=row_idx, column=col_idx, value="Physical")
            elif key == "comments_export":
                # Build export string from stored comment list (newest first, no timestamps)
                c = ws.cell(row=row_idx, column=col_idx,
                            value=comments_to_export(entry.get("_comments", [])))
            else:
                c = ws.cell(row=row_idx, column=col_idx, value=str(value) if value != "" else "")
            c.fill   = fill
            c.border = BORDER
            c.alignment = Alignment(vertical="center", wrap_text=False)
            if col_idx <= 2:
                c.fill = PatternFill("solid", fgColor="F2F2F2")
                c.font = Font(name="Arial", size=10, bold=(col_idx == 1))
            else:
                c.font = Font(name="Arial", size=10, color="1F4E79")
            if key in ("pct_complete", "remaining_dur") and value != "":
                c.alignment = Alignment(horizontal="right", vertical="center")

    # USERDATA sheet — do not modify this section, P6 is very particular about it
    wu  = wb.create_sheet("USERDATA")
    wu.column_dimensions["A"].width = 60
    b2  = Border(left=Side(style="thin", color="B0B8C8"), right=Side(style="thin", color="B0B8C8"),
                 top=Side(style="thin",  color="B0B8C8"), bottom=Side(style="thin", color="B0B8C8"))
    # Row 1: field key identifier (no spaces — required by P6)
    r1 = wu.cell(row=1, column=1, value="user_data")
    r1.font   = Font(name="Arial", bold=True, size=9, color="4472C4")
    r1.fill   = PatternFill("solid", fgColor="D9E1F2")
    r1.border = b2
    # Row 2: section label
    r2 = wu.cell(row=2, column=1, value="UserSettings Do Not Edit")
    r2.font   = Font(name="Arial", bold=True, size=11, color="1C3557")
    r2.fill   = PatternFill("solid", fgColor="D9E1F2")
    r2.border = b2
    # Row 3: settings values
    r3 = wu.cell(row=3, column=1,
                 value=USER_DATA)
    r3.font      = Font(name="Arial", size=10)
    r3.fill      = PatternFill("solid", fgColor="F8F9FB")
    r3.alignment = Alignment(vertical="top", wrap_text=True)
    r3.border    = b2
    wu.row_dimensions[3].height = 80

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL IMPORT
# ══════════════════════════════════════════════════════════════════════════════

P6_KEY_MAP = {
    "task_code": "activity_id", "task_name": "activity_name",
    "status_code": "activity_status", "act_start_date": "actual_start",
    "act_end_date": "actual_finish", "complete_pct": "pct_complete",
    "remain_drtn_hr_cnt": "remaining_dur", "complete_pct_type": "complete_pct_type",
    "wbs_id": "wbs_id", "user_field_910": "comments_import",
}
HEADER_KEY_MAP = {
    "activity id": "activity_id", "activity name": "activity_name",
    "activity status": "activity_status", "actual start": "actual_start",
    "actual finish": "actual_finish", "duration % complete": "pct_complete",
    "remaining duration": "remaining_dur", "percent complete type": "complete_pct_type",
    "wbs code": "wbs_id", "comments": "comments_import",
}

def read_p6_excel(file_bytes: bytes) -> tuple:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    warnings_list = []
    sheet_name = next((s for s in wb.sheetnames if s.upper() == "TASK"), None)
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
        warnings_list.append(f"No TASK sheet found — reading from '{sheet_name}' instead.")
    ws = wb[sheet_name]
    rows_iter = list(ws.iter_rows(values_only=True))
    if not rows_iter:
        return [], ["The sheet appears to be empty."]
    col_map, data_start = {}, 1
    for row_idx in range(min(3, len(rows_iter))):
        row = rows_iter[row_idx]
        mapping = {}
        for col_idx, cell_val in enumerate(row):
            if cell_val is None:
                continue
            cs = str(cell_val).strip().lower()
            if cs in P6_KEY_MAP:
                dk = P6_KEY_MAP[cs]
                if dk != "complete_pct_type":  # always force Physical on export; ignore on import
                    mapping[col_idx] = dk
            elif cs in HEADER_KEY_MAP:
                dk = HEADER_KEY_MAP[cs]
                if dk != "complete_pct_type":
                    mapping[col_idx] = dk
        if mapping:
            col_map, data_start = mapping, row_idx + 1
            if row_idx == 0 and len(rows_iter) > 1:
                next_str = [str(v).strip().lower() for v in rows_iter[1] if v is not None]
                if any(h in HEADER_KEY_MAP for h in next_str):
                    data_start = 2
            break
    if not col_map:
        return [], ["Could not detect column headers."]
    entries = []
    for row in rows_iter[data_start:]:
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        entry = {
            "activity_id": "", "activity_name": "", "activity_status": "",
            "actual_start": "", "actual_finish": "", "pct_complete": "",
            "remaining_dur": "", "complete_pct_type": "Physical", "wbs_id": "",
            "comments_import": "",
        }
        for col_idx, data_key in col_map.items():
            if col_idx >= len(row):
                continue
            raw_val = row[col_idx]
            if data_key in DATE_KEYS:
                entry[data_key] = normalise_imported_date(raw_val)
            elif data_key == "pct_complete":
                vs = str(raw_val).replace("%", "").strip() if raw_val is not None else ""
                try:
                    entry[data_key] = str(int(float(vs))) if vs else ""
                except ValueError:
                    entry[data_key] = vs
            else:
                entry[data_key] = "" if raw_val is None else str(raw_val).strip()
        if not entry["activity_id"]:
            continue
        if not entry["complete_pct_type"]:
            entry["complete_pct_type"] = "Physical"
        entry["_submitted_at"] = datetime.now().strftime("%d/%m/%Y %H:%M")
        entries.append(entry)
    return entries, warnings_list

# ══════════════════════════════════════════════════════════════════════════════
# DATE INPUT WIDGET HELPER
# Uses st.datetime_input (Streamlit >= 1.43).
# Returns a datetime, or None if the user has not enabled an optional field.
# ══════════════════════════════════════════════════════════════════════════════

def datetime_inputs(label: str, key: str, required: bool = True,
                    default_dt: datetime | None = None) -> datetime | None:
    """
    Render a single datetime picker (st.datetime_input).
    For optional fields a checkbox gates the widget; returns None when unchecked.
    """
    default_val = default_dt if default_dt else datetime.combine(date.today(), time(8, 0))

    if not required:
        enabled = st.checkbox(f"Set {label}", key=f"{key}_enabled",
                              value=(default_dt is not None))
        if not enabled:
            return None

    return st.datetime_input(label, value=default_val, key=f"{key}_dt",
                              step=60 * 15)   # 15-minute steps

# ══════════════════════════════════════════════════════════════════════════════
# PAGE SETUP
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="P6 Asbuilt Collector", page_icon="🏗️", layout="wide")

# ══════════════════════════════════════════════════════════════════════════════
# AUTH — SESSION STATE & LOGIN SCREEN
# ══════════════════════════════════════════════════════════════════════════════

if "authenticated" not in st.session_state:
    st.session_state.update({"authenticated": False, "username": "",
                              "display_name": "", "role": ""})

if not st.session_state.authenticated:
    st.title("P6 Asbuilt Collector")
    st.caption("Sign in to continue")
    st.divider()

    _, col_m, _ = st.columns([1, 1.1, 1])
    with col_m:
        with st.container(border=True):
            st.subheader("Sign In")
            username = st.text_input("Username", placeholder="Enter your username")
            password = st.text_input("Password", type="password", placeholder="Enter your password")
            if st.button("Log In", type="primary", use_container_width=True):
                user = USERS.get(username)
                if user and user["hash"] == hashlib.sha256(password.encode()).hexdigest():
                    st.session_state.update({
                        "authenticated": True, "username": username,
                        "display_name": user["name"], "role": user["role"],
                    })
                    st.rerun()
                else:
                    st.error("Incorrect username or password.")
    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR (authenticated)
# ══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.title("🏗️ P6 Asbuilt")
    st.divider()
    st.write(f"**{st.session_state.display_name}**")
    st.caption(ROLE_LABEL.get(st.session_state.role, st.session_state.role))
    if st.button("Log Out", use_container_width=True):
        st.session_state.update({"authenticated": False, "username": "",
                                  "display_name": "", "role": ""})
        st.rerun()
    st.divider()
    st.caption("Entries: p6_asbuilt_store.json")
    st.caption("Photos:  p6_images/")

# ══════════════════════════════════════════════════════════════════════════════
# HEADER & DYNAMIC TABS
# ══════════════════════════════════════════════════════════════════════════════

st.title("🏗️ Primavera P6 — Asbuilt Data Collector")
st.caption("Submit and update asbuilt progress entries, then export a P6-compatible spreadsheet.")
st.divider()
logo=Path("Tricertus_logo.jpg")
st.logo(logo,size="large")

TAB_DEFS = [
    ("📋  View All Entries",  "view"),
    ("📝  Submit / Update",   "submit"),
    ("📤  Import from Excel", "import"),
    ("📥  Export to Excel",   "export"),
    ("📸  Photo Log",         "photos"),
]
visible   = [(lbl, perm) for lbl, perm in TAB_DEFS if has_permission(perm)]
tab_objs  = st.tabs([lbl for lbl, _ in visible])
tab_index = {perm: tab_objs[i] for i, (_, perm) in enumerate(visible)}

# ══════════════════════════════════════════════════════════════════════════════
# TAB: VIEW ALL ENTRIES
# ══════════════════════════════════════════════════════════════════════════════

with tab_index["view"]:
    entries = load_entries()
    st.subheader(f"All Entries ({len(entries)})")

    if not entries:
        st.info("No entries yet.")
    else:
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total",       len(entries))
        m2.metric("Completed",   sum(1 for e in entries if e.get("activity_status") == "Completed"))
        m3.metric("In Progress", sum(1 for e in entries if e.get("activity_status") == "In Progress"))
        m4.metric("Not Started", sum(1 for e in entries if e.get("activity_status") == "Not Started"))
        st.divider()

        # ── Sort controls ──────────────────────────────────────────────────
        sort_col, dir_col = st.columns([3, 1])
        with sort_col:
            sort_by = st.selectbox(
                "Sort by",
                options=["WBS Code", "Actual Start", "Actual Finish", "Activity ID"],
                key="view_sort_by",
            )
        with dir_col:
            sort_asc = st.radio(
                "Order", options=["↑ Asc", "↓ Desc"],
                key="view_sort_dir", horizontal=True,
            ) == "↑ Asc"

        def wbs_key(e: dict):
            parts = e.get("wbs_id", "").split(".")
            segments = []
            for p in parts:
                try:
                    segments.append((0, int(p)))
                except ValueError:
                    segments.append((1, p.lower()))
            return segments or [(1, "")]

        def date_key(field: str):
            def _key(e: dict):
                dt = iso_to_dt(e.get(field, ""))
                return dt if dt else (datetime.min if sort_asc else datetime.max)
            return _key

        if sort_by == "WBS Code":
            sorted_entries = sorted(entries, key=wbs_key, reverse=not sort_asc)
        elif sort_by == "Actual Start":
            sorted_entries = sorted(entries, key=date_key("actual_start"), reverse=not sort_asc)
        elif sort_by == "Actual Finish":
            sorted_entries = sorted(entries, key=date_key("actual_finish"), reverse=not sort_asc)
        else:
            sorted_entries = sorted(entries, key=lambda e: e.get("activity_id", "").upper(), reverse=not sort_asc)

        st.divider()

        can_edit   = has_permission("submit")
        can_delete = has_permission("submit")

        for entry in sorted_entries:
            i = entries.index(entry)
            status    = entry.get("activity_status", "")
            act_id    = entry.get("activity_id", "")
            act_name  = entry.get("activity_name", "")
            wbs       = entry.get("wbs_id", "—")
            pct       = entry.get("pct_complete", "0")
            rem       = entry.get("remaining_dur", "—")
            a_start   = display_dt(entry.get("actual_start", ""))
            a_finish  = display_dt(entry.get("actual_finish", ""))
            subm_at   = entry.get("_submitted_at", "")
            submitter = entry.get("_submitted_by", "")
            n_comments = len(entry.get("_comments", []))

            with st.container(border=True):
                # ── Header row ─────────────────────────────────────────────
                head_left, head_left2, head_right = st.columns([2, 2, 1])
                with head_left:
                    st.write(f"Activity ID: {act_id}")
                with head_left2:
                    st.write(f"Activity Name: {act_name}")
                with head_right:
                    st.write(status)

                # ── Detail row ─────────────────────────────────────────────
                c1, c2, c3, c4, c5 = st.columns([1, 1, 1, 0.3, 0.4], gap="xsmall")
                c1.write("WBS");      c1.write(wbs)
                c2.write("Start");    c2.write(a_start)
                c3.write("Finish");   c3.write(a_finish)
                c4.metric("% Complete", f"{pct}%")
                c5.metric("Remaining", f"{rem} days" if rem and rem != "—" else "—")

                # ── Footer ─────────────────────────────────────────────────
                footer = f"Last updated: {subm_at}"
                if submitter:
                    footer += f"  ·  By: {submitter}"
                if n_comments:
                    footer += f"  ·  💬 {n_comments} comment{'s' if n_comments != 1 else ''}"
                st.caption(footer)

                # ── Inline edit expander (submit/admin only) ───────────────
                if can_edit:
                    with st.expander("✏️  Edit name / Add comment"):
                        # Activity name
                        st.write("**Edit Activity Name**")
                        new_name = st.text_input(
                            "Activity Name",
                            value=act_name,
                            key=f"edit_name_{i}",
                            label_visibility="collapsed",
                        ).strip()

                        st.divider()

                        # Existing comments
                        st.write("**Comments**")
                        existing_comments = entry.get("_comments", [])
                        if existing_comments:
                            for c in existing_comments:
                                st.write(f"**{c['at']}** — {c['by']}")
                                st.write(c["text"])
                                st.divider()
                        else:
                            st.caption("No comments yet.")

                        new_comment_text = st.text_area(
                            "Add comment",
                            placeholder="Enter progress notes, observations, or issues...",
                            height=100,
                            key=f"view_comment_{i}",
                            label_visibility="collapsed",
                        ).strip()

                        # Save button — only active if something changed
                        name_changed    = new_name != act_name and new_name != ""
                        comment_entered = bool(new_comment_text)

                        if st.button(
                            "💾  Save changes",
                            key=f"edit_save_{i}",
                            type="primary",
                            disabled=not (name_changed or comment_entered),
                        ):
                            updated = entry.copy()
                            if name_changed:
                                updated["activity_name"] = new_name
                            if comment_entered:
                                new_record = {
                                    "text": new_comment_text,
                                    "by":   st.session_state.display_name,
                                    "at":   datetime.now().strftime("%d/%m/%Y %H:%M"),
                                }
                                updated["_comments"] = [new_record] + existing_comments
                            updated["_submitted_at"] = datetime.now().strftime("%d/%m/%Y %H:%M")
                            updated["_submitted_by"] = st.session_state.display_name
                            entries[i] = updated
                            save_entries(entries)
                            st.success("Saved.")
                            st.rerun()

                # ── Delete button ──────────────────────────────────────────
                if can_delete and st.button(f"🗑 Delete {act_id}", key=f"del_{i}"):
                    entries.pop(i)
                    save_entries(entries)
                    st.rerun()

        st.dataframe(sorted_entries)

# ══════════════════════════════════════════════════════════════════════════════
# TAB: SUBMIT / UPDATE
# ══════════════════════════════════════════════════════════════════════════════

if "submit" in tab_index:
    with tab_index["submit"]:
        entries   = load_entries()
        known_ids = {e["activity_id"].upper(): e for e in entries}

        st.subheader("Submit or Update an Asbuilt Entry")
        st.caption("Enter the Activity ID — if it already exists the name is filled automatically.")

        col_id, col_wbs = st.columns(2)
        with col_id:
            activity_id_raw = st.text_input("Activity ID *", placeholder="e.g. A1000").strip()
        with col_wbs:
            wbs_input = st.text_input("WBS ID *", placeholder="e.g. 1.2.3").strip()

        existing = known_ids.get(activity_id_raw.upper()) if activity_id_raw else None
        if existing:
            st.info(
                f"**Existing entry found:** {existing['activity_name']}  \n"
                f"Status: **{existing['activity_status']}** | "
                f"**{existing.get('pct_complete', 0)}%** complete  \n"
                f"Submitting will **update** this entry.", icon="ℹ️",
            )

        if existing:
            st.text_input("Activity Name", value=existing["activity_name"], disabled=True)
            activity_name = existing["activity_name"]
        else:
            activity_name = st.text_input(
                "Activity Name *", placeholder="e.g. Concrete Pour - Foundations"
            ).strip()

        activity_status = st.selectbox("Activity Status *", STATUS_OPTIONS)

        def _existing_dt(key: str) -> datetime | None:
            return iso_to_dt(existing.get(key, "")) if existing else None

        actual_start_dt  = None
        actual_finish_dt = None
        pct_complete     = 0
        remaining_dur    = ""

        if activity_status == "Not Started":
            st.info(
                "% Complete is set to 0 and Remaining Duration left blank automatically "
                "for 'Not Started' activities.", icon="ℹ️"
            )

        elif activity_status == "In Progress":
            actual_start_dt = datetime_inputs(
                "Actual Start *", key="start_ip", required=True,
                default_dt=_existing_dt("actual_start"),
            )
            col_p, col_r = st.columns(2)
            with col_p:
                pct_complete = st.number_input(
                    "Duration % Complete *", min_value=0, max_value=99, step=5,
                    value=int(existing.get("pct_complete", 0)) if existing else 0,
                )
            with col_r:
                remaining_dur = st.text_input(
                    "Remaining Duration (days) *", placeholder="e.g. 5",
                    value=existing.get("remaining_dur", "") if existing else "",
                ).strip()

        elif activity_status == "Completed":
            actual_start_dt = datetime_inputs(
                "Actual Start *", key="start_c", required=True,
                default_dt=_existing_dt("actual_start"),
            )
            actual_finish_dt = datetime_inputs(
                "Actual Finish *", key="finish_c", required=True,
                default_dt=_existing_dt("actual_finish"),
            )
            pct_complete  = 100
            remaining_dur = "0"
            st.info("% Complete set to 100 and Remaining Duration to 0 automatically.", icon="✅")

        # ── Comments section ──────────────────────────────────────────────
        st.divider()
        st.subheader("Comments")

        existing_comments = existing.get("_comments", []) if existing else []
        if existing_comments:
            st.caption(f"{len(existing_comments)} existing comment{'s' if len(existing_comments) != 1 else ''} stored for this activity:")
            for c in existing_comments:
                st.write(f"**{c['at']}** — {c['by']}")
                st.write(c["text"])
                st.divider()

        new_comment_text = st.text_area(
            "Add a new comment (optional)",
            placeholder="Enter progress notes, observations, or issues...",
            height=120,
            key="submit_new_comment",
        ).strip()

        if st.button("Submit Entry", type="primary"):
            errors = []
            if not activity_id_raw:                                                      errors.append("Activity ID is required.")
            if not wbs_input:                                                            errors.append("WBS ID is required.")
            if not existing and not activity_name:                                       errors.append("Activity Name is required for new activities.")
            if activity_status in ("In Progress", "Completed") and not actual_start_dt: errors.append("Actual Start is required.")
            if activity_status == "Completed" and not actual_finish_dt:                  errors.append("Actual Finish is required when status is Completed.")
            if activity_status == "In Progress" and not remaining_dur:                   errors.append("Remaining Duration is required when In Progress.")
            if actual_start_dt and actual_finish_dt and actual_finish_dt < actual_start_dt:
                errors.append("Actual Finish cannot be before Actual Start.")

            if errors:
                for e in errors:
                    st.error(e)
            else:
                # Build updated comment list — prepend new comment if provided
                updated_comments = list(existing.get("_comments", [])) if existing else []
                if new_comment_text:
                    updated_comments.insert(0, {
                        "text": new_comment_text,
                        "by":   st.session_state.display_name,
                        "at":   datetime.now().strftime("%d/%m/%Y %H:%M"),
                    })

                entry = {
                    "activity_id":       activity_id_raw,
                    "activity_name":     activity_name,
                    "activity_status":   activity_status,
                    "actual_start":      dt_to_iso(actual_start_dt)  if actual_start_dt  else "",
                    "actual_finish":     dt_to_iso(actual_finish_dt) if actual_finish_dt else "",
                    "pct_complete":      str(pct_complete),
                    "remaining_dur":     str(remaining_dur),
                    "complete_pct_type": "Physical",
                    "wbs_id":            wbs_input,
                    "_comments":         updated_comments,
                    "_submitted_at":     datetime.now().strftime("%d/%m/%Y %H:%M"),
                    "_submitted_by":     st.session_state.display_name,
                }
                entries, action = upsert_entry(entries, entry)
                save_entries(entries)
                icon = "✅" if action == "saved" else "🔄"
                st.success(f"{icon} Entry **{action}** successfully!")
                with st.expander("View saved data"):
                    display = entry.copy()
                    if display["actual_start"]:  display["actual_start"]  = display_dt(display["actual_start"])
                    if display["actual_finish"]: display["actual_finish"] = display_dt(display["actual_finish"])
                    st.json(display)

# ══════════════════════════════════════════════════════════════════════════════
# TAB: IMPORT FROM EXCEL
# ══════════════════════════════════════════════════════════════════════════════

if "import" in tab_index:
    with tab_index["import"]:
        entries   = load_entries()
        known_ids = {e["activity_id"].upper(): e for e in entries}

        st.subheader("Import from P6 Excel File")
        st.caption("Upload a P6-format XLSX. Conflicts will be shown for your decision before anything is saved.")

        uploaded = st.file_uploader("Choose a P6 XLSX file", type=["xlsx", "xls"])

        if uploaded:
            with st.spinner("Reading Excel file..."):
                imported_rows, warnings_list = read_p6_excel(uploaded.read())

            for w in warnings_list:
                st.warning(w)

            if not imported_rows:
                st.error("No valid rows found. Check the file has a TASK sheet with correct headers.")
            else:
                clean     = [r for r in imported_rows if r["activity_id"].upper() not in known_ids]
                conflicts = [r for r in imported_rows if r["activity_id"].upper() in known_ids]
                st.success(
                    f"Found **{len(imported_rows)}** rows: **{len(clean)}** new, "
                    f"**{len(conflicts)}** conflict{'s' if len(conflicts) != 1 else ''}."
                )

                resolutions = {}
                if conflicts:
                    st.divider()
                    st.subheader("⚠️ Conflicts — Activity ID already exists")
                    st.caption("Review each conflict and choose an action before confirming the import.")
                    for row in conflicts:
                        aid      = row["activity_id"].upper()
                        existing = known_ids[aid]
                        label    = f"🔁  {row['activity_id']}  —  {row.get('activity_name', existing.get('activity_name', ''))}"
                        with st.expander(label, expanded=True):
                            fields = [
                                ("Status",          "activity_status", False),
                                ("Actual Start",     "actual_start",    True),
                                ("Actual Finish",    "actual_finish",   True),
                                ("% Complete",       "pct_complete",    False),
                                ("Remaining (days)", "remaining_dur",   False),
                                ("WBS",              "wbs_id",          False),
                            ]
                            # Existing stored comments summary
                            existing_comment_count = len(existing.get("_comments", []))
                            incoming_comment_str   = row.get("comments_import", "").strip()
                            c_cur, c_new = st.columns(2)
                            with c_cur:
                                st.write("**Current (stored)**")
                                for lbl, k, is_date in fields:
                                    val = existing.get(k, "") or ""
                                    st.write(f"- **{lbl}:** {display_dt(val) if is_date else (val or '—')}")
                                st.write(f"- **Comments:** {existing_comment_count} stored comment{'s' if existing_comment_count != 1 else ''}")
                            with c_new:
                                st.write("**Incoming (from file)**")
                                for lbl, k, is_date in fields:
                                    val = row.get(k, "") or ""
                                    st.write(f"- **{lbl}:** {display_dt(val) if is_date else (val or '—')}")
                                st.write(f"- **Comments (user_field_910):** {incoming_comment_str or '—'}")

                            choice = st.radio(
                                "Activity data", ["Overwrite with incoming", "Keep current"],
                                key=f"conflict_{aid}", horizontal=True,
                            )
                            resolutions[aid] = "overwrite" if choice == "Overwrite with incoming" else "skip"

                            # Separate resolution for comments if both sides have data
                            if incoming_comment_str and existing_comment_count > 0:
                                comment_choice = st.radio(
                                    "Comments",
                                    ["Append imported comments to existing", "Keep existing only", "Replace with imported only"],
                                    key=f"comment_conflict_{aid}", horizontal=True,
                                )
                                resolutions[f"comment_{aid}"] = comment_choice
                            elif incoming_comment_str:
                                resolutions[f"comment_{aid}"] = "Append imported comments to existing"
                            else:
                                resolutions[f"comment_{aid}"] = "Keep existing only"

                st.divider()
                ow    = sum(1 for v in resolutions.values() if v == "overwrite")
                sk    = sum(1 for v in resolutions.values() if v == "skip")
                parts = [f"{len(clean)} new entries will be added"]
                if ow: parts.append(f"{ow} existing entries will be overwritten")
                if sk: parts.append(f"{sk} conflicts will be skipped")
                st.info("  ·  ".join(parts))

                if st.button("✅  Confirm Import", type="primary"):
                    entries = load_entries()
                    added = overwritten = skipped = 0
                    for row in clean:
                        # Convert any imported comment string to structured list
                        imp_str = row.pop("comments_import", "") or ""
                        row["_comments"] = import_string_to_comments(imp_str, st.session_state.display_name)
                        row["_submitted_by"] = st.session_state.display_name
                        entries.append(row)
                        added += 1
                    for row in conflicts:
                        aid = row["activity_id"].upper()
                        if resolutions.get(aid) == "overwrite":
                            idx = next(i for i, e in enumerate(entries) if e["activity_id"].upper() == aid)
                            imp_str = row.pop("comments_import", "") or ""
                            comment_res = resolutions.get(f"comment_{aid}", "Keep existing only")
                            existing_comments = entries[idx].get("_comments", [])
                            imported_comments = import_string_to_comments(imp_str, st.session_state.display_name)
                            if comment_res == "Append imported comments to existing":
                                # Imported go after existing (existing are newer)
                                row["_comments"] = existing_comments + imported_comments
                            elif comment_res == "Replace with imported only":
                                row["_comments"] = imported_comments
                            else:
                                row["_comments"] = existing_comments
                            row["_submitted_by"] = st.session_state.display_name
                            entries[idx] = row
                            overwritten += 1
                        else:
                            skipped += 1
                    save_entries(entries)
                    msg = f"Import complete: **{added}** added"
                    if overwritten: msg += f", **{overwritten}** overwritten"
                    if skipped:     msg += f", **{skipped}** skipped"
                    st.success(msg)
                    st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# TAB: EXPORT TO EXCEL
# ══════════════════════════════════════════════════════════════════════════════

if "export" in tab_index:
    with tab_index["export"]:
        entries = load_entries()
        st.subheader("Export P6-Ready Excel File")
        st.write("The exported file is formatted for direct import into Primavera P6:")
        st.write("- **Row 1:** P6 internal field key names (`task_code`, `act_start_date`, etc.)")
        st.write("- **Row 2:** Column headers · **Sheet name:** `TASK`")
        st.write("- **Date format:** `DD/MM/YYYY HH:MM` stored as proper Excel datetime cells")
        st.write("- **% Complete:** Plain integer")
        st.write("- **Complete Type:** Always Physical as to not overide remaining durations with calulated values from percent complete")
        st.divider()

        if not entries:
            st.warning("No entries to export yet.")
        else:
            st.info(f"{len(entries)} {'entry' if len(entries) == 1 else 'entries'} ready to export.")
            excel_bytes = build_excel(entries)
            fname = f"p6_asbuilt_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            st.download_button(
                label="⬇️  Download P6-Ready XLSX",
                data=excel_bytes,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
            st.divider()
            st.write("**How to import into P6:**")
            st.write("1. Edit WBS codes using CNTR-G replacing the old project name with the current one")
            st.write("2. Open **Primavera P6 Professional**")
            st.write("3. `File` → `Import` → `Spreadsheet (XLSX)` → **Next**")
            st.write("4. Browse to the downloaded file → **Next**")
            st.write("5. Select **Activities** → choose your project → **Finish**")
            st.write("6. Press **F9** to reschedule")

# ══════════════════════════════════════════════════════════════════════════════
# TAB: PHOTO LOG
# ══════════════════════════════════════════════════════════════════════════════

if "photos" in tab_index:
    with tab_index["photos"]:
        ensure_photo_dir()
        entries           = load_entries()
        known_ids_ordered = [e["activity_id"] for e in entries]
        known_map         = {e["activity_id"].upper(): e for e in entries}
        can_upload        = has_permission("photos")

        # ── Upload section (readwrite / admin only) ────────────────────────
        if can_upload:
            st.subheader("Upload Progress Photo")

            up_col1, up_col2 = st.columns(2)
            with up_col1:
                if known_ids_ordered:
                    selected_id = st.selectbox(
                        "Activity *",
                        options=known_ids_ordered,
                        format_func=lambda aid: f"{aid}  —  {known_map[aid.upper()]['activity_name']}",
                        key="photo_upload_activity",
                    )
                else:
                    st.warning("No activities yet — submit some entries first.")
                    selected_id = None
            with up_col2:
                photo_date = st.date_input(
                    "Date of Photo *",
                    value=date.today(),
                    format="DD/MM/YYYY",
                    key="photo_upload_date",
                )

            comment = st.text_input(
                "Comment",
                placeholder="e.g. North elevation — formwork complete (max 100 characters)",
                max_chars=100,
                key="photo_upload_comment",
            )

            uploaded_file = st.file_uploader(
                "Choose image *",
                type=["jpg", "jpeg", "png", "webp", "gif"],
                key="photo_upload_file",
            )

            if uploaded_file and selected_id:
                st.image(uploaded_file, caption="Preview", width=400)

            if st.button("📸  Save Photo", type="primary",
                         disabled=(not uploaded_file or not selected_id)):
                file_bytes    = uploaded_file.read()
                activity_name = known_map[selected_id.upper()]["activity_name"]
                add_photo(
                    activity_id   = selected_id,
                    activity_name = activity_name,
                    photo_date    = photo_date,
                    comment       = comment,
                    file_bytes    = file_bytes,
                    original_name = uploaded_file.name,
                    uploaded_by   = st.session_state.display_name,
                )
                st.success(f"✅ Photo saved for **{selected_id}** — {activity_name}")
                st.rerun()

            st.divider()

        # ── Gallery section (all roles) ────────────────────────────────────
        st.subheader("Photo Gallery")
        photos = load_photos()

        if not photos:
            st.info("No photos uploaded yet.")
        else:
            filter_col1, filter_col2 = st.columns([2, 3])
            with filter_col1:
                all_ids   = sorted({p["activity_id"] for p in photos})
                filter_id = st.selectbox(
                    "Filter by Activity",
                    options=["— All —"] + all_ids,
                    key="photo_filter_id",
                )
            with filter_col2:
                filter_text = st.text_input(
                    "Search comments",
                    placeholder="Type to search…",
                    key="photo_filter_text",
                ).strip().lower()

            filtered = photos
            if filter_id != "— All —":
                filtered = [p for p in filtered if p["activity_id"] == filter_id]
            if filter_text:
                filtered = [p for p in filtered
                            if filter_text in p.get("comment", "").lower()
                            or filter_text in p["activity_id"].lower()
                            or filter_text in p.get("activity_name", "").lower()]

            filtered = sorted(filtered, key=lambda p: p["photo_date"], reverse=True)

            st.caption(f"Showing {len(filtered)} of {len(photos)} photo{'s' if len(photos) != 1 else ''}")
            st.divider()

            if not filtered:
                st.info("No photos match the current filter.")
            else:
                COLS = 3
                for row_start in range(0, len(filtered), COLS):
                    cols = st.columns(COLS)
                    for col_idx, photo in enumerate(filtered[row_start:row_start + COLS]):
                        with cols[col_idx]:
                            img_path = PHOTO_DIR / photo["filename"]
                            if img_path.exists():
                                st.image(str(img_path), use_container_width=True)
                            else:
                                st.warning("Image file missing")

                            # Format photo date for display
                            try:
                                photo_dt_str = datetime.strptime(
                                    photo["photo_date"], "%Y-%m-%d"
                                ).strftime("%d/%m/%Y")
                            except ValueError:
                                photo_dt_str = photo["photo_date"]

                            activity_status = known_map.get(
                                photo["activity_id"].upper(), {}
                            ).get("activity_status", "")

                            st.write(f"**`{photo['activity_id']}`**  {photo.get('activity_name', '')}")
                            st.write(f"Status: {activity_status}")
                            st.write(f"📅 {photo_dt_str}")
                            if photo.get("comment"):
                                st.caption(photo["comment"])
                            st.caption(
                                f"Uploaded {photo['uploaded_at']} by {photo.get('uploaded_by', '—')}"
                            )

                            if can_upload:
                                if st.button("🗑 Delete", key=f"photo_del_{photo['id']}"):
                                    delete_photo(photo["id"])
                                    st.rerun()

                            st.divider()
