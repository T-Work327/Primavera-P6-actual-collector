"""
Primavera P6 Asbuilt Data Collector — Streamlit Web App
=========================================================
Run with:
    streamlit run p6_asbuilt_app.py

Requirements:
    pip install streamlit openpyxl

──────────────────────────────────────────────────────────
USER MANAGEMENT
──────────────────────────────────────────────────────────
Edit the USERS dictionary below to add, remove, or change
passwords and roles. Passwords are stored as SHA-256 hashes.

To generate a hash for a new password, run in Python:
    import hashlib
    print(hashlib.sha256("yourpassword".encode()).hexdigest())

Roles:
  viewer    — View entries only
  engineer  — View + Submit + Import + Photos + Site Walk
  admin     — All engineer permissions + Export + Settings
  developer — All admin permissions + Manage Users
──────────────────────────────────────────────────────────
"""

import io
import json
import uuid
import zipfile
import bcrypt
from datetime import date, datetime, time, timedelta
from pathlib import Path

import streamlit as st

try:
    from PIL import Image, ImageOps
    _PILLOW = True
except ImportError:
    _PILLOW = False

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    st.error("openpyxl is required.  Run:  pip install openpyxl")
    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
# USER DEFINITIONS
# ══════════════════════════════════════════════════════════════════════════════
#
# Roles:
#   viewer    — Read only
#   engineer  — View + Submit + Import + Photos + Site Walk
#   admin     — All engineer permissions + Export + Settings + Notifications
#   developer — All admin permissions + Manage Users (project assignment)
#
# Passwords are stored as bcrypt hashes in .streamlit/secrets.toml:
#   [passwords]
#   admin_hash      = "$2b$08$..."
#   admin2_hash     = "$2b$08$..."
#   engineer_hash   = "$2b$08$..."
#   engineer2_hash  = "$2b$08$..."
#   viewer_hash     = "$2b$08$..."
#   developer_hash  = "$2b$08$..."
#
# To generate a hash:
#   import bcrypt
#   print(bcrypt.hashpw("yourpassword".encode(), bcrypt.gensalt(rounds=8)).decode())

def _h(pw: str, rounds: int = 8) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt(rounds=rounds)).decode()

def _hcheck(pw: str, hsh: str) -> bool:
    return bcrypt.checkpw(pw.encode(), hsh.encode())

# Load hashes from secrets
_pw = st.secrets["passwords"]

USERS = {
    #  username        hash                          role          display name
    "admin":      {"hash": _pw["admin_hash"],      "role": "admin",     "name": "Administrator"},
    "admin2":     {"hash": _pw["admin2_hash"],     "role": "admin",     "name": "Administrator 2"},
    "engineer":   {"hash": _pw["engineer_hash"],   "role": "engineer",  "name": "Site Engineer"},
    "engineer2":  {"hash": _pw["engineer2_hash"],  "role": "engineer",  "name": "Site Engineer 2"},
    "viewer":     {"hash": _pw["viewer_hash"],     "role": "viewer",    "name": "Project Viewer"},
    "developer":  {"hash": _pw["developer_hash"],  "role": "developer", "name": "Developer"},
}

# ── Role Permission Matrix ─────────────────────────────────────────────────────
# Each role name maps directly to a set of permission keys.
# To change what a role can do, edit its set here — all users with that role
# are updated instantly without touching individual user records.
PERMISSIONS = {
    "viewer":    {"view"},
    "engineer":  {"view", "submit", "import", "photos", "sitewalk"},
    "admin":     {"view", "submit", "import", "export", "photos",
                  "settings", "sitewalk", "manage_users"},
    "developer": {"view", "submit", "import", "export", "photos",
                  "settings", "sitewalk", "manage_users",
                  "tab_order", "tab_visibility"},
}

def has_permission(perm: str) -> bool:
    return perm in PERMISSIONS.get(st.session_state.get("role", ""), set())

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

DATA_FILE  = Path("p6_asbuilt_store.json")
PHOTO_DIR  = Path("p6_images")
PHOTO_FILE   = Path("p6_photo_log.json")
ASSIGN_FILE   = Path("p6_photo_assignments.json")
PROJ_SETTINGS  = Path("p6_project_settings.json")
NOTIF_FILE     = Path("p6_notifications.json")
TAB_VIS_FILE   = Path("p6_tab_visibility.json")
LOOKAHEADDAYS= 14

USER_DATA = "DurationQtyType=QT_Day\nShowAsPercentage=0\nSmallScaleQtyType=QT_Hour\nDateFormat=dd/mm/yyyy\nCurrencyFormat=US Dollar"

STATUS_OPTIONS = ["Not Started", "In Progress", "Completed"]

STATUS_COLOUR = {
    "Not Started": "#6b7280",
    "In Progress":  "#d97706",
    "Completed":    "#16a34a",
}

ROLE_LABEL = {
    "viewer":    "Viewer",
    "engineer":  "Engineer",
    "admin":     "Admin",
    "developer": "Developer",
}

# P6 internal field key names (row 1 of TASK sheet)
P6_FIELD_KEYS = [
    "task_code", "task_name", "status_code", "act_start_date",
    "act_end_date", "complete_pct", "remain_drtn_hr_cnt",
    "complete_pct_type", "wbs_id", "user_field_813", "start_date", "task_type",
]

# Column definitions: (display header, column width, data key)
P6_COLUMNS = [
    ("Activity ID",          14, "activity_id"),
    ("Activity Name",        36, "activity_name"),
    ("Activity Status",      16, "activity_status"),
    ("Actual Start",         20, "actual_start"),
    ("Actual Finish",        20, "actual_finish"),
    ("Duration % Complete",  20, "pct_complete"),
    ("Remaining Duration",   20, "remaining_dur"),
    ("Percent Complete Type",20, "complete_pct_type"),
    ("WBS Code",             36, "wbs_id"),
    ("Comments",             50, "comments_export"),
    ("Predicted Start",      20, "predicted_start"),
    ("Task Type",            20, "task_type"),
]

DATE_KEYS = {"actual_start", "actual_finish", "predicted_start"}

# ══════════════════════════════════════════════════════════════════════════════
# DATE HELPERS
# Dates stored in JSON as ISO strings: "YYYY-MM-DDTHH:MM:00"
# ══════════════════════════════════════════════════════════════════════════════

def dt_to_iso(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%dT%H:%M:00")

def iso_to_dt(value: str) -> datetime | None:
    if not value or str(value).strip() == "":
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:00", "%Y-%m-%dT%H:%M",
                "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d",
                "%d-%b-%y %H:%M", "%d-%b-%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt)
        except ValueError:
            continue
    return None

def display_dt(value: str) -> str:
    dt = iso_to_dt(value)
    return dt.strftime("%d/%m/%Y %H:%M") if dt else "—"

def normalise_imported_date(raw_val) -> str:
    if raw_val is None or str(raw_val).strip() == "":
        return ""
    if isinstance(raw_val, datetime):
        return dt_to_iso(raw_val)
    dt = iso_to_dt(str(raw_val))
    return dt_to_iso(dt) if dt else str(raw_val).strip()


# ══════════════════════════════════════════════════════════════════════════════
# COMMENT HELPERS
# Comments are stored on each entry as a list of dicts:
#   _comments: [{"text": "...", "by": "...", "at": "DD/MM/YYYY HH:MM"}, ...]
# Newest first.  Exported to P6 as a single ';'-separated string (no timestamps).
# ══════════════════════════════════════════════════════════════════════════════

def comments_to_export(comments: list[dict]) -> str:
    """Flatten comment list → '; '-joined string, newest first, no timestamps."""
    return "; ".join(c["text"] for c in comments if c.get("text", "").strip())

def import_string_to_comments(raw: str, imported_by: str) -> list[dict]:
    """
    Split a '; '-separated import string into individual comment records.
    Each segment gets the same import timestamp and is marked as imported.
    Order is preserved (P6 exports newest first, so we keep that).
    """
    if not raw or not raw.strip():
        return []
    segments = [s.strip() for s in raw.split(";") if s.strip()]
    ts = datetime.now().strftime("%d/%m/%Y %H:%M")
    return [{"text": seg, "by": f"{imported_by} (imported)", "at": ts}
            for seg in segments]

def merge_imported_comments(imported: list[dict],
                             existing: list[dict]) -> list[dict]:
    """Return only imported comments whose text is not already in existing."""
    existing_texts = {c.get("text","").strip().lower() for c in existing}
    return [c for c in imported
            if c.get("text","").strip().lower() not in existing_texts]


# ══════════════════════════════════════════════════════════════════════════════
# STORAGE
# ══════════════════════════════════════════════════════════════════════════════

def load_entries() -> list[dict]:
    if DATA_FILE.exists():
        try:
            return json.loads(DATA_FILE.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return []
    return []

def save_entries(entries: list[dict]) -> None:
    DATA_FILE.write_text(json.dumps(entries, ensure_ascii=False, indent=2), encoding="utf-8")

def upsert_entry(entries: list[dict], new: dict) -> tuple:
    """Scoped to project — same activity_id allowed across different projects."""
    new_aid     = new.get("activity_id", "").upper()
    new_project = get_project_from_wbs(new.get("wbs_id", ""))
    idx = next(
        (i for i, e in enumerate(entries)
         if e.get("activity_id", "").upper() == new_aid
         and get_project_from_wbs(e.get("wbs_id", "")) == new_project),
        None,
    )
    if idx is not None:
        entries[idx] = new
        return entries, "updated"
    entries.append(new)
    return entries, "saved"

# ══════════════════════════════════════════════════════════════════════════════
# DUPLICATE DETECTION
# ══════════════════════════════════════════════════════════════════════════════

# Fields compared when deciding whether incoming data is identical to stored.
# Identity fields (activity_id, name, wbs) and metadata (_submitted_at etc.)
# are intentionally excluded — we only care about progress data changing.
_PROGRESS_FIELDS = (
    "activity_status", "actual_start", "actual_finish",
    "pct_complete", "remaining_dur", "predicted_start", "task_type",
)

def is_exact_duplicate(incoming: dict, stored: dict) -> bool:
    """Return True if all progress fields are identical between incoming and stored."""
    for field in _PROGRESS_FIELDS:
        # Normalise: strip whitespace, treat None and "" as equivalent
        a = str(incoming.get(field) or "").strip()
        b = str(stored.get(field)   or "").strip()
        if a != b:
            return False
    return True

@st.cache_data(show_spinner=False)
def build_photo_backup() -> bytes:
    """
    Build a ZIP of all image files plus the two JSON metadata files.
    Returns bytes ready for st.download_button.
    Cache cleared on upload/delete via build_photo_backup.clear().
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        if PHOTO_FILE.exists():
            zf.write(PHOTO_FILE, PHOTO_FILE.name)
        if ASSIGN_FILE.exists():
            zf.write(ASSIGN_FILE, ASSIGN_FILE.name)
        if PHOTO_DIR.exists():
            for img_path in PHOTO_DIR.iterdir():
                if img_path.is_file():
                    zf.write(img_path, f"{PHOTO_DIR.name}/{img_path.name}")
    return buf.getvalue()


def restore_photo_backup(zip_bytes: bytes) -> tuple[int, int, list[str]]:
    """
    Restore a photo library from a backup ZIP.
    Extracts image files into PHOTO_DIR, overwrites the two JSON metadata files.
    Returns (photos_restored, images_restored, warnings).
    Existing files not in the backup are left untouched.
    """
    ensure_photo_dir()
    warnings_list = []
    photos_restored = images_restored = 0

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = zf.namelist()

        # Restore metadata JSON files
        for json_name, dest_path in [
            (PHOTO_FILE.name,  PHOTO_FILE),
            (ASSIGN_FILE.name, ASSIGN_FILE),
        ]:
            if json_name in names:
                dest_path.write_bytes(zf.read(json_name))
                if json_name == PHOTO_FILE.name:
                    try:
                        photos_restored = len(json.loads(zf.read(json_name)))
                    except Exception:
                        pass
            else:
                warnings_list.append(f"{json_name} not found in backup.")

        # Restore image files
        img_prefix = f"{PHOTO_DIR.name}/"
        for name in names:
            if name.startswith(img_prefix) and not name.endswith("/"):
                img_filename = name[len(img_prefix):]
                dest = PHOTO_DIR / img_filename
                dest.write_bytes(zf.read(name))
                images_restored += 1

    return photos_restored, images_restored, warnings_list


# ══════════════════════════════════════════════════════════════════════════════
# PHOTO STORAGE
#
# Two-file model:
#   p6_photo_log.json         — one record per image file (no activity link)
#   p6_photo_assignments.json — many-to-many: {photo_id, activity_id}
#
# This means one image file is stored once and can be assigned to any number
# of activities without duplication.
# ══════════════════════════════════════════════════════════════════════════════

def ensure_photo_dir() -> None:
    PHOTO_DIR.mkdir(exist_ok=True)

# ── Photos (image records) ─────────────────────────────────────────────────

def load_photos() -> list[dict]:
    if PHOTO_FILE.exists():
        try:
            return json.loads(PHOTO_FILE.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return []
    return []

def save_photos(photos: list[dict]) -> None:
    PHOTO_FILE.write_text(json.dumps(photos, ensure_ascii=False, indent=2), encoding="utf-8")

THUMB_SIZE = (400, 400)   # max thumbnail dimensions

def upload_photo(photo_date: date, comment: str,
                 file_bytes: bytes, original_name: str,
                 uploaded_by: str) -> dict:
    """Save image + thumbnail and create a photo record. Does NOT assign to any activity."""
    ensure_photo_dir()
    base_id  = uuid.uuid4().hex
    ext      = Path(original_name).suffix.lower() or ".jpg"
    filename = f"{base_id}{ext}"
    thumb    = f"{base_id}_thumb.jpg"
    dest     = PHOTO_DIR / filename
    dest_t   = PHOTO_DIR / thumb

    if _PILLOW and ext in (".jpg", ".jpeg", ".png", ".webp"):
        img = Image.open(io.BytesIO(file_bytes))
        img = ImageOps.exif_transpose(img)
        img.save(dest)
        # Generate thumbnail — convert to RGB so PNG/WEBP save as JPEG cleanly
        t = img.copy()
        t.thumbnail(THUMB_SIZE, Image.LANCZOS)
        t.convert("RGB").save(dest_t, "JPEG", quality=75, optimize=True)
    else:
        dest.write_bytes(file_bytes)
        thumb = ""   # no thumbnail for GIF/unsupported

    record = {
        "id":          base_id,
        "photo_date":  photo_date.isoformat(),
        "comment":     comment.strip(),
        "filename":    filename,
        "thumb":       thumb,
        "uploaded_at": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "uploaded_by": uploaded_by,
    }
    photos = load_photos()
    photos.append(record)
    save_photos(photos)
    return record

def delete_photo_file(photo_id: str) -> None:
    """Delete the image file and all its assignments."""
    photos = load_photos()
    record = next((p for p in photos if p["id"] == photo_id), None)
    if record:
        img_path = PHOTO_DIR / record["filename"]
        if img_path.exists():
            img_path.unlink()
        thumb = record.get("thumb", "")
        if thumb:
            t_path = PHOTO_DIR / thumb
            if t_path.exists():
                t_path.unlink()
        save_photos([p for p in photos if p["id"] != photo_id])
        # Remove all assignments for this photo
        assignments = load_assignments()
        save_assignments([a for a in assignments if a["photo_id"] != photo_id])

# ── Assignments (many-to-many link) ───────────────────────────────────────

def load_assignments() -> list[dict]:
    if ASSIGN_FILE.exists():
        try:
            return json.loads(ASSIGN_FILE.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return []
    return []

def save_assignments(assignments: list[dict]) -> None:
    ASSIGN_FILE.write_text(
        json.dumps(assignments, ensure_ascii=False, indent=2), encoding="utf-8"
    )

@st.cache_data(show_spinner=False)
def load_image_bytes(filename: str) -> bytes | None:
    """Load image file bytes once and cache. Cache cleared on upload/delete."""
    if not filename:
        return None
    path = PHOTO_DIR / filename
    return path.read_bytes() if path.exists() else None

def assign_photo(photo_id: str, activity_ids: list[str], assigned_by: str,
                 entries: list | None = None) -> None:
    """Add assignments for a photo to a list of activities (skip duplicates).
    Stores wbs_id so photos resolve correctly when same activity_id exists
    in multiple projects.
    """
    if entries is None:
        entries = load_entries()
    wbs_lookup = {e.get("activity_id","").upper(): e.get("wbs_id","") for e in entries}
    assignments = load_assignments()
    existing    = {
        (a["photo_id"], a["activity_id"].upper(),
         get_project_from_wbs(a.get("wbs_id","")))
        for a in assignments
    }
    new_records = []
    for aid in activity_ids:
        wbs     = wbs_lookup.get(aid.upper(), "")
        project = get_project_from_wbs(wbs)
        if (photo_id, aid.upper(), project) not in existing:
            new_records.append({
                "photo_id":    photo_id,
                "activity_id": aid,
                "wbs_id":      wbs,
                "assigned_at": datetime.now().strftime("%d/%m/%Y %H:%M"),
                "assigned_by": assigned_by,
            })
    if new_records:
        assignments.extend(new_records)
        save_assignments(assignments)
        if "photo_assignments" in st.session_state:
            st.session_state["photo_assignments"] = assignments

def unassign_photo(photo_id: str, activity_id: str, wbs_id: str = "") -> None:
    """Remove a single photo→activity assignment scoped to project.
    If wbs_id is provided, only removes the assignment matching that project.
    """
    target_project = get_project_from_wbs(wbs_id) if wbs_id else None
    assignments = load_assignments()
    updated = []
    for a in assignments:
        if a["photo_id"] == photo_id and a["activity_id"].upper() == activity_id.upper():
            if target_project is None:
                continue
            if get_project_from_wbs(a.get("wbs_id","")) == target_project:
                continue
        updated.append(a)
    save_assignments(updated)
    if "photo_assignments" in st.session_state:
        st.session_state["photo_assignments"] = updated

def photos_for_activity(activity_id: str) -> list[dict]:
    """Return all photo records assigned to a given activity."""
    assignments = load_assignments()
    photo_map   = {p["id"]: p for p in load_photos()}
    return [
        photo_map[a["photo_id"]]
        for a in assignments
        if a["activity_id"].upper() == activity_id.upper()
        and a["photo_id"] in photo_map
    ]

def activities_for_photo(photo_id: str) -> list[str]:
    """Return list of activity_ids assigned to a given photo."""
    return [a["activity_id"] for a in load_assignments()
            if a["photo_id"] == photo_id]

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════

# P6 internal status codes — must NOT contain spaces (0x20) or P6 rejects the import.
STATUS_TO_P6 = {
    "Not Started": "TK_NotStart",
    "In Progress":  "TK_Active",
    "Completed":    "TK_Complete",
}

THIN   = Side(style="thin", color="B0B8C8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

@st.cache_data
def build_excel(entries: list[dict], project_name: str = "") -> bytes:
    """Build P6-ready XLSX.
    If project_name is supplied, the WBS prefix on every row is replaced with it.
    e.g. project_name="ProjectB" turns "ProjectA.1.2.3" → "ProjectB.1.2.3"
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TASK"
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

    # Row 1 — P6 internal field keys
    for col_idx, (key, (_, width, _)) in enumerate(zip(P6_FIELD_KEYS, P6_COLUMNS), start=1):
        c = ws.cell(row=1, column=col_idx, value=key)
        c.font      = Font(name="Arial", italic=True, color="4472C4", size=9)
        c.fill      = PatternFill("solid", fgColor="D9E1F2")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Row 2 — Human-readable column headers
    for col_idx, (header, _, _) in enumerate(P6_COLUMNS, start=1):
        c = ws.cell(row=2, column=col_idx, value=header)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", fgColor="1C3557")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = BORDER

    # Data rows
    for row_idx, entry in enumerate(entries, start=3):
        fill = PatternFill("solid", fgColor="EEF2F8" if row_idx % 2 == 0 else "FFFFFF")
        for col_idx, (_, _, key) in enumerate(P6_COLUMNS, start=1):
            value = entry.get(key, "")
            if key in DATE_KEYS:
                dt_val = iso_to_dt(value)
                c = ws.cell(row=row_idx, column=col_idx, value=dt_val)
                if dt_val:
                    c.number_format = "DD/MM/YYYY HH:MM"
            elif key == "complete_pct_type":
                c = ws.cell(row=row_idx, column=col_idx, value="Physical")
            elif key == "comments_export":
                # Build export string from stored comment list (newest first, no timestamps)
                c = ws.cell(row=row_idx, column=col_idx,
                            value=comments_to_export(entry.get("_comments", [])))
            elif key == "wbs_id" and project_name and value:
                # Replace existing prefix with the supplied project name
                new_wbs = project_name.strip() + "." + strip_wbs_prefix(str(value))
                c = ws.cell(row=row_idx, column=col_idx, value=new_wbs)
            else:
                c = ws.cell(row=row_idx, column=col_idx, value=str(value) if value != "" else "")
            c.fill   = fill
            c.border = BORDER
            c.alignment = Alignment(vertical="center", wrap_text=False)
            if col_idx <= 2:
                c.fill = PatternFill("solid", fgColor="F2F2F2")
                c.font = Font(name="Arial", size=10, bold=(col_idx == 1))
            else:
                c.font = Font(name="Arial", size=10, color="1F4E79")
            if key in ("pct_complete", "remaining_dur") and value != "":
                c.alignment = Alignment(horizontal="right", vertical="center")

    # USERDATA sheet — do not modify this section, P6 is very particular about it
    wu  = wb.create_sheet("USERDATA")
    wu.column_dimensions["A"].width = 60
    b2  = Border(left=Side(style="thin", color="B0B8C8"), right=Side(style="thin", color="B0B8C8"),
                 top=Side(style="thin",  color="B0B8C8"), bottom=Side(style="thin", color="B0B8C8"))
    # Row 1: field key identifier (no spaces — required by P6)
    r1 = wu.cell(row=1, column=1, value="user_data")
    r1.font   = Font(name="Arial", bold=True, size=9, color="4472C4")
    r1.fill   = PatternFill("solid", fgColor="D9E1F2")
    r1.border = b2
    # Row 2: section label
    r2 = wu.cell(row=2, column=1, value="UserSettings Do Not Edit")
    r2.font   = Font(name="Arial", bold=True, size=11, color="1C3557")
    r2.fill   = PatternFill("solid", fgColor="D9E1F2")
    r2.border = b2
    # Row 3: settings values
    r3 = wu.cell(row=3, column=1,
                 value=USER_DATA)
    r3.font      = Font(name="Arial", size=10)
    r3.fill      = PatternFill("solid", fgColor="F8F9FB")
    r3.alignment = Alignment(vertical="top", wrap_text=True)
    r3.border    = b2
    wu.row_dimensions[3].height = 80

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL IMPORT
# ══════════════════════════════════════════════════════════════════════════════

P6_KEY_MAP = {
    "task_code": "activity_id", "task_name": "activity_name",
    "status_code": "activity_status", "act_start_date": "actual_start",
    "act_end_date": "actual_finish", "complete_pct": "pct_complete",
    "remain_drtn_hr_cnt": "remaining_dur", "complete_pct_type": "complete_pct_type",
    "wbs_id": "wbs_id", "user_field_813": "comments_import",
    "start_date": "predicted_start",
    "task_type": "task_type",
}
HEADER_KEY_MAP = {
    "activity id": "activity_id", "activity name": "activity_name",
    "activity status": "activity_status", "actual start": "actual_start",
    "actual finish": "actual_finish", "duration % complete": "pct_complete",
    "remaining duration": "remaining_dur", "percent complete type": "complete_pct_type",
    "wbs code": "wbs_id", "comments": "comments_import",
    "predicted start": "predicted_start",
    "task type": "task_type",
}

def read_p6_excel(file_bytes: bytes) -> tuple:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    warnings_list = []
    sheet_name = next((s for s in wb.sheetnames if s.upper() == "TASK"), None)
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
        warnings_list.append(f"No TASK sheet found — reading from '{sheet_name}' instead.")
    ws = wb[sheet_name]
    rows_iter = list(ws.iter_rows(values_only=True))
    if not rows_iter:
        return [], ["The sheet appears to be empty."]
    col_map, data_start = {}, 1
    for row_idx in range(min(3, len(rows_iter))):
        row = rows_iter[row_idx]
        mapping = {}
        for col_idx, cell_val in enumerate(row):
            if cell_val is None:
                continue
            cs = str(cell_val).strip().lower()
            if cs in P6_KEY_MAP:
                dk = P6_KEY_MAP[cs]
                if dk != "complete_pct_type":  # always force Physical on export; ignore on import
                    mapping[col_idx] = dk
            elif cs in HEADER_KEY_MAP:
                dk = HEADER_KEY_MAP[cs]
                if dk != "complete_pct_type":
                    mapping[col_idx] = dk
        if mapping:
            col_map, data_start = mapping, row_idx + 1
            if row_idx == 0 and len(rows_iter) > 1:
                next_str = [str(v).strip().lower() for v in rows_iter[1] if v is not None]
                if any(h in HEADER_KEY_MAP for h in next_str):
                    data_start = 2
            break
    if not col_map:
        return [], ["Could not detect column headers."]
    entries = []
    for row in rows_iter[data_start:]:
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        entry = {
            "activity_id": "", "activity_name": "", "activity_status": "",
            "actual_start": "", "actual_finish": "", "pct_complete": "",
            "remaining_dur": "", "complete_pct_type": "Physical", "wbs_id": "",
            "comments_import": "", "predicted_start": "", "task_type": "",
        }
        for col_idx, data_key in col_map.items():
            if col_idx >= len(row):
                continue
            raw_val = row[col_idx]
            if data_key in DATE_KEYS:
                entry[data_key] = normalise_imported_date(raw_val)
            elif data_key == "pct_complete":
                vs = str(raw_val).replace("%", "").strip() if raw_val is not None else ""
                try:
                    entry[data_key] = str(int(float(vs))) if vs else ""
                except ValueError:
                    entry[data_key] = vs
            elif data_key == "remaining_dur":
                if raw_val is None or str(raw_val).strip() == "":
                    entry[data_key] = ""
                else:
                    try:
                        # Store as plain integer — no trailing .0
                        entry[data_key] = str(int(float(str(raw_val).strip())))
                    except ValueError:
                        entry[data_key] = str(raw_val).strip()
            elif data_key == "comments_import":
                # Store as plain string; semicolons preserved for import_string_to_comments
                entry[data_key] = "" if raw_val is None else str(raw_val).strip()
            else:
                entry[data_key] = "" if raw_val is None else str(raw_val).strip()
        if not entry["activity_id"]:
            continue
        if not entry["complete_pct_type"]:
            entry["complete_pct_type"] = "Physical"
        entry["_submitted_at"] = datetime.now().strftime("%d/%m/%Y %H:%M")
        entries.append(entry)
    return entries, warnings_list


# ══════════════════════════════════════════════════════════════════════════════
# MICROSOFT PROJECT IMPORT
# ══════════════════════════════════════════════════════════════════════════════
# MS Project exports don't carry stable Activity IDs, so rows are matched to
# stored activities by (Name + WBS suffix).
# P6 prefixes WBS with the project name:  "ProjectX.1.2.3"
# MS Project stores WBS without prefix:   "1.2.3"
# We strip everything up to and including the first '.' before comparing.

MSP_KEY_MAP = {
    # MS Project XML/XLSX field headers (lowercase)
    "task name":          "activity_name",
    "name":               "activity_name",
    "wbs":                "wbs_id",
    "outline number":     "wbs_id",
    "% complete":         "pct_complete",
    "percent complete":   "pct_complete",
    "% work complete":    "pct_complete",
    "actual start":       "actual_start",
    "actual finish":      "actual_finish",
    "actual duration":    "remaining_dur",
    "remaining duration": "remaining_dur",
    "duration":           "remaining_dur",
    "status":             "activity_status",
    "notes":              "comments_import",
}

# MS Project status strings → our internal values
MSP_STATUS_MAP = {
    "complete":     "Completed",
    "completed":    "Completed",
    "in progress":  "In Progress",
    "not started":  "Not Started",
    "future task":  "Not Started",
    "on schedule":  "In Progress",
    "late":         "In Progress",
    "":             "Not Started",
}

def strip_wbs_prefix(wbs) -> str:
    """Strip the P6 project-name prefix from a stored WBS code.
    'ProjectX.1.2.3'  →  '1.2.3'
    '1.2.3'           →  '1.2.3'  (already clean, first segment is numeric)
    """
    wbs = str(wbs or "")
    if not wbs:
        return ""
    parts = wbs.strip().split(".", 1)
    if len(parts) == 2 and not parts[0].isdigit():
        return parts[1]
    return wbs.strip()


def strip_msp_wbs(wbs: str) -> str:
    """Normalise an MSP WBS code for comparison against stored P6 WBS.
    MSP has one extra level vs P6, and single-segment codes are WBS titles.
    '1.2.3.4'  →  '1.2.3'   (drop last segment)
    '1'        →  ''          (WBS title row — ignored)
    '1.2'      →  '1'
    """
    if not wbs:
        return ""
    segments = wbs.strip().split(".")
    if len(segments) <= 1:
        return ""
    return ".".join(segments[:-1])

def get_project_from_wbs(wbs) -> str:
    """Extract the project name from a WBS code.
    'ProjectA.1.2.3' -> 'ProjectA'
    '1.2.3'          -> '(Unassigned)'
    """
    wbs = str(wbs or "")
    if not wbs:
        return "(Unassigned)"
    parts = wbs.strip().split(".", 1)
    if len(parts) == 2 and not parts[0].isdigit():
        return parts[0]
    return "(Unassigned)"


def get_all_projects(entries: list[dict]) -> list[str]:
    """Return sorted unique project names from all entries."""
    return sorted({get_project_from_wbs(e.get("wbs_id", "")) for e in entries})


def filter_by_project(entries: list[dict], project: str) -> list[dict]:
    """Filter entries to those belonging to the given project name."""
    if not project:
        return entries
    return [e for e in entries if get_project_from_wbs(e.get("wbs_id", "")) == project]


def make_wbs_with_project(project: str, numeric_wbs: str) -> str:
    """Prepend project name to a numeric WBS. Avoids double-prefixing."""
    if not project or project == "(Unassigned)":
        return numeric_wbs.strip()
    numeric = numeric_wbs.strip()
    if numeric.startswith(project + "."):
        return numeric
    return f"{project}.{numeric}" if numeric else project


def rename_project(entries: list[dict], old_name: str, new_name: str) -> tuple[list[dict], int]:
    """
    Rename a project across all stored entries by replacing the WBS prefix,
    and also updates wbs_id stored in photo assignment records to keep the
    photo log in sync.
    Returns (updated_entries, count_changed).
    """
    changed = 0
    for entry in entries:
        wbs = entry.get("wbs_id", "")
        if not wbs:
            continue
        if get_project_from_wbs(wbs) == old_name:
            numeric = strip_wbs_prefix(wbs)
            entry["wbs_id"] = f"{new_name}.{numeric}" if numeric else new_name
            changed += 1

    # Update wbs_id in assignment records so photo→activity links stay in sync
    assignments = load_assignments()
    asgn_changed = 0
    for a in assignments:
        if get_project_from_wbs(a.get("wbs_id","")) == old_name:
            numeric = strip_wbs_prefix(a.get("wbs_id",""))
            a["wbs_id"] = f"{new_name}.{numeric}" if numeric else new_name
            asgn_changed += 1
    if asgn_changed:
        save_assignments(assignments)
        for _k in ("photo_assignments","photo_to_aids","aid_to_pids","_assign_sig"):
            st.session_state.pop(_k, None)

    return entries, changed


# ══════════════════════════════════════════════════════════════════════════════
# PROJECT SETTINGS  (report date, future: holidays)
# Stored as {project_name: {report_date: "YYYY-MM-DD"}}
# ══════════════════════════════════════════════════════════════════════════════

def load_project_settings() -> dict:
    if PROJ_SETTINGS.exists():
        try:
            return json.loads(PROJ_SETTINGS.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return {}
    return {}

def save_project_settings(settings: dict) -> None:
    PROJ_SETTINGS.write_text(
        json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8"
    )

def get_report_date(project: str) -> date | None:
    """Return the stored report date for a project, or None."""
    settings = load_project_settings()
    raw = settings.get(project, {}).get("report_date")
    if not raw:
        return None
    try:
        return datetime.strptime(raw, "%Y-%m-%d").date()
    except ValueError:
        return None

def set_report_date(project: str, report_dt: date) -> None:
    settings = load_project_settings()
    settings.setdefault(project, {})["report_date"] = report_dt.isoformat()
    save_project_settings(settings)


def get_last_walk_date(project: str) -> date | None:
    """Return the stored last site walk date for a project, or None."""
    settings = load_project_settings()
    raw = settings.get(project, {}).get("last_walk_date")
    if not raw:
        return None
    try:
        return datetime.strptime(raw, "%Y-%m-%d").date()
    except ValueError:
        return None

def set_last_walk_date(project: str, walk_dt: date) -> None:
    settings = load_project_settings()
    settings.setdefault(project, {})["last_walk_date"] = walk_dt.isoformat()
    save_project_settings(settings)


# ══════════════════════════════════════════════════════════════════════════════
# NOTIFICATIONS
# ══════════════════════════════════════════════════════════════════════════════

def load_notifications() -> list[dict]:
    if NOTIF_FILE.exists():
        try:
            return json.loads(NOTIF_FILE.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return []
    return []

def save_notifications(notifs: list[dict]) -> None:
    NOTIF_FILE.write_text(
        json.dumps(notifs, ensure_ascii=False, indent=2), encoding="utf-8"
    )

def get_project_admin_recipients(project: str) -> list[str]:
    """Return usernames of admin/developer users who have access to this project."""
    recipients = []
    for username, user in USERS.items():
        role = user.get("role", "")
        if role in ("admin", "developer"):
            if user_can_access_project(username, project):
                recipients.append(username)
    return recipients


def create_notification(created_by: str, project: str, title: str, body: str,
                        recipients: list[str] | None = None,
                        rows: list[dict] | None = None,
                        status: str = "info",
                        edits: dict | None = None,
                        rejection_note: str = "") -> str:
    """
    Create a notification. Returns the new notification's id.

    status:
      "info"      — plain notification, no action needed (default)
      "pending"   — awaiting admin approval; edits holds the proposed changes
      "approved"  — approved and applied
      "rejected"  — rejected and sent back to the engineer with a note

    rows is an optional list of dicts shown as a table.
    edits is the activity_id -> field-changes dict awaiting approval.
    """
    if recipients is None:
        recipients = get_project_admin_recipients(project)
    notifs = load_notifications()
    notif_id = uuid.uuid4().hex
    notifs.append({
        "id":             notif_id,
        "created_at":     datetime.now().strftime("%d/%m/%Y %H:%M"),
        "created_by":     created_by,
        "project":        project,
        "title":          title,
        "body":           body,
        "rows":           rows or [],
        "recipients":     recipients,
        "read_by":        [],
        "status":         status,
        "edits":          edits or {},
        "rejection_note": rejection_note,
    })
    save_notifications(notifs)
    return notif_id


def update_notification(notif_id: str, **fields) -> None:
    """Update arbitrary fields on an existing notification record."""
    notifs = load_notifications()
    for n in notifs:
        if n["id"] == notif_id:
            n.update(fields)
    save_notifications(notifs)

# ══════════════════════════════════════════════════════════════════════════════
# TAB VISIBILITY
# ══════════════════════════════════════════════════════════════════════════════

def load_tab_visibility() -> dict:
    if TAB_VIS_FILE.exists():
        try:
            return json.loads(TAB_VIS_FILE.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return {}
    return {}

def save_tab_visibility(vis: dict) -> None:
    TAB_VIS_FILE.write_text(
        json.dumps(vis, ensure_ascii=False, indent=2), encoding="utf-8"
    )

def load_tab_order() -> list[str]:
    """Return stored tab order as list of perm keys, or [] for default."""
    return load_tab_visibility().get("tab_order", [])

def save_tab_order(order: list[str]) -> None:
    vis = load_tab_visibility()
    vis["tab_order"] = order
    save_tab_visibility(vis)


def is_tab_visible(perm: str, username: str, role: str) -> bool:
    """Return True if this tab should be shown for this user.
    Developer always sees all permitted tabs.
    Checks role default then user override.
    """
    if role == "developer":
        return True
    vis = load_tab_visibility()
    role_visible = vis.get("roles", {}).get(role, {}).get(perm, True)
    user_vis = vis.get("users", {}).get(username, {})
    if perm in user_vis:
        return user_vis[perm]
    return role_visible


def mark_notification_read(notif_id: str, username: str) -> None:
    notifs = load_notifications()
    for n in notifs:
        if n["id"] == notif_id and username not in n["read_by"]:
            n["read_by"].append(username)
    save_notifications(notifs)

def delete_notification(notif_id: str) -> None:
    save_notifications([n for n in load_notifications() if n["id"] != notif_id])

def unread_notifications(username: str) -> list[dict]:
    """Return notifications unread by this user and addressed to them."""
    return [
        n for n in load_notifications()
        if username not in n.get("read_by", [])
        and (not n.get("recipients") or username in n.get("recipients", []))
    ]


def get_allowed_users(project: str) -> list[str]:
    """Return list of usernames allowed on this project.
    Empty list means no restriction — all users are allowed.
    """
    settings = load_project_settings()
    return settings.get(project, {}).get("allowed_users", [])

def set_allowed_users(project: str, usernames: list[str]) -> None:
    settings = load_project_settings()
    settings.setdefault(project, {})["allowed_users"] = usernames
    save_project_settings(settings)

def auto_assign_new_projects(username: str,
                              before_projects: set[str],
                              after_entries: list[dict]) -> None:
    """
    After saving entries, detect any newly created projects (present in
    after_entries but not in before_projects) and automatically assign
    the creating user to them.
    Skips (Unassigned) and projects that already have explicit access rules.
    """
    after_projects = set(get_all_projects(after_entries))
    new_projects   = after_projects - before_projects - {"(Unassigned)"}
    for proj in new_projects:
        existing = get_allowed_users(proj)
        if not existing:   # no rules yet — set the creator as sole allowed user
            set_allowed_users(proj, [username])

def user_can_access_project(username: str, project: str) -> bool:
    """Return True if username is allowed to access project.
    If allowed_users is empty, all users are permitted (no restriction set).
    '(Unassigned)' is always accessible to everyone.
    """
    if project == "(Unassigned)":
        return True
    allowed = get_allowed_users(project)
    if not allowed:          # no restriction configured
        return True
    return username in allowed

def get_accessible_projects(username: str, all_projects: list[str]) -> list[str]:
    """Filter a list of project names to those the user can access."""
    return [p for p in all_projects if user_can_access_project(username, p)]


# ══════════════════════════════════════════════════════════════════════════════
# WORKING DAYS & DURATION CALCULATIONS
# Weekends excluded; public holidays can be added to HOLIDAYS set later.
# ══════════════════════════════════════════════════════════════════════════════

HOLIDAYS: set[date] = set()   # add dates here when needed, e.g. date(2026,1,1)

def working_days_between(start: date, end: date) -> int:
    """Count working days (Mon–Fri, excluding HOLIDAYS) from start up to but not
    including end.  Returns 0 if end <= start."""
    from datetime import timedelta as _td
    if end <= start:
        return 0
    total = 0
    cur   = start
    while cur < end:
        if cur.weekday() < 5 and cur not in HOLIDAYS:
            total += 1
        cur += _td(days=1)
    return total

def _add_working_days(start: date, days: int) -> date:
    """Return the date that is `days` working days after start."""
    if days <= 0:
        return start
    cur   = start
    added = 0
    while added < days:
        cur = date.fromordinal(cur.toordinal() + 1)
        if cur.weekday() < 5 and cur not in HOLIDAYS:
            added += 1
    return cur

def expected_finish_date(report_date: date, remaining_dur: str) -> date | None:
    """Calculate expected finish = report_date + remaining working days."""
    try:
        days = int(float(remaining_dur))
    except (ValueError, TypeError):
        return None
    if days < 0:
        return None
    return _add_working_days(report_date, days)

def calc_duration_pct(actual_start_iso: str, expected_finish: date,
                      report_date: date) -> int | None:
    """
    Duration % complete = working days elapsed / total working days * 100.
    Elapsed = actual_start → report_date
    Total   = actual_start → expected_finish
    Returns integer 0-99 (never 100 — only Completed activities are 100%).
    Returns None if data is insufficient.
    """
    start_dt = iso_to_dt(actual_start_iso)
    if start_dt is None:
        return None
    start_d = start_dt.date()
    if report_date <= start_d or expected_finish <= start_d:
        return None
    # Include the report date itself in elapsed (count up to day after)
    from datetime import timedelta as _td
    elapsed = working_days_between(start_d, report_date + _td(days=1))
    total   = working_days_between(start_d, expected_finish + _td(days=1))
    if total == 0:
        return None
    pct = min(99, int(elapsed / total * 100))
    return pct

def recalculate_project(entries: list[dict], project: str,
                        report_date: date) -> tuple[list[dict], int]:
    """
    Recalculate duration % complete for all In Progress activities in project.
    Only updates pct_complete — all other fields are unchanged.
    Returns (entries, count_updated).
    """
    updated = 0
    for entry in entries:
        if entry.get("activity_status") != "In Progress":
            continue
        if get_project_from_wbs(entry.get("wbs_id","")) != project:
            continue
        ef = expected_finish_date(report_date, entry.get("remaining_dur",""))
        if ef is None:
            continue
        pct = calc_duration_pct(entry.get("actual_start",""), ef, report_date)
        if pct is not None:
            entry["pct_complete"] = str(pct)
            updated += 1
    return entries, updated


def read_msp_excel(file_bytes: bytes) -> tuple:
    """
    Read a Microsoft Project XLSX export.
    Returns (rows, warnings) where rows use our internal field names.
    Rows do NOT have activity_id set — matching is done in the UI.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    warnings_list = []
    # Use the first sheet (MSP exports vary in sheet naming)
    ws = wb[wb.sheetnames[0]]
    rows_iter = list(ws.iter_rows(values_only=True))
    if not rows_iter:
        return [], ["The sheet appears to be empty."]

    # Detect header row
    col_map, data_start = {}, 1
    for row_idx in range(min(5, len(rows_iter))):
        row = rows_iter[row_idx]
        mapping = {}
        for col_idx, cell_val in enumerate(row):
            if cell_val is None:
                continue
            cs = str(cell_val).strip().lower()
            if cs in MSP_KEY_MAP:
                mapping[col_idx] = MSP_KEY_MAP[cs]
        if mapping:
            col_map, data_start = mapping, row_idx + 1
            break

    if not col_map:
        return [], ["Could not detect Microsoft Project column headers. "
                    "Expected columns like 'Task Name', 'WBS', '% Complete', "
                    "'Actual Start', 'Actual Finish'."]

    if "activity_name" not in col_map.values():
        warnings_list.append("No Task Name / Name column found.")
    if "wbs_id" not in col_map.values():
        warnings_list.append("No WBS / Outline Number column found — name-only matching will be used.")

    entries = []
    for row in rows_iter[data_start:]:
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        entry = {
            "activity_id":       "",   # filled by matching logic in the UI
            "activity_name":     "",
            "activity_status":   "",
            "actual_start":      "",
            "actual_finish":     "",
            "pct_complete":      "",
            "remaining_dur":     "",
            "complete_pct_type": "Physical",
            "wbs_id":            "",
            "comments_import":   "",
        }
        for col_idx, data_key in col_map.items():
            if col_idx >= len(row):
                continue
            raw_val = row[col_idx]
            if data_key in DATE_KEYS:
                entry[data_key] = normalise_imported_date(raw_val)
            elif data_key == "pct_complete":
                vs = str(raw_val).replace("%", "").strip() if raw_val is not None else ""
                try:
                    entry[data_key] = str(int(float(vs))) if vs else ""
                except ValueError:
                    entry[data_key] = vs
            elif data_key == "activity_status":
                raw_str = str(raw_val).strip().lower() if raw_val else ""
                entry[data_key] = MSP_STATUS_MAP.get(raw_str, "Not Started")
            elif data_key == "remaining_dur":
                # MSP duration strings like "5 days", "5d", "5" — extract number
                if raw_val is None:
                    entry[data_key] = ""
                else:
                    try:
                        # Store as plain integer — no trailing .0
                        entry[data_key] = str(int(float(str(raw_val).strip())))
                    except ValueError:
                        entry[data_key] = str(raw_val).strip()
            else:
                entry[data_key] = "" if raw_val is None else str(raw_val).strip()

        # Skip summary rows (WBS with no sub-level or name only rows)
        if not entry["activity_name"]:
            continue

        # Strip P6 prefix from WBS for consistent comparison
        entry["wbs_id"] = strip_msp_wbs(entry["wbs_id"])
        entry["_submitted_at"] = datetime.now().strftime("%d/%m/%Y %H:%M")
        entries.append(entry)

    return entries, warnings_list


def match_msp_to_stored(msp_rows: list[dict], stored: list[dict]) -> tuple:
    """
    Match MSP rows to stored activities by (name + WBS suffix).
    Returns:
      matched       — list of (msp_row, stored_entry)   — exactly one match
      unmatched     — list of msp_row                    — no stored match found
      duplicates    — list of (msp_row, [stored_entries]) — ambiguous (2+ matches)
    """
    # Build lookup: (name.lower(), wbs_suffix.lower()) → [stored entries]
    lookup: dict[tuple, list] = {}
    for e in stored:
        key = (
            e.get("activity_name", "").strip().lower(),
            strip_wbs_prefix(str(e.get("wbs_id", "") or "")).lower(),
        )
        lookup.setdefault(key, []).append(e)

    matched, unmatched, duplicates = [], [], []
    for row in msp_rows:
        key = (
            row.get("activity_name", "").strip().lower(),
            str(row.get("wbs_id", "") or "").lower(),
        )
        hits = lookup.get(key, [])
        if len(hits) == 1:
            matched.append((row, hits[0]))
        elif len(hits) == 0:
            unmatched.append(row)
        else:
            duplicates.append((row, hits))

    return matched, unmatched, duplicates

# ══════════════════════════════════════════════════════════════════════════════
# WBS OFFSET DETECTION
# ══════════════════════════════════════════════════════════════════════════════

def detect_wbs_offset(unmatched: list[dict], stored: list[dict]) -> list[dict]:
    """
    For each unmatched MSP row, check whether the same activity name exists in
    stored with a WBS that differs only in the last numeric segment by ±1.

    Returns a list of offset suggestions, each a dict:
      {
        "msp_row":      the unmatched MSP row,
        "stored_entry": the stored entry found at the adjacent WBS,
        "msp_wbs":      e.g. "1.2.3"
        "stored_wbs":   e.g. "1.2.2"   (stored WBS after stripping prefix)
        "depth":        the segment index where they differ (0-based),
        "delta":        +1 or -1  (MSP value minus stored value at that depth),
      }

    Only suggests when the name match is unique and unambiguous.
    """
    # Build name → list of stored entries
    name_lookup: dict[str, list] = {}
    for e in stored:
        name_lookup.setdefault(e.get("activity_name","").strip().lower(), []).append(e)

    suggestions = []
    for row in unmatched:
        msp_name = row.get("activity_name","").strip().lower()
        msp_wbs  = row.get("wbs_id","").strip()
        if not msp_wbs:
            continue
        msp_segs = msp_wbs.split(".")
        if len(msp_segs) < 2:
            continue

        candidates = name_lookup.get(msp_name, [])
        if not candidates:
            continue

        for delta in (+1, -1):
            for depth in range(len(msp_segs) - 1, -1, -1):
                try:
                    adj_val = int(msp_segs[depth]) - delta  # stored value = msp - delta
                    if adj_val < 0:
                        continue
                except ValueError:
                    continue

                adj_segs = msp_segs[:]
                adj_segs[depth] = str(adj_val)
                adj_wbs = ".".join(adj_segs)

                # Compare against stored WBS after stripping P6 prefix
                for e in candidates:
                    stored_wbs_clean = strip_wbs_prefix(str(e.get("wbs_id","") or "")).lower()
                    if stored_wbs_clean == adj_wbs.lower():
                        suggestions.append({
                            "msp_row":      row,
                            "stored_entry": e,
                            "msp_wbs":      msp_wbs,
                            "stored_wbs":   stored_wbs_clean,
                            "depth":        depth,
                            "delta":        delta,
                        })

    return suggestions


def apply_wbs_offset(stored: list[dict], prefix: str, depth: int,
                     delta: int, from_val: int) -> tuple[list[dict], int]:
    """
    Shift WBS codes in stored entries where:
      - The WBS prefix up to `depth` matches `prefix`
      - The segment at `depth` is >= from_val  (for positive delta)
        or <= from_val  (for negative delta)

    `prefix` is the dot-joined segments BEFORE depth, e.g. "1.2" for depth 2.
    `delta`  is the amount to add to segment at depth (+1 or -1).
    `from_val` is the stored segment value where the shift starts.

    Returns (updated_entries, count_changed).
    """
    changed = 0
    for entry in stored:
        raw_wbs   = entry.get("wbs_id", "")
        clean_wbs = strip_wbs_prefix(raw_wbs)
        segs      = clean_wbs.split(".")
        if len(segs) <= depth:
            continue
        # Check prefix matches
        if prefix and ".".join(segs[:depth]) != prefix:
            continue
        try:
            seg_val = int(segs[depth])
        except ValueError:
            continue
        # Only shift entries at or beyond the insertion/deletion point
        if delta > 0 and seg_val < from_val:
            continue
        if delta < 0 and seg_val > from_val:
            continue

        new_segs       = segs[:]
        new_segs[depth] = str(seg_val + delta)
        new_clean_wbs  = ".".join(new_segs)

        # Reconstruct: if original had a prefix, re-attach it
        original_segs = raw_wbs.split(".")
        if len(original_segs) > len(segs):
            # There was a non-numeric prefix segment
            prefix_part  = ".".join(original_segs[:len(original_segs)-len(segs)])
            entry["wbs_id"] = prefix_part + "." + new_clean_wbs
        else:
            entry["wbs_id"] = new_clean_wbs
        changed += 1

    return stored, changed


# ══════════════════════════════════════════════════════════════════════════════
# DATE INPUT WIDGET HELPER
# Uses st.datetime_input (Streamlit >= 1.43).
# Returns a datetime, or None if the user has not enabled an optional field.
# ══════════════════════════════════════════════════════════════════════════════

def datetime_inputs(label: str, key: str, required: bool = True,
                    default_dt: datetime | None = None) -> datetime | None:
    """
    Render a single datetime picker (st.datetime_input).
    For optional fields a checkbox gates the widget; returns None when unchecked.
    """
    default_val = default_dt if default_dt else datetime.combine(date.today(), time(8, 0))

    if not required:
        enabled = st.checkbox(f"Set {label}", key=f"{key}_enabled",
                              value=(default_dt is not None))
        if not enabled:
            return None

    return st.datetime_input(label, value=default_val, key=f"{key}_dt",
                              step=60 * 15)   # 15-minute steps

# ══════════════════════════════════════════════════════════════════════════════
# PAGE SETUP
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="P6 Asbuilt Collector", page_icon="🏗️", layout="wide")

# ── Mobile camera helper ──────────────────────────────────────────────────────
# Injects accept="image/*" capture="environment" onto Streamlit file uploaders
# so mobile/tablet devices open the camera app directly.
# Call inject_camera_uploader(key) right after the st.file_uploader widget.

def inject_camera_uploader(uploader_key: str) -> None:
    """Patch the hidden <input type=file> to open the camera on mobile."""
    st.markdown(
        f"""<script>
        (function() {{
            function patch() {{
                var inputs = window.parent.document.querySelectorAll('input[type=file]');
                inputs.forEach(function(el) {{
                    el.setAttribute('accept', 'image/*');
                    el.setAttribute('capture', 'environment');
                }});
            }}
            patch();
            setTimeout(patch, 500);
            setTimeout(patch, 1500);
        }})();
        </script>""",
        unsafe_allow_html=True,
    )



# ══════════════════════════════════════════════════════════════════════════════
# AUTH — SESSION STATE & LOGIN SCREEN
# ══════════════════════════════════════════════════════════════════════════════

if "authenticated" not in st.session_state:
    st.session_state.update({"authenticated": False, "username": "",
                              "display_name": "", "role": ""})

if not st.session_state.authenticated:
    st.title("P6 Asbuilt Collector")
    st.caption("Sign in to continue")
    st.divider()

    _, col_m, _ = st.columns([1, 1.1, 1])
    with col_m:
        with st.container(border=True):
            st.subheader("Sign In")
            username = st.text_input("Username", placeholder="Enter your username")
            password = st.text_input("Password", type="password", placeholder="Enter your password")
            if st.button("Log In", type="primary", width='stretch'):
                user = USERS.get(username)
                if user and _hcheck(password, user["hash"]):
                    st.session_state.update({
                        "authenticated": True, "username": username,
                        "display_name": user["name"], "role": user["role"],
                    })
                    st.rerun()
                else:
                    st.error("Incorrect username or password.")
    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR (authenticated)
# ══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.title("🏗️ P6 Asbuilt")
    st.divider()
    st.write(f"**{st.session_state.display_name}**")
    st.caption(ROLE_LABEL.get(st.session_state.role, st.session_state.role))
    if st.button("Log Out", width='stretch'):
        st.session_state.update({"authenticated": False, "username": "",
                                  "display_name": "", "role": ""})
        st.rerun()
    st.divider()

    # ── Shared project selector ────────────────────────────────────────────
    _all_entries_for_projects = load_entries()
    _all_projects_raw = get_all_projects(_all_entries_for_projects)
    # Filter to only projects this user can access
    _accessible = get_accessible_projects(
        st.session_state.get("username", ""), _all_projects_raw
    )
    _projects = _accessible  # no "All Projects" option

    if "selected_project" not in st.session_state or             st.session_state["selected_project"] not in _projects:
        st.session_state["selected_project"] = _projects[0] if _projects else ""

    st.session_state["selected_project"] = st.selectbox(
        "📁 Project",
        options=_projects if _projects else ["— No projects —"],
        index=_projects.index(st.session_state["selected_project"])
              if st.session_state["selected_project"] in _projects else 0,
        key="sidebar_project_select",
    )
    _sel_project = st.session_state["selected_project"]

    # ── Report Date (per-project) ──────────────────────────────────────────
    if _sel_project and _sel_project != "— No projects —":
        st.divider()
        _stored_report_date = get_report_date(_sel_project)
        _report_date_val    = _stored_report_date or date.today()

        _new_report_date = st.date_input(
            "📅 Report Date",
            value=_report_date_val,
            format="DD/MM/YYYY",
            key="sidebar_report_date",
            help="Date of the last progress report. Used to calculate expected "
                 "finish and duration % complete for In Progress activities.",
        )

        # Trigger recalculation when the date actually changes
        _prev_key = f"_prev_report_date_{_sel_project}"
        if _new_report_date != st.session_state.get(_prev_key):
            st.session_state[_prev_key] = _new_report_date
            set_report_date(_sel_project, _new_report_date)
            _recalc_entries = load_entries()
            _recalc_entries, _n_updated = recalculate_project(
                _recalc_entries, _sel_project, _new_report_date
            )
            if _n_updated:
                save_entries(_recalc_entries)
                st.caption(f"♻️ {_n_updated} activities recalculated.")

        if _stored_report_date:
            st.caption(f"Report date: {_stored_report_date.strftime('%d/%m/%Y')}")

    st.divider()
    st.caption("Entries: p6_asbuilt_store.json")
    st.caption("Photos:  p6_images/")
    st.divider()
    if has_permission("settings"):
        _unread = unread_notifications(st.session_state.get("username",""))
        st.caption(
            f"🔔  {len(_unread)} unread notification{'s' if len(_unread) != 1 else ''}"
            if _unread else "🔔  No new notifications"
        )
    st.divider()
    if st.button("🔄  Refresh", width='stretch',
                 help="Reload data from disk to see updates from other users."):
        # Clear data caches so fresh data is loaded from disk
        load_image_bytes.clear()
        build_photo_backup.clear()
        for _k in ("photo_list", "photo_map", "photo_assignments",
                   "photo_to_aids", "aid_to_pids", "_assign_sig"):
            st.session_state.pop(_k, None)
        st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# HEADER & DYNAMIC TABS
# ══════════════════════════════════════════════════════════════════════════════

#st.title("🏗️ Primavera P6 — Asbuilt Data Collector")

st.session_state["selected_project"] = st.selectbox(
    "📁 Project",
    options=_projects if _projects else ["— No projects —"],
    index=_projects.index(st.session_state["selected_project"])
          if st.session_state["selected_project"] in _projects else 0,
    key="title_project_select",
)
_sel_project = st.session_state["selected_project"]
st.divider()
logo=Path("Tricertus_logo.jpg")
st.logo(logo,size="large")

TAB_DEFS = [
    ("📋  View All Entries",  "view"),
    ("📝  Add New Activity",  "submit"),
    ("📤  Import from Excel", "import"),
    ("📥  Export to Excel",   "export"),
    ("📸  Photo Log",         "photos"),
    ("⚙️  Settings",          "settings"),
    (" Site Walk",         "sitewalk"),
]
_sw_username = st.session_state.get("username", "")
_sw_role     = st.session_state.get("role", "")

_stored_order = load_tab_order()
if _stored_order:
    _tab_dict      = {perm: lbl for lbl, perm in TAB_DEFS}
    _ordered_perms = _stored_order + [p for p in [t[1] for t in TAB_DEFS]
                                       if p not in _stored_order]
    _TAB_DEFS_ORDERED = [(_tab_dict[p], p) for p in _ordered_perms if p in _tab_dict]
else:
    _TAB_DEFS_ORDERED = TAB_DEFS

visible = [
    (lbl, perm) for lbl, perm in _TAB_DEFS_ORDERED
    if has_permission(perm)
    and is_tab_visible(perm, _sw_username, _sw_role)
]
tab_objs  = st.tabs([lbl for lbl, _ in visible])
tab_index = {perm: tab_objs[i] for i, (_, perm) in enumerate(visible)}

# ══════════════════════════════════════════════════════════════════════════════
# GANTT CHART BUILDER — reusable SVG generator
# ══════════════════════════════════════════════════════════════════════════════

def build_gantt_svg(activities: list[dict], report_date=None,
                    title: str = "") -> str:
    """
    Build a Gantt chart SVG string from a list of activity dicts.
    Each dict needs: activity_id, activity_name, activity_status,
                     actual_start, actual_finish, predicted_start,
                     remaining_dur
    Returns an SVG string or "" if no bars can be drawn.
    """
    from datetime import timedelta as _td2

    _today2 = date.today()
    _COLOURS2 = {
        "Completed":   "#16a34a",
        "In Progress": "#d97706",
        "Not Started": "#6b7280",
    }

    bars = []
    for act in activities:
        status = act.get("activity_status","")
        start = end = None
        if status == "Completed":
            s = iso_to_dt(act.get("actual_start",""))
            e = iso_to_dt(act.get("actual_finish",""))
            if s: start = s.date()
            if e: end   = e.date()
        elif status == "In Progress":
            rpt = report_date or _today2
            start = rpt
            try:    rem = int(float(act.get("remaining_dur","0") or 0))
            except: rem = 0
            end = _add_working_days(rpt, rem)
        elif status == "Not Started":
            p = iso_to_dt(act.get("predicted_start",""))
            if p:
                start = p.date()
                try:    rem = int(float(act.get("remaining_dur","0") or 0))
                except: rem = 0
                end = _add_working_days(start, rem)
        if start and end and end >= start:
            bars.append({
                "aid":    act.get("activity_id",""),
                "name":   act.get("activity_name",""),
                "status": status,
                "start":  start,
                "end":    end,
            })

    if not bars:
        return ""

    all_starts   = [b["start"] for b in bars]
    all_ends     = [b["end"]   for b in bars]
    chart_start  = min(all_starts)
    chart_end    = max(all_ends)
    total_days   = max((chart_end - chart_start).days + 1, 1)

    BAR_H   = 36;  ROW_GAP  = 10
    LABEL_W = 180; CHART_W  = 500
    HEADER_H= 52;  ROW_H    = BAR_H + ROW_GAP
    SVG_H   = HEADER_H + len(bars) * ROW_H + 20
    SVG_W   = LABEL_W + CHART_W + 10
    MAX_CH  = 24

    def _xg(d):
        return LABEL_W + int(((d - chart_start).days / total_days) * CHART_W)

    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{SVG_W}" height="{SVG_H}" '
        f'style="font-family:Arial,sans-serif;background:#f8f9fb">',
        f'<defs><clipPath id="lc2"><rect x="0" y="0" width="{LABEL_W-4}" height="{SVG_H}"/>'
        f'</clipPath></defs>',
    ]
    if title:
        parts.append(
            f'<text x="{SVG_W//2}" y="16" font-size="11" fill="#1C3557" '
            f'font-weight="bold" text-anchor="middle">{title}</text>'
        )

    # Separator
    parts.append(f'<line x1="{LABEL_W}" y1="0" x2="{LABEL_W}" y2="{SVG_H}" '
                 f'stroke="#d1d5db" stroke-width="1"/>')

    # Grid + date labels
    MIN_GAP   = 55
    last_lx   = -999
    cur       = chart_start
    while cur <= chart_end:
        if cur.weekday() == 0:
            gx = _xg(cur)
            parts.append(f'<line x1="{gx}" y1="{HEADER_H}" x2="{gx}" y2="{SVG_H}" '
                         f'stroke="#e2e8f0" stroke-width="1"/>')
            if gx - last_lx >= MIN_GAP:
                parts.append(f'<text x="{gx+2}" y="{HEADER_H-4}" font-size="9" '
                              f'fill="#94a3b8">{cur.strftime("%d %b")}</text>')
                last_lx = gx
        cur += _td2(days=1)

    # Bars
    for bi, bar in enumerate(bars):
        y   = HEADER_H + bi * ROW_H + ROW_GAP // 2
        x1  = _xg(bar["start"])
        x2  = _xg(bar["end"])
        bw  = max(x2 - x1, 4)
        col = _COLOURS2.get(bar["status"], "#6b7280")
        bg  = "#eef2f8" if bi % 2 == 0 else "#ffffff"
        parts.append(f'<rect x="0" y="{y}" width="{SVG_W}" height="{BAR_H}" fill="{bg}"/>')
        name = bar["name"]
        if len(name) <= MAX_CH:
            parts.append(f'<text x="4" y="{y+BAR_H//2+4}" font-size="10" fill="#1C3557" '
                         f'font-weight="bold" clip-path="url(#lc2)">{name}</text>')
        else:
            mid   = MAX_CH
            split = name.rfind(" ", 0, mid) or mid
            l1    = name[:split].strip()
            l2    = (name[split:].strip())[:MAX_CH-1] + ("…" if len(name[split:].strip()) > MAX_CH-1 else "")
            parts.append(f'<text x="4" y="{y+BAR_H//2-3}" font-size="10" fill="#1C3557" '
                         f'font-weight="bold" clip-path="url(#lc2)">{l1}</text>')
            parts.append(f'<text x="4" y="{y+BAR_H//2+10}" font-size="10" fill="#1C3557" '
                         f'font-weight="bold" clip-path="url(#lc2)">{l2}</text>')
        parts.append(f'<rect x="{x1}" y="{y+3}" width="{bw}" height="{BAR_H-6}" rx="3" '
                     f'fill="{col}" opacity="0.85"/>')

    # Overlay lines
    if chart_start <= _today2 <= chart_end:
        tx = _xg(_today2)
        parts.append(f'<line x1="{tx}" y1="0" x2="{tx}" y2="{SVG_H}" '
                     f'stroke="#ef4444" stroke-width="2" stroke-dasharray="4,3"/>')
        parts.append(f'<text x="{tx+2}" y="{HEADER_H//2}" font-size="9" '
                     f'fill="#ef4444" font-weight="bold">Today</text>')
    if report_date and chart_start <= report_date <= chart_end:
        rx = _xg(report_date)
        parts.append(f'<line x1="{rx}" y1="0" x2="{rx}" y2="{SVG_H}" '
                     f'stroke="#1d4ed8" stroke-width="2" stroke-dasharray="6,3"/>')
        parts.append(f'<text x="{rx+2}" y="{HEADER_H//2+12}" font-size="9" '
                     f'fill="#1d4ed8" font-weight="bold">Data date</text>')

    parts.append("</svg>")
    return "".join(parts)



# ══════════════════════════════════════════════════════════════════════════════
# TAB: VIEW ALL ENTRIES
# ══════════════════════════════════════════════════════════════════════════════

if "view" in tab_index:
    with tab_index["view"]:
        entries = load_entries()
        _sel_project = st.session_state.get("selected_project", "")
        entries = filter_by_project(entries, _sel_project)
        st.subheader(f"All Entries ({len(entries)}) — {_sel_project}")
    
        if not entries:
            st.info(f"No entries for project '{_sel_project}'." if _sel_project
                    else "No project selected.")
        else:
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Total",       len(entries))
            m2.metric("Completed",   sum(1 for e in entries if e.get("activity_status") == "Completed"))
            m3.metric("In Progress", sum(1 for e in entries if e.get("activity_status") == "In Progress"))
            m4.metric("Not Started", sum(1 for e in entries if e.get("activity_status") == "Not Started"))
            st.divider()
    
            # ── Search + Sort controls ─────────────────────────────────────────
            search_col, sort_col, dir_col = st.columns([3, 2, 1])
            with search_col:
                search_text = st.text_input(
                    "Search",
                    placeholder="Activity ID or name…",
                    key="view_search",
                ).strip().lower()
            with sort_col:
                sort_by = st.selectbox(
                    "Sort by",
                    options=["WBS Code", "Actual Start", "Actual Finish", "Activity ID"],
                    key="view_sort_by",
                )
            with dir_col:
                sort_asc = st.radio(
                    "Order", options=["↑ Asc", "↓ Desc"],
                    key="view_sort_dir", horizontal=True,
                ) == "↑ Asc"
    
            # Apply search filter before sorting
            if search_text:
                entries = [
                    e for e in entries
                    if search_text in str(e.get("activity_id",   "") or "").lower()
                    or search_text in str(e.get("activity_name", "") or "").lower()
                    or search_text in str(e.get("wbs_id",        "") or "").lower()
                ]
    
            def wbs_key(e: dict):
                parts = str(e.get("wbs_id", "") or "").split(".")
                segments = []
                for p in parts:
                    try:
                        segments.append((0, int(p)))
                    except ValueError:
                        segments.append((1, p.lower()))
                return segments or [(1, "")]
    
            def date_key(field: str):
                def _key(e: dict):
                    dt = iso_to_dt(e.get(field, ""))
                    return dt if dt else (datetime.min if sort_asc else datetime.max)
                return _key
    
            if sort_by == "WBS Code":
                sorted_entries = sorted(entries, key=wbs_key, reverse=not sort_asc)
            elif sort_by == "Actual Start":
                sorted_entries = sorted(entries, key=date_key("actual_start"), reverse=not sort_asc)
            elif sort_by == "Actual Finish":
                sorted_entries = sorted(entries, key=date_key("actual_finish"), reverse=not sort_asc)
            else:
                sorted_entries = sorted(entries, key=lambda e: e.get("activity_id", "").upper(), reverse=not sort_asc)
    
            st.divider()
    
            can_edit   = has_permission("submit")
            can_delete = has_permission("submit")
    
            # ── Pagination ─────────────────────────────────────────────────────
            PAGE_SIZE = 25
            total_pages = max(1, (len(sorted_entries) + PAGE_SIZE - 1) // PAGE_SIZE)
            page = st.number_input(
                f"Page (1 – {total_pages})",
                min_value=1, max_value=total_pages, value=1, step=1,
                key="view_page",
            ) - 1  # zero-based
            page_entries = sorted_entries[page * PAGE_SIZE : (page + 1) * PAGE_SIZE]
            total_label  = f"{len(sorted_entries)} match{'es' if len(sorted_entries) != 1 else ''}" if search_text else f"{len(sorted_entries)} entries"
            st.caption(
                f"Showing {page * PAGE_SIZE + 1}–{min((page + 1) * PAGE_SIZE, len(sorted_entries))} "
                f"of {total_label}"
            )
            st.divider()
    
            # Keyed by (activity_id, project) to support same IDs across projects
            id_to_index = {
                (e.get("activity_id","").upper(),
                 get_project_from_wbs(e.get("wbs_id",""))): idx
                for idx, e in enumerate(entries)
            }
    
            for entry in page_entries:
                i     = id_to_index.get(
                    (entry.get("activity_id","").upper(),
                     get_project_from_wbs(entry.get("wbs_id",""))), 0)
                status    = entry.get("activity_status", "")
                act_id    = entry.get("activity_id", "")
                act_name  = entry.get("activity_name", "")
                wbs       = entry.get("wbs_id", "—")
                pct       = entry.get("pct_complete", "0")
                rem       = entry.get("remaining_dur", "—")
                a_start   = display_dt(entry.get("actual_start", ""))
                a_finish  = display_dt(entry.get("actual_finish", ""))
                subm_at   = entry.get("_submitted_at", "")
                submitter = entry.get("_submitted_by", "")
                n_comments = len(entry.get("_comments", []))
    
                with st.container(border=True):
                    # ── Header row ─────────────────────────────────────────────
                    head_left, head_left2, head_right = st.columns([2, 2, 1])
                    with head_left:
                        st.write(f"Activity ID: {act_id}")
                    with head_left2:
                        st.write(f"Activity Name: {act_name}")
                    with head_right:
                        st.write(status)
    
                    # ── Detail row ─────────────────────────────────────────────
                    c1, c2, c3, c4, c5 = st.columns([1, 1, 1, 0.3, 0.4], gap="xsmall")
                    c1.write("WBS");      c1.write(wbs)
                    c2.write("Start");    c2.write(a_start)
                    # Show expected finish for In Progress if report date is set
                    _rpt_view = get_report_date(get_project_from_wbs(entry.get("wbs_id","")))
                    _exp_f_str = ""
                    if status == "In Progress" and _rpt_view and entry.get("remaining_dur"):
                        _ef_view = expected_finish_date(_rpt_view, entry.get("remaining_dur",""))
                        if _ef_view:
                            _exp_f_str = f" (exp. {_ef_view.strftime('%d/%m/%Y')})"
                    c3.write("Finish");   c3.write(a_finish + _exp_f_str if _exp_f_str else a_finish)
                    c4.metric("% Complete", f"{pct}%")
                    c5.metric("Remaining", f"{rem} days" if rem and rem != "—" else "—")
    
                    # ── Footer ─────────────────────────────────────────────────
                    footer = f"Last updated: {subm_at}"
                    if submitter:
                        footer += f"  ·  By: {submitter}"
                    if n_comments:
                        footer += f"  ·  💬 {n_comments} comment{'s' if n_comments != 1 else ''}"
                    st.caption(footer)
    
                    # ── Inline edit expander (submit/admin only) ───────────────
                    if can_edit:
                        with st.expander("✏️  Edit name / Add comment"):
                            # Activity name
                            st.write("**Edit Activity Name**")
                            new_name = st.text_input(
                                "Activity Name",
                                value=act_name,
                                key=f"edit_name_{i}",
                                label_visibility="collapsed",
                            ).strip()
    
                            st.divider()
    
                            # Existing comments
                            st.write("**Comments**")
                            existing_comments = entry.get("_comments", [])
                            if existing_comments:
                                for c in existing_comments:
                                    st.write(f"**{c['at']}** — {c['by']}")
                                    st.write(c["text"])
                                    st.divider()
                            else:
                                st.caption("No comments yet.")
    
                            new_comment_text = st.text_area(
                                "Add comment",
                                placeholder="Enter progress notes, observations, or issues...",
                                height=100,
                                key=f"view_comment_{i}",
                                label_visibility="collapsed",
                            ).strip()
    
                            # Save button — only active if something changed
                            name_changed    = new_name != act_name and new_name != ""
                            comment_entered = bool(new_comment_text)
    
                            if st.button(
                                "💾  Save changes",
                                key=f"edit_save_{i}",
                                type="primary",
                                disabled=not (name_changed or comment_entered),
                            ):
                                updated = entry.copy()
                                if name_changed:
                                    updated["activity_name"] = new_name
                                if comment_entered:
                                    new_record = {
                                        "text": new_comment_text,
                                        "by":   st.session_state.display_name,
                                        "at":   datetime.now().strftime("%d/%m/%Y %H:%M"),
                                    }
                                    updated["_comments"] = [new_record] + existing_comments
                                updated["_submitted_at"] = datetime.now().strftime("%d/%m/%Y %H:%M")
                                updated["_submitted_by"] = st.session_state.display_name
                                entries[i] = updated
                                save_entries(entries)
                                st.success("Saved.")
                                st.rerun()
    
                    # ── Delete button ──────────────────────────────────────────
                    if can_delete and st.button(f"🗑 Delete {act_id}", key=f"del_{i}"):
                        entries.pop(i)
                        save_entries(entries)
                        st.rerun()
    
            st.dataframe(sorted_entries)
    
    # ══════════════════════════════════════════════════════════════════════════════
    # TAB: Add new activity
    # ══════════════════════════════════════════════════════════════════════════════
    
if "submit" in tab_index:
    with tab_index["submit"]:
        _sel_project = st.session_state.get("selected_project", "")

        st.subheader("➕ Add New Activity Request")
        st.write(
            "Use this form to request a new activity be added to the project. "
            "A notification will be sent to the project admins who will add it "
            "to Primavera P6 and import it."
        )
        if not _sel_project:
            st.warning("Select a project in the sidebar first.")
            st.stop()

        st.caption(f"Project: **{_sel_project}**")
        st.divider()

        # ── Core fields ───────────────────────────────────────────────────────
        _today       = date.today()
        
        req_name = st.text_input(
            "Activity Name *",
            placeholder="e.g. Concrete Pour — Level 3 Slab",
            key="req_activity_name",
        ).strip()

        req_type = st.text_input(
            "Type of Activity *",
            placeholder="e.g. RFI, Rework, Resubmittal, Variation, NCR",
            key="req_activity_type",
        ).strip()

        req_area = st.text_input(
            "Area of Work *",
            placeholder="e.g. North Wing, Level 3, Grid C-D",
            key="req_area",
        ).strip()

        req_status = st.selectbox(
            "Status *",
            STATUS_OPTIONS,
            key="req_status",
        )
        

        rq_start_dt = rq_finish_dt = None
        rq_pct = 0
        rq_rem = 0
        


        if req_status == "Not Started":
            st.info("% Complete set to 0 automatically.", icon="ℹ️")
            rq_exp = st.text_input(
                    "Expected Duration (days) *",
                    value=str(""),
                    placeholder="e.g. 5",
                    key="req_status_NS"
                    ).strip()
                        
        elif req_status == "Completed":
            rq_start_dt = datetime_inputs(
                "Actual Start *", key="rq_start", required=True,
                default_dt=_today,
            )
            rq_finish_dt = datetime_inputs(
                "Actual Finish *", key="rq_finish", required=True,
                default_dt=_today,
            )
            rq_pct = 100
            rq_rem = "0"
            st.info("% Complete set to 100, Remaining to 0.", icon="✅")    
            
        elif req_status == "In Progress":
                rq_start_dt = datetime_inputs(
                    "Actual Start *", key="req_status_IP")
        
                # ── Input mode selector ───────────────────────────────
                _ip_mode_rq = st.radio(
                "Update by",
                ["Duration % Complete", "Remaining Duration", "Expected Finish",
                 "Physical % Complete"],
                horizontal=True,
                key="rq_ip_mode",
                )

                rq_pct = int(0)
                rq_rem = str("")
                _rpt = _today

                if _ip_mode_rq == "Duration % Complete":
                    rq_pct = st.number_input(
                        "Duration % Complete *",
                        min_value=0, max_value=99, step=5,
                        value=rq_pct,
                        key="rq_pct",
                        )
                    if rq_start_dt and rq_pct > 0:
                        _elapsed  = working_days_between(rq_start_dt.date(), _rpt) + 1
                        _total    = int(_elapsed / (rq_pct / 100))
                        _rem_calc = max(0, _total - _elapsed)
                        rq_rem    = str(_rem_calc)
                        _ef_calc  = _add_working_days(_rpt, _rem_calc)
                        st.caption(
                            f"💡 Remaining: **{_rem_calc} days**  ·  "
                            f"Expected finish: **{_ef_calc.strftime('%d/%m/%Y')}**"
                        )
    
                elif _ip_mode_rq == "Remaining Duration":
                    rq_rem = st.text_input(
                        "Remaining Duration (days) *",
                        value=rq_rem,
                        placeholder="e.g. 5",
                        key="rq_rem",
                        ).strip()
                    if rq_rem and rq_start_dt:
                        _ef_calc  = expected_finish_date(_rpt, rq_rem)
                        _pct_calc = calc_duration_pct(
                            dt_to_iso(rq_start_dt), _ef_calc, _rpt
                        ) if _ef_calc else None
                        rq_pct = _pct_calc if _pct_calc is not None else rq_pct
                        if _ef_calc:
                            st.caption(
                                f"💡 % Complete: **{rq_pct}%**  ·  "
                                f"Expected finish: **{_ef_calc.strftime('%d/%m/%Y')}**"
                            )
        
                elif _ip_mode_rq == "Expected Finish":
                    _cur_ef = _add_working_days(
                        _rpt, int(float(rq_rem)) if rq_rem else 0
                    ) if rq_rem else _rpt
                    _ef_input = st.date_input(
                        "Expected Finish *",
                        value=_cur_ef,
                        format="DD/MM/YYYY",
                        key="rq_ef",
                    )
                    _rem_calc = working_days_between(_rpt, _ef_input)
                    rq_rem    = str(_rem_calc)
                    _pct_calc = calc_duration_pct(
                        dt_to_iso(rq_start_dt), _ef_input, _rpt
                    ) if rq_start_dt else None
                    rq_pct = _pct_calc if _pct_calc is not None else rq_pct
                    st.caption(
                        f"💡 Remaining: **{_rem_calc} days**  ·  "
                        f"% Complete: **{rq_pct}%**"
                    )
        
                else:  # Physical % Complete
                    st.caption(
                        "Enter the physical % complete directly. "
                        "Then select how to specify the remaining schedule."
                    )
                    rq_pct = st.number_input(
                        "Physical % Complete *",
                        min_value=0, max_value=99, step=1,
                        value=rq_pct,
                        key="rq_phys_pct",
                    )
                    # Second selector — how to supply remaining schedule info
                    _phys_sub = st.selectbox(
                        "Also specify",
                        ["Remaining Duration", "Expected Finish",
                         "Duration % Complete"],
                        key="rq_phys_sub",
                    )
                    if _phys_sub == "Remaining Duration":
                        rq_rem = st.text_input(
                            "Remaining Duration (days)",
                            value=rq_rem,
                            placeholder="e.g. 5",
                            key="rq_phys_rem",
                        ).strip()
                        if rq_rem:
                            _ef_calc = expected_finish_date(_rpt, rq_rem)
                            if _ef_calc:
                                st.caption(
                                    f"💡 Expected finish: "
                                    f"**{_ef_calc.strftime('%d/%m/%Y')}**"
                                )
        
                    elif _phys_sub == "Expected Finish":
                        _cur_ef = _add_working_days(
                            _rpt, int(float(rq_rem)) if rq_rem else 0
                        ) if rq_rem else _rpt
                        _ef_input = st.date_input(
                            "Expected Finish",
                            value=_cur_ef,
                            format="DD/MM/YYYY",
                            key="rq_phys_ef",
                        )
                        _rem_calc = working_days_between(_rpt, _ef_input)
                        rq_rem    = str(_rem_calc)
                        st.caption(
                            f"💡 Remaining: **{_rem_calc} days**"
                        )
        
                    else:  # Duration % Complete
                        _dur_pct = st.number_input(
                            "Duration % Complete",
                            min_value=0, max_value=99, step=5,
                            value=int(0),
                            key="rq_phys_dur",
                        )
                        if rq_start_dt and _dur_pct > 0:
                            _elapsed  = working_days_between(
                                rq_start_dt.date(), _rpt) + 1
                            _total    = int(_elapsed / (_dur_pct / 100))
                            _rem_calc = max(0, _total - _elapsed)
                            rq_rem    = str(_rem_calc)
                            _ef_calc  = _add_working_days(_rpt, _rem_calc)
                            st.caption(
                                f"💡 Remaining: **{_rem_calc} days**  ·  "
                                f"Expected finish: **{_ef_calc.strftime('%d/%m/%Y')}**"
                            )

        
        
        
        
        # ── Comment + attachments ─────────────────────────────────────────────
        st.divider()
        st.subheader("Comment & Attachments")

        req_comment = st.text_area(
            "Comment *",
            placeholder="Describe the activity, reason for addition, location details, etc.",
            height=140,
            key="req_comment",
        ).strip()

        req_files = st.file_uploader(
            "Attach supporting files (PDF or images, optional)",
            type=["pdf","jpg","jpeg","png","webp"],
            accept_multiple_files=True,
            key="req_attachments",
        )

        # ── Submit ────────────────────────────────────────────────────────────
        st.divider()
        if st.button("📤  Send Request to Admin", type="primary"):
            errs = []
            if not req_name:    errs.append("Activity Name is required.")
            if not req_type:    errs.append("Type of Activity is required.")
            if not req_area:    errs.append("Area of Work is required.")
            if not req_comment: errs.append("Comment is required.")
            if errs:
                for e in errs:
                    st.error(e)
            else:
                # Save attachments to disk and collect metadata
                ensure_photo_dir()
                _att_meta = []
                for _f in (req_files or []):
                    _ext  = Path(_f.name).suffix.lower() or ".bin"
                    _fname = f"req_{uuid.uuid4().hex}{_ext}"
                    _fpath = PHOTO_DIR / _fname
                    _fbytes = _f.read()
                    _fpath.write_bytes(_fbytes)
                    _att_meta.append({
                        "original_name": _f.name,
                        "filename":      _fname,
                        "size_kb":       round(len(_fbytes) / 1024, 1),
                    })

                # Build notification body
                _notif_body = (
                    f"**New activity request** from "
                    f"**{st.session_state.display_name}**\n\n"
                    f"- **Activity Name:** {req_name}\n"
                    f"- **Type:** {req_type}\n"
                    f"- **Area:** {req_area}\n"
                    f"- **Status:** {req_status}\n\n"
                    f"**Comment:**\n{req_comment}"
                )
                if _att_meta:
                    _notif_body += (
                        f"\n\n**Attachments ({len(_att_meta)}):** "
                        + ", ".join(a["original_name"] for a in _att_meta)
                    )

                create_notification(
                    created_by  = st.session_state.display_name,
                    project     = _sel_project,
                    title       = f"New Activity Request — {req_name} — {_sel_project}",
                    body        = _notif_body,
                    rows        = [
                        {"Field": "Activity Name", "Value": req_name},
                        {"Field": "Type",          "Value": req_type},
                        {"Field": "Area of Work",  "Value": req_area},
                        {"Field": "Status",        "Value": req_status},
                        {"Field": "remaining Duration", "Value":rq_rem},
                        {"Field": "Start date", "Value": rq_start_dt.strftime('%Y-%m-%d') if rq_start_dt else None}, 
                        {"Field": "Actual finish", "Value": rq_finish_dt.strftime('%Y-%m-%d') if rq_start_dt else None}, 
                        {"Field": "Attachments",   "Value":
                            ", ".join(a["original_name"] for a in _att_meta)
                            if _att_meta else "None"},
                    ],
                    status      = "info",
                    recipients  = get_project_admin_recipients(_sel_project),
                    edits       = {"attachments": _att_meta},
                )

                st.success(
                    f"✅ Request sent! Admins for **{_sel_project}** have been notified. "
                    f"The activity will be added to P6 and imported once approved."
                )
                if _att_meta:
                    st.caption(
                        f"{len(_att_meta)} file{'s' if len(_att_meta) != 1 else ''} "
                        f"uploaded: "
                        + ", ".join(
                            f"{a['original_name']} ({a['size_kb']} KB)"
                            for a in _att_meta
                        )
                    )

if "import" in tab_index:
    with tab_index["import"]:
        entries   = load_entries()
        _sel_project = st.session_state.get("selected_project", "")
        _entries_for_match = filter_by_project(entries, _sel_project)
        known_ids = {e["activity_id"].upper(): e for e in _entries_for_match}

        if _sel_project:
            st.info(f"Importing into project: **{_sel_project}**", icon="📁")

        st.subheader("Import from Excel")

        import_mode = st.radio(
            "Source format",
            options=["Primavera P6 Export", "Microsoft Project Export"],
            horizontal=True,
            key="import_mode",
        )

        if import_mode == "Primavera P6 Export":
            st.caption("Upload a P6-format XLSX (TASK sheet). Conflicts matched by Activity ID.")
        else:
            st.caption(
                "Upload a Microsoft Project XLSX export. "
                "Activities are matched to stored entries by **Name + WBS**. "
                "MS Project WBS codes are compared after stripping the P6 project prefix."
            )

        uploaded = st.file_uploader("Choose an Excel file", type=["xlsx", "xls"], key="import_file")

        if uploaded:
            file_bytes = uploaded.read()
            with st.spinner("Reading Excel file..."):
                if import_mode == "Primavera P6 Export":
                    imported_rows, warnings_list = read_p6_excel(file_bytes)
                else:
                    imported_rows, warnings_list = read_msp_excel(file_bytes)

            for w in warnings_list:
                st.warning(w)

            if not imported_rows:
                st.error("No valid rows found. Check the file format and column headers.")

            elif import_mode == "Microsoft Project Export":
                # ── MSP matching flow ──────────────────────────────────────
                st.divider()
                matched_all, unmatched, duplicates = match_msp_to_stored(imported_rows, _entries_for_match)

                # Drop matched rows where the incoming data is identical to stored
                matched          = [(r, s) for r, s in matched_all
                                    if not is_exact_duplicate(r, s)]
                identical_skipped = len(matched_all) - len(matched)

                msg = (
                    f"Found **{len(imported_rows)}** rows: "
                    f"**{len(matched)}** matched with changes, "
                    f"**{len(unmatched)}** unmatched, "
                    f"**{len(duplicates)}** ambiguous"
                )
                if identical_skipped:
                    msg += f", **{identical_skipped}** identical (skipped)"
                st.success(msg)

                # ── Duplicate name+WBS conflicts ──────────────────────────
                if duplicates:
                    st.divider()
                    st.subheader("⚠️ Ambiguous matches — multiple activities share the same Name and WBS")
                    st.caption(
                        "The activities below could not be matched uniquely. "
                        "Rename one of the stored activities to make it unique, "
                        "then re-import."
                    )
                    for msp_row, hits in duplicates:
                        with st.expander(
                            f"❓  {msp_row['activity_name']}  —  WBS: {msp_row['wbs_id']}",
                            expanded=True,
                        ):
                            st.write(f"**{len(hits)} stored activities match this name and WBS:**")
                            for h in hits:
                                st.write(
                                    f"- `{h['activity_id']}`  {h['activity_name']}  "
                                    f"WBS: {h['wbs_id']}  Status: {h['activity_status']}"
                                )
                            st.warning(
                                "Rename one of these activities in the Submit/Update tab "
                                "or View All Entries before re-importing."
                            )

                # ── WBS offset detection ─────────────────────────────────
                # Run before rendering unmatched cards so suggestions can be
                # applied to stored entries before the user handles each row.
                wbs_applied_offsets = {}  # suggestion_key → bool (applied)

                if unmatched:
                    offset_suggestions = detect_wbs_offset(unmatched, _entries_for_match)

                    # Deduplicate suggestions by (depth, delta, prefix, from_val)
                    # so we only show one prompt per unique shift
                    seen_shifts  = {}   # shift_key → suggestion
                    for sg in offset_suggestions:
                        segs     = sg["stored_wbs"].split(".")
                        prefix   = ".".join(segs[:sg["depth"]])
                        from_val = int(segs[sg["depth"]])
                        shift_key = (prefix, sg["depth"], sg["delta"], from_val)
                        if shift_key not in seen_shifts:
                            seen_shifts[shift_key] = sg

                    if seen_shifts:
                        st.divider()
                        st.subheader("🔀 Possible WBS offset detected")
                        st.caption(
                            "The following unmatched activities were found at adjacent WBS codes "
                            "with the same name — this may indicate a WBS level was inserted or "
                            "deleted, shifting all subsequent codes. Review each suggestion and "
                            "apply the shift if correct. **Applying a shift updates the stored "
                            "WBS codes immediately and cannot be undone here — save a backup first.**"
                        )

                        for shift_key, sg in seen_shifts.items():
                            prefix, depth, delta, from_val = shift_key
                            direction = "up (+1)" if delta > 0 else "down (−1)"
                            parent    = prefix if prefix else "root"
                            affected  = [
                                e for e in entries
                                if (lambda c, s, d, fv, dl:
                                    len(s) > d
                                    and (not c or ".".join(s[:d]) == c)
                                    and (int(s[d]) >= fv if dl > 0 else int(s[d]) <= fv)
                                )(
                                    prefix,
                                    strip_wbs_prefix(e.get("wbs_id","")).split("."),
                                    depth,
                                    from_val,
                                    delta,
                                )
                            ]

                            with st.expander(
                                f"Shift WBS segment {depth+1} {direction} "
                                f"under '{parent}' from position {from_val} onwards "
                                f"— affects ~{len(affected)} activities",
                                expanded=True,
                            ):
                                c_msp, c_stored = st.columns(2)
                                with c_msp:
                                    st.write("**Unmatched MSP activity:**")
                                    st.write(f"- Name: {sg['msp_row'].get('activity_name','')}")
                                    st.write(f"- WBS: `{sg['msp_wbs']}`")
                                with c_stored:
                                    st.write("**Stored activity found at adjacent WBS:**")
                                    st.write(f"- Name: {sg['stored_entry'].get('activity_name','')}")
                                    st.write(f"- Stored WBS: `{sg['stored_entry'].get('wbs_id','')}`")
                                    st.write(f"- ID: `{sg['stored_entry'].get('activity_id','')}`")

                                st.write(f"**Activities that would be shifted ({len(affected)}):**")
                                preview = affected[:5]
                                for e in preview:
                                    old_w = strip_wbs_prefix(e.get("wbs_id",""))
                                    segs2 = old_w.split(".")
                                    segs2[depth] = str(int(segs2[depth]) + delta)
                                    st.caption(
                                        f"- `{e['activity_id']}` {e['activity_name']}  "
                                        f"{old_w} → {'.'.join(segs2)}"
                                    )
                                if len(affected) > 5:
                                    st.caption(f"  … and {len(affected)-5} more")

                                if st.button(
                                    f"✅ Apply this WBS shift",
                                    key=f"apply_wbs_shift_{'_'.join(str(x) for x in shift_key)}",
                                    type="primary",
                                ):
                                    entries, n_changed = apply_wbs_offset(
                                        entries, prefix, depth, delta, from_val
                                    )
                                    save_entries(entries)
                                    # Rebuild known_ids after the shift
                                    known_ids = {e["activity_id"].upper(): e for e in entries}
                                    st.success(
                                        f"WBS shift applied — {n_changed} activities updated. "
                                        f"Re-run the import to see updated matches."
                                    )
                                    st.rerun()

                # ── Unmatched rows — add as new or manually map to existing ─
                # new_activity_ids:   {row_idx → new Activity ID string}
                # manual_overwrites:  {row_idx → stored activity_id to overwrite}
                new_activity_ids  = {}
                manual_overwrites = {}

                if unmatched:
                    st.divider()
                    st.subheader(f"➕ {len(unmatched)} unmatched rows")
                    st.caption(
                        "These rows had no automatic Name + WBS match. "
                        "For each row you can: **Add as new** (enter a new Activity ID), "
                        "**Overwrite existing** (manually select the stored activity it should update), "
                        "or leave both blank to skip."
                    )

                    # Build a readable option list for the overwrite selectbox
                    overwrite_options = ["— Select stored activity —"] + [
                        f"{e['activity_id']}  —  {e['activity_name']}  (WBS: {e['wbs_id']})"
                        for e in entries
                    ]
                    # Map display string back to activity_id
                    option_to_aid = {
                        f"{e['activity_id']}  —  {e['activity_name']}  (WBS: {e['wbs_id']})": e["activity_id"]
                        for e in entries
                    }

                    for idx, row in enumerate(unmatched):
                        with st.container(border=True):
                            # Row info
                            st.write(f"**{row['activity_name']}**")
                            st.caption(
                                f"WBS: {row['wbs_id'] or '—'}"
                                f"  ·  Status: {row.get('activity_status','—')}"
                                f"  ·  % Complete: {row.get('pct_complete','—')}"
                                f"  ·  Start: {display_dt(row.get('actual_start',''))}"
                                f"  ·  Finish: {display_dt(row.get('actual_finish',''))}"
                            )

                            # Mode toggle
                            mode = st.radio(
                                "Action",
                                ["Skip", "Add as new", "Overwrite existing"],
                                key=f"unmatched_mode_{idx}",
                                horizontal=True,
                            )

                            if mode == "Add as new":
                                c_id, c_status = st.columns([2, 1])
                                with c_id:
                                    proposed_id = st.text_input(
                                        "New Activity ID",
                                        placeholder="e.g. A1050",
                                        key=f"new_act_id_{idx}",
                                    ).strip()
                                with c_status:
                                    st.write("")  # spacer
                                    if proposed_id:
                                        if proposed_id.upper() in known_ids:
                                            st.error("ID already exists.")
                                        elif proposed_id.upper() in {
                                            v.upper() for v in new_activity_ids.values() if v
                                        }:
                                            st.error("ID used twice above.")
                                        else:
                                            st.success("✓")
                                            new_activity_ids[idx] = proposed_id

                            elif mode == "Overwrite existing":
                                st.caption(
                                    "Use this when the MSP name differs slightly from the stored name "
                                    "(e.g. 'ActivityName (detail)' vs 'ActivityName'). "
                                    "Only progress fields will be updated — the stored name and ID are kept."
                                )
                                selected_opt = st.selectbox(
                                    "Select stored activity to overwrite",
                                    options=overwrite_options,
                                    key=f"overwrite_select_{idx}",
                                )
                                if selected_opt != "— Select stored activity —":
                                    target_aid = option_to_aid[selected_opt]
                                    # Warn if this target is already being overwritten by another row
                                    already_used = [
                                        i for i, a in manual_overwrites.items()
                                        if a.upper() == target_aid.upper() and i != idx
                                    ]
                                    if already_used:
                                        st.error(
                                            f"This activity is already targeted by another row above."
                                        )
                                    else:
                                        manual_overwrites[idx] = target_aid
                                        stored_e = known_ids[target_aid.upper()]
                                        c_cur, c_new = st.columns(2)
                                        with c_cur:
                                            st.write("**Stored (will be kept):**")
                                            st.write(f"- Name: {stored_e.get('activity_name','')}")
                                            st.write(f"- Status: {stored_e.get('activity_status','—')}")
                                            st.write(f"- Start: {display_dt(stored_e.get('actual_start',''))}")
                                            st.write(f"- Finish: {display_dt(stored_e.get('actual_finish',''))}")
                                            st.write(f"- % Complete: {stored_e.get('pct_complete','—')}")
                                        with c_new:
                                            st.write("**Incoming (will overwrite progress):**")
                                            st.write(f"- Name: {row.get('activity_name','')}")
                                            st.write(f"- Status: {row.get('activity_status','—')}")
                                            st.write(f"- Start: {display_dt(row.get('actual_start',''))}")
                                            st.write(f"- Finish: {display_dt(row.get('actual_finish',''))}")
                                            st.write(f"- % Complete: {row.get('pct_complete','—')}")

                # ── Matched rows resolution ───────────────────────────────
                msp_resolutions = {}
                if matched:
                    st.divider()
                    st.subheader(f"✅ {len(matched)} matched rows — review and confirm")
                    st.caption("Each MSP row has been matched to a stored activity. Review the changes below.")

                    fields = [
                        ("Status",          "activity_status",  False),
                        ("Actual Start",     "actual_start",     True),
                        ("Actual Finish",    "actual_finish",    True),
                        ("% Complete",       "pct_complete",     False),
                        ("Remaining (days)", "remaining_dur",    False),
                        ("Predicted Start",  "predicted_start",  True),
                        ("Task Type",        "task_type",        False),
                    ]
                    for msp_row, stored_entry in matched:
                        aid = stored_entry["activity_id"].upper()
                        label = (
                            f"🔁  `{stored_entry['activity_id']}`  "
                            f"{stored_entry['activity_name']}  —  WBS: {stored_entry['wbs_id']}"
                        )
                        with st.expander(label, expanded=False):
                            c_cur, c_new = st.columns(2)
                            with c_cur:
                                st.write("**Current (stored)**")
                                for lbl, k, is_date in fields:
                                    val = stored_entry.get(k, "") or ""
                                    st.write(f"- **{lbl}:** {display_dt(val) if is_date else (val or '—')}")
                            with c_new:
                                st.write("**Incoming (MS Project)**")
                                for lbl, k, is_date in fields:
                                    val = msp_row.get(k, "") or ""
                                    st.write(f"- **{lbl}:** {display_dt(val) if is_date else (val or '—')}")

                            choice = st.radio(
                                "Action",
                                ["Overwrite with incoming", "Keep current"],
                                key=f"msp_conflict_{aid}",
                                horizontal=True,
                            )
                            msp_resolutions[aid] = {
                                "action":  "overwrite" if choice == "Overwrite with incoming" else "skip",
                                "msp_row": msp_row,
                                "stored":  stored_entry,
                            }

                    st.divider()
                    ow = sum(1 for v in msp_resolutions.values() if v["action"] == "overwrite")
                    sk = sum(1 for v in msp_resolutions.values() if v["action"] == "skip")
                    parts = []
                    if ow: parts.append(f"{ow} activities will be updated")
                    if sk: parts.append(f"{sk} matches will be skipped")
                    if duplicates: parts.append(f"{len(duplicates)} ambiguous rows skipped")
                    st.info("  ·  ".join(parts) if parts else "Nothing to import.")

                    # ── Validation ───────────────────────────────────────
                    _ids_entered   = list(new_activity_ids.values())
                    _id_dupes      = len(_ids_entered) != len(set(i.upper() for i in _ids_entered))
                    _id_clash      = any(i.upper() in known_ids for i in _ids_entered)
                    _ow_dupes      = len(manual_overwrites) != len(set(
                                         a.upper() for a in manual_overwrites.values()))
                    _can_confirm   = bool(msp_resolutions) or bool(new_activity_ids) or bool(manual_overwrites)
                    _blocked       = _id_dupes or _id_clash or _ow_dupes

                    # ── Summary ───────────────────────────────────────────
                    ow       = sum(1 for v in msp_resolutions.values() if v["action"] == "overwrite")
                    sk       = sum(1 for v in msp_resolutions.values() if v["action"] == "skip")
                    n_new    = len(new_activity_ids)
                    n_manual = len(manual_overwrites)
                    n_skip   = len(unmatched) - n_new - n_manual
                    parts    = []
                    if ow:       parts.append(f"{ow} auto-matched activities will be updated")
                    if sk:       parts.append(f"{sk} auto-matches skipped")
                    if n_new:    parts.append(f"{n_new} new activities will be added")
                    if n_manual: parts.append(f"{n_manual} manual overwrites will be applied")
                    if n_skip:   parts.append(f"{n_skip} unmatched rows skipped")
                    if duplicates: parts.append(f"{len(duplicates)} ambiguous rows skipped")
                    st.info("  ·  ".join(parts) if parts else "Nothing to import.")

                    if _id_dupes:  st.error("Duplicate Activity IDs in the 'Add as new' fields above.")
                    if _id_clash:  st.error("One or more new Activity IDs already exist in the store.")
                    if _ow_dupes:  st.error("The same stored activity is targeted by more than one manual overwrite.")

                    if st.button(
                        "✅  Confirm MSP Import", type="primary",
                        disabled=(not _can_confirm or _blocked),
                    ):
                        entries       = load_entries()
                        entries_index = {
                            (e.get("activity_id","").upper(),
                             get_project_from_wbs(e.get("wbs_id",""))): idx
                            for idx, e in enumerate(entries)
                        }
                        updated = added = manually_updated = 0

                        # Helper: merge MSP progress into a stored entry in-place
                        def _apply_msp(stored: dict, msp_row: dict) -> dict:
                            stored["activity_status"] = msp_row.get("activity_status") or stored["activity_status"]
                            stored["actual_start"]    = msp_row.get("actual_start")    or stored["actual_start"]
                            stored["actual_finish"]   = msp_row.get("actual_finish")   or stored["actual_finish"]
                            stored["pct_complete"]    = msp_row.get("pct_complete")    or stored["pct_complete"]
                            stored["remaining_dur"]   = msp_row.get("remaining_dur")   or stored["remaining_dur"]
                            imp_str = msp_row.pop("comments_import", "") or ""
                            if imp_str:
                                _imp_cmts   = import_string_to_comments(imp_str, st.session_state.display_name)
                                _exist_cmts = stored.get("_comments", [])
                                _new_only   = merge_imported_comments(_imp_cmts, _exist_cmts)
                                if _new_only:
                                    stored["_comments"] = _new_only + _exist_cmts
                            stored["_submitted_at"] = datetime.now().strftime("%d/%m/%Y %H:%M")
                            stored["_submitted_by"] = st.session_state.display_name
                            return stored

                        # Auto-matched updates
                        for aid, res in msp_resolutions.items():
                            if res["action"] == "overwrite":
                                _proj_res = get_project_from_wbs(
                                    res["stored"].get("wbs_id",""))
                                idx = entries_index.get((aid, _proj_res))
                                if idx is not None:
                                    entries[idx] = _apply_msp(entries[idx], res["msp_row"])
                                    updated += 1

                        # Manual overwrites from unmatched rows
                        for row_idx, target_aid in manual_overwrites.items():
                            _stored_m = known_ids.get(target_aid.upper(), {})
                            _proj_m   = get_project_from_wbs(_stored_m.get("wbs_id",""))
                            idx = entries_index.get((target_aid.upper(), _proj_m))
                            if idx is not None:
                                entries[idx] = _apply_msp(entries[idx], unmatched[row_idx])
                                manually_updated += 1

                        # New activities from unmatched rows
                        for idx, new_id in new_activity_ids.items():
                            if not new_id:
                                continue
                            row     = unmatched[idx]
                            imp_str = row.pop("comments_import", "") or ""
                            entries.append({
                                "activity_id":       new_id,
                                "activity_name":     row.get("activity_name", ""),
                                "activity_status":   row.get("activity_status", "Not Started"),
                                "actual_start":      row.get("actual_start", ""),
                                "actual_finish":     row.get("actual_finish", ""),
                                "pct_complete":      row.get("pct_complete", "0"),
                                "remaining_dur":     row.get("remaining_dur", 0),
                                "complete_pct_type": "Physical",
                                "wbs_id":            row.get("wbs_id", ""),
                                "_comments":         import_string_to_comments(imp_str, st.session_state.display_name),
                                "_submitted_at":     datetime.now().strftime("%d/%m/%Y %H:%M"),
                                "_submitted_by":     st.session_state.display_name,
                            })
                            added += 1

                        _before_projs_msp = set(get_all_projects(load_entries()))
                        save_entries(entries)
                        auto_assign_new_projects(
                            st.session_state.get("username",""),
                            _before_projs_msp, entries
                        )
                        msg = "MSP import complete:"
                        if updated:          msg += f" **{updated}** auto-matched updated"
                        if manually_updated: msg += f", **{manually_updated}** manually overwritten"
                        if added:            msg += f", **{added}** new activities added"
                        st.success(msg)
                        st.rerun()

            else:
                # ── P6 matching flow — scoped to (activity_id, project) ────
                def _row_key(r: dict) -> tuple:
                    return (r["activity_id"].upper(),
                            get_project_from_wbs(r.get("wbs_id", "")))

                clean     = [r for r in imported_rows if _row_key(r) not in known_ids
                             # known_ids is keyed by activity_id.upper() only (project-filtered)
                             # re-check using full entries for cross-project safety
                             ]

                # Rebuild a full cross-project lookup keyed by (aid, project)
                _all_entries     = load_entries()
                _full_index      = {
                    (e.get("activity_id","").upper(),
                     get_project_from_wbs(e.get("wbs_id",""))): e
                    for e in _all_entries
                }
                clean     = [r for r in imported_rows if _row_key(r) not in _full_index]
                all_match = [r for r in imported_rows if _row_key(r) in _full_index]

                # Split into genuine changes vs identical (skip silently)
                conflicts = [r for r in all_match
                             if not is_exact_duplicate(r, _full_index[_row_key(r)])]
                duplicates_skipped = len(all_match) - len(conflicts)

                msg = (
                    f"Found **{len(imported_rows)}** rows: **{len(clean)}** new, "
                    f"**{len(conflicts)}** conflict{'s' if len(conflicts) != 1 else ''}"
                )
                if duplicates_skipped:
                    msg += f", **{duplicates_skipped}** identical (skipped)"
                st.success(msg)

                # resolutions keyed by (aid, project) tuple to avoid cross-project clashes
                resolutions = {}
                if conflicts:
                    st.divider()
                    st.subheader("⚠️ Conflicts — Activity ID + Project already exists")
                    st.caption("Review each conflict and choose an action before confirming the import.")
                    for row in conflicts:
                        rkey     = _row_key(row)
                        aid      = row["activity_id"].upper()
                        project  = rkey[1]
                        existing = _full_index[rkey]
                        label    = (f"🔁  {row['activity_id']}  [{project}]  —  "
                                    f"{row.get('activity_name', existing.get('activity_name', ''))}")
                        with st.expander(label, expanded=True):
                            fields = [
                                ("Status",          "activity_status",  False),
                                ("Actual Start",     "actual_start",     True),
                                ("Actual Finish",    "actual_finish",    True),
                                ("% Complete",       "pct_complete",     False),
                                ("Remaining (days)", "remaining_dur",    False),
                                ("Predicted Start",  "predicted_start",  True),
                                ("Task Type",        "task_type",        False),
                                ("WBS",              "wbs_id",           False),
                            ]
                            existing_comment_count = len(existing.get("_comments", []))
                            incoming_comment_str   = row.get("comments_import", "").strip()
                            c_cur, c_new = st.columns(2)
                            with c_cur:
                                st.write("**Current (stored)**")
                                for lbl, k, is_date in fields:
                                    val = existing.get(k, "") or ""
                                    st.write(f"- **{lbl}:** {display_dt(val) if is_date else (val or '—')}")
                                st.write(f"- **Comments:** {existing_comment_count} stored comment{'s' if existing_comment_count != 1 else ''}")
                            with c_new:
                                st.write("**Incoming (from file)**")
                                for lbl, k, is_date in fields:
                                    val = row.get(k, "") or ""
                                    st.write(f"- **{lbl}:** {display_dt(val) if is_date else (val or '—')}")
                                st.write(f"- **Comments (user_field_813):** {incoming_comment_str or '—'}")

                            # Use a safe string key for st widgets (no special chars)
                            widget_key = f"{aid}__{project}"
                            choice = st.radio(
                                "Activity data", ["Overwrite with incoming", "Keep current"],
                                key=f"conflict_{widget_key}", horizontal=True,
                            )
                            resolutions[rkey] = "overwrite" if choice == "Overwrite with incoming" else "skip"

                            if incoming_comment_str and existing_comment_count > 0:
                                comment_choice = st.radio(
                                    "Comments",
                                    ["Append imported comments to existing",
                                     "Keep existing only",
                                     "Replace with imported only"],
                                    key=f"comment_conflict_{widget_key}", horizontal=True,
                                )
                                resolutions[("comment",) + rkey] = comment_choice
                            elif incoming_comment_str:
                                resolutions[("comment",) + rkey] = "Append imported comments to existing"
                            else:
                                resolutions[("comment",) + rkey] = "Keep existing only"

                st.divider()
                ow    = sum(1 for k, v in resolutions.items()
                            if isinstance(k, tuple) and k[0] != "comment" and v == "overwrite")
                sk    = sum(1 for k, v in resolutions.items()
                            if isinstance(k, tuple) and k[0] != "comment" and v == "skip")
                parts = [f"{len(clean)} new entries will be added"]
                if ow: parts.append(f"{ow} existing entries will be overwritten")
                if sk: parts.append(f"{sk} conflicts will be skipped")
                st.info("  ·  ".join(parts))

                if st.button("✅  Confirm Import", type="primary"):
                    entries = load_entries()
                    added = overwritten = skipped = 0

                    for row in clean:
                        imp_str = row.pop("comments_import", "") or ""
                        row["_comments"]     = import_string_to_comments(imp_str, st.session_state.display_name)
                        row["_submitted_by"] = st.session_state.display_name
                        entries.append(row)
                        added += 1

                    # Build full (aid, project) index after appending new rows
                    entries_index = {
                        (e.get("activity_id","").upper(),
                         get_project_from_wbs(e.get("wbs_id",""))): idx
                        for idx, e in enumerate(entries)
                    }

                    for row in conflicts:
                        rkey = _row_key(row)
                        if resolutions.get(rkey) == "overwrite":
                            idx = entries_index.get(rkey)
                            if idx is not None:
                                imp_str      = row.pop("comments_import", "") or ""
                                comment_res  = resolutions.get(("comment",) + rkey, "Keep existing only")
                                stored_cmts  = entries[idx].get("_comments", [])
                                imported_cmts = import_string_to_comments(imp_str, st.session_state.display_name)
                                if comment_res == "Append imported comments to existing":
                                    new_only = merge_imported_comments(imported_cmts, stored_cmts)
                                    row["_comments"] = stored_cmts + new_only
                                elif comment_res == "Replace with imported only":
                                    row["_comments"] = imported_cmts
                                else:
                                    row["_comments"] = stored_cmts
                                row["_submitted_by"] = st.session_state.display_name
                                entries[idx] = row
                                overwritten += 1
                        else:
                            skipped += 1

                    save_entries(entries)
                    msg = f"Import complete: **{added}** added"
                    if overwritten: msg += f", **{overwritten}** overwritten"
                    if skipped:     msg += f", **{skipped}** skipped"
                    st.success(msg)
                    st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# TAB: EXPORT TO EXCEL
# ══════════════════════════════════════════════════════════════════════════════

if "export" in tab_index:
    with tab_index["export"]:
        entries = load_entries()
        st.subheader("Export P6-Ready Excel File")
        st.write("The exported file is formatted for direct import into Primavera P6:")
        st.write("- **Row 1:** P6 internal field key names (`task_code`, `act_start_date`, etc.)")
        st.write("- **Row 2:** Column headers · **Sheet name:** `TASK`")
        st.write("- **Date format:** `DD/MM/YYYY HH:MM` stored as proper Excel datetime cells")
        st.write("- **% Complete:** Plain integer")
        st.write("- **Complete Type:** Always Physical as to not overide remaining durations with calulated values from percent complete")
        st.divider()

        _sel_project = st.session_state.get("selected_project", "")
        entries = filter_by_project(entries, _sel_project)

        if not entries:
            st.warning(f"No entries to export for project '{_sel_project}'.")
        else:
            st.info(f"{len(entries)} {'entry' if len(entries) == 1 else 'entries'} ready to export ({_sel_project}).")

            # ── Project name / WBS prefix ─────────────────────────────────
            # Detect the current prefix from the first entry that has one
            _current_prefix = ""
            for _e in entries:
                _wbs = str(_e.get("wbs_id", "") or "")
                if _wbs and "." in _wbs:
                    _first_seg = _wbs.split(".")[0]
                    if not _first_seg.isdigit():
                        _current_prefix = _first_seg
                        break

            rename_wbs = st.checkbox(
                "Update project name in WBS codes",
                value=False,
                key="export_rename_wbs",
                help="Use this when the P6 project name has changed since activities were entered.",
            )

            project_name_out = ""
            if rename_wbs:
                if _current_prefix:
                    st.caption(f"Current prefix detected: **{_current_prefix}**")
                else:
                    st.caption("No non-numeric prefix detected in stored WBS codes.")

                project_name_out = st.text_input(
                    "New P6 project name *",
                    placeholder="e.g. ProjectB",
                    key="export_project_name",
                ).strip()

                st.warning(
                    "⚠️ The project name entered here must exactly match the project name "
                    "in Primavera P6, including capitalisation. If it does not match, "
                    "the import into P6 may fail or assign activities to the wrong WBS."
                )

                if project_name_out and _current_prefix:
                    st.caption(
                        f"Preview: **{_current_prefix}.1.2.3** → "
                        f"**{project_name_out}.1.2.3**"
                    )

            st.divider()
            _ready = not rename_wbs or bool(project_name_out)
            if not _ready:
                st.error("Enter the new project name above before downloading.")

            excel_bytes = build_excel(entries, project_name=project_name_out)
            fname = f"p6_asbuilt_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            st.download_button(
                label="⬇️  Download P6-Ready XLSX",
                data=excel_bytes,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                width='stretch',
                disabled=not _ready,
            )
            st.divider()
            st.write("**How to import into P6:**")
            st.write("1. Open **Primavera P6 Professional**")
            st.write("2. `File` → `Import` → `Spreadsheet (XLSX)` → **Next**")
            st.write("3. Browse to the downloaded file → **Next**")
            st.write("4. Select **Activities** → choose your project → **Finish**")
            st.write("5. Press **F9** to reschedule")

# ══════════════════════════════════════════════════════════════════════════════
# TAB: PHOTO LOG
# ══════════════════════════════════════════════════════════════════════════════

if "photos" in tab_index:
    with tab_index["photos"]:
        ensure_photo_dir()
        entries           = load_entries()
        # Filter activity list to projects the current user can access
        _accessible_photo = get_accessible_projects(
            st.session_state.get("username", ""), get_all_projects(entries)
        )
        known_ids_ordered = [
            e["activity_id"] for e in entries
            if get_project_from_wbs(e.get("wbs_id","")) in _accessible_photo
            or get_project_from_wbs(e.get("wbs_id","")) == "(Unassigned)"
        ]
        # Keyed by (activity_id.upper(), project) to disambiguate same IDs across projects
        known_map = {
            (e["activity_id"].upper(),
             get_project_from_wbs(e.get("wbs_id",""))): e
            for e in entries
        }
        can_upload        = has_permission("photos")

        # ── Cache photos list and all lookups in session_state ─────────────
        # These are only reloaded from disk when a photo is uploaded or
        # deleted (st.rerun with cache cleared). All other interactions
        # — dropdown changes, filter changes, assignment changes — read
        # entirely from session_state, avoiding any disk I/O or list rebuilds.

        if "photo_list" not in st.session_state:
            st.session_state["photo_list"] = load_photos()
        photos = st.session_state["photo_list"]

        if "photo_assignments" not in st.session_state:
            st.session_state["photo_assignments"] = load_assignments()
        assignments = st.session_state["photo_assignments"]

        # Lookups are rebuilt only when assignments change (cheap, in-memory)
        # We store them in session_state so a dropdown change doesn't rebuild them
        # Hash photo_id+activity_id pairs so any change — not just count — triggers rebuild
        assign_sig = hash(tuple((a["photo_id"], a["activity_id"]) for a in assignments))
        if (st.session_state.get("_assign_sig") != assign_sig
                or "photo_to_aids" not in st.session_state):
            _pta, _atp, _pm = {}, {}, {}
            for a in assignments:
                _proj_a  = get_project_from_wbs(a.get("wbs_id",""))
                _aid_key = (a["activity_id"].upper(), _proj_a)
                _pta.setdefault(a["photo_id"], []).append(_aid_key)
                _atp.setdefault(_aid_key, []).append(a["photo_id"])
            for p in photos:
                _pm[p["id"]] = p
            st.session_state["photo_to_aids"] = _pta
            st.session_state["aid_to_pids"]   = _atp
            st.session_state["photo_map"]     = _pm
            st.session_state["_assign_sig"]   = assign_sig

        photo_to_aids = st.session_state["photo_to_aids"]
        aid_to_pids   = st.session_state["aid_to_pids"]
        photo_map     = st.session_state["photo_map"]

        # ── Step 1: Upload ─────────────────────────────────────────────────
        if can_upload:
            st.subheader("Step 1 — Upload Photo")
            st.caption("Upload the image and set its date and comment. You will assign it to activities in Step 2.")

            up_col1, up_col2 = st.columns(2)
            with up_col1:
                photo_date = st.date_input(
                    "Date of Photo *",
                    value=date.today(),
                    format="DD/MM/YYYY",
                    key="photo_upload_date",
                )
            with up_col2:
                comment = st.text_input(
                    "Comment",
                    placeholder="e.g. North elevation — formwork complete (max 100 chars)",
                    max_chars=100,
                    key="photo_upload_comment",
                )

            uploaded_file = st.file_uploader(
                "Choose image *",
                type=["jpg", "jpeg", "png", "webp", "gif"],
                key="photo_upload_file",
            )
            inject_camera_uploader("photo_upload_file")

            if uploaded_file:
                # Read bytes once — reused for both preview and upload
                _file_bytes = uploaded_file.read()
                if _PILLOW:
                    _prev = ImageOps.exif_transpose(Image.open(io.BytesIO(_file_bytes)))
                    st.image(_prev, caption="Preview", width=400)
                else:
                    st.image(io.BytesIO(_file_bytes), caption="Preview", width=400)

            if st.button("📤  Upload Photo", type="primary", disabled=not uploaded_file):
                record = upload_photo(
                    photo_date    = photo_date,
                    comment       = comment,
                    file_bytes    = _file_bytes,
                    original_name = uploaded_file.name,
                    uploaded_by   = st.session_state.display_name,
                )
                st.success(f"✅ Photo uploaded — now assign it to activities in Step 2.")
                # Store the new photo id so Step 2 pre-selects it
                st.session_state["last_uploaded_photo_id"] = record["id"]
                load_image_bytes.clear()
                build_photo_backup.clear()
                for _k in ("photo_list","photo_map","photo_to_aids","aid_to_pids","_assign_sig"):
                    st.session_state.pop(_k, None)
                st.rerun()

            st.divider()

            # ── Step 2: Assign ─────────────────────────────────────────────
            st.subheader("Step 2 — Assign Photo to Activities")
            st.caption("Select a photo from the library and assign it to one or more activities.")

            if not photos:
                st.info("No photos uploaded yet — upload one above first.")
            elif not known_ids_ordered:
                st.warning("No activities found — submit some entries first.")
            else:
                # Photo selector — default to most recently uploaded
                last_id   = st.session_state.get("last_uploaded_photo_id", photos[-1]["id"])
                photo_ids = [p["id"] for p in sorted(photos, key=lambda p: p["uploaded_at"], reverse=True)]

                def photo_label(pid: str) -> str:
                    p = photo_map.get(pid, {})
                    try:
                        dt_str = datetime.strptime(p.get("photo_date",""), "%Y-%m-%d").strftime("%d/%m/%Y")
                    except ValueError:
                        dt_str = p.get("photo_date", "")
                    n = len(photo_to_aids.get(pid, []))
                    suffix = f" — {n} assignment{'s' if n != 1 else ''}"
                    comment_preview = (p.get("comment","")[:30] + "…") if p.get("comment","") else "no comment"
                    return f"{dt_str}  ·  {comment_preview}{suffix}"

                default_idx = photo_ids.index(last_id) if last_id in photo_ids else 0
                selected_photo_id = st.selectbox(
                    "Select photo",
                    options=photo_ids,
                    index=default_idx,
                    format_func=photo_label,
                    key="assign_photo_select",
                )

                selected_photo = photo_map.get(selected_photo_id, {})

                # Show thumbnail only — avoids loading full image in Step 2
                _thumb = selected_photo.get("thumb", "")
                _thumb_bytes = load_image_bytes(_thumb) if _thumb else load_image_bytes(selected_photo.get("filename",""))
                if _thumb_bytes:
                    st.image(_thumb_bytes, width=200)

                # Current assignments for this photo
                # photo_to_aids stores (aid.upper(), project) tuples
                current_aids = set(photo_to_aids.get(selected_photo_id, []))

                # Multi-select for activities
                assign_ids = st.multiselect(
                    "Assign to activities",
                    options=known_ids_ordered,
                    default=[aid for aid in known_ids_ordered
                             if any(t[0] == aid.upper() for t in current_aids)],
                    format_func=lambda aid: (get_project_from_wbs(
                        next((e.get("wbs_id","") for e in entries
                              if e["activity_id"].upper() == aid.upper()), "")
                    ))+" "+(
                        known_map.get(
                            (aid.upper(), get_project_from_wbs(
                                next((e.get("wbs_id","") for e in entries
                                      if e["activity_id"].upper() == aid.upper()), "")
                            )), {}
                        ).get("activity_name", aid)
                    ),
                    key="assign_activity_select",
                )

                if st.button("💾  Save Assignments", type="primary"):
                    # Add new assignments
                    new_aids = [aid for aid in assign_ids
                               if not any(t[0] == aid.upper() for t in current_aids)]
                    if new_aids:
                        assign_photo(selected_photo_id, new_aids,
                                     st.session_state.display_name, entries)
                    removed_aids = [aid for aid in known_ids_ordered
                                    if any(t[0] == aid.upper() for t in current_aids)
                                    and aid not in assign_ids]
                    for aid in removed_aids:
                        _rm_wbs = next((e.get("wbs_id","") for e in entries
                                        if e["activity_id"].upper() == aid.upper()), "")
                        unassign_photo(selected_photo_id, aid, _rm_wbs)
                    # session_state updated inside assign/unassign — no rerun needed
                    st.success(
                        f"Assignments saved — "
                        f"{len(new_aids)} added, {len(removed_aids)} removed."
                    )

            st.divider()

        # ── Gallery (all roles) ────────────────────────────────────────────
        st.subheader("Photo Gallery")

        # Backup — available to all roles from any connected device
        with st.expander("💾  Backup Photo Library"):
            st.caption(
                "Downloads a ZIP of all image files, photo metadata, and "
                "assignment records. Use this to keep an off-device copy "
                "or to migrate the library to another server."
            )
            _backup_bytes = build_photo_backup()
            _backup_fname = f"p6_photo_backup_{datetime.now().strftime('%Y%m%d_%H%M')}.zip"
            st.download_button(
                label="⬇️  Download Photo Backup",
                data=_backup_bytes,
                file_name=_backup_fname,
                mime="application/zip",
                width='stretch',
            )
            _n_photos = len(photos)
            _size_mb  = round(len(_backup_bytes) / 1_048_576, 1)
            st.caption(f"{_n_photos} photo{'s' if _n_photos != 1 else ''} · {_size_mb} MB")


        # Restore — admin only
        if has_permission("export"):
            with st.expander("📂  Restore from Backup  (Admin only)"):
                st.warning(
                    "⚠️ Restoring will overwrite the current photo metadata and replace "
                    "any image files that exist in the backup. "
                    "Image files on the server that are **not** in the backup are kept. "
                    "**Download a fresh backup first if you want to preserve the current state.**"
                )
                restore_file = st.file_uploader(
                    "Upload backup ZIP",
                    type=["zip"],
                    key="photo_restore_upload",
                )
                if restore_file:
                    zip_bytes = restore_file.read()
                    # Preview contents
                    try:
                        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as _zf:
                            _names    = _zf.namelist()
                            _n_imgs   = sum(1 for n in _names
                                            if n.startswith(f"{PHOTO_DIR.name}/")
                                            and not n.endswith("/"))
                            _has_log  = PHOTO_FILE.name  in _names
                            _has_asgn = ASSIGN_FILE.name in _names
                        st.write("**Backup contents:**")
                        st.write(f"- Photo log JSON: {'✓' if _has_log  else '✗ missing'}")
                        st.write(f"- Assignments JSON: {'✓' if _has_asgn else '✗ missing'}")
                        st.write(f"- Image files: {_n_imgs}")
                    except Exception as _e:
                        st.error(f"Could not read ZIP: {_e}")
                        zip_bytes = None

                    if zip_bytes and st.button(
                        "📂  Confirm Restore", type="primary", key="photo_restore_confirm"
                    ):
                        with st.spinner("Restoring…"):
                            n_photos, n_imgs, restore_warns = restore_photo_backup(zip_bytes)
                        for w in restore_warns:
                            st.warning(w)
                        # Invalidate all caches so restored data loads fresh
                        load_image_bytes.clear()
                        build_photo_backup.clear()
                        for _k in ("photo_list", "photo_map", "photo_assignments",
                                   "photo_to_aids", "aid_to_pids", "_assign_sig"):
                            st.session_state.pop(_k, None)
                        st.success(
                            f"Restore complete — {n_photos} photo records, "
                            f"{n_imgs} image files restored."
                        )
                        st.rerun()

        st.divider()

        if not photos:
            st.info("No photos uploaded yet.")
        else:
            # Filter controls
            _sel_project = st.session_state.get("selected_project", "")
            _accessible_photo_projs = get_accessible_projects(
                st.session_state.get("username", ""), get_all_projects(entries)
            )
            f_col0, f_col1, f_col2 = st.columns([2, 2, 3])
            with f_col0:
                _photo_proj = st.selectbox(
                    "Filter by Project",
                    options=_accessible_photo_projs,
                    index=_accessible_photo_projs.index(_sel_project)
                          if _sel_project in _accessible_photo_projs else 0,
                    key="photo_filter_project",
                )
            with f_col1:
                # Only show activity IDs that have at least one assignment,
                # filtered to the selected project
                _project_aids = {
                    e["activity_id"] for e in filter_by_project(entries, _photo_proj)
                } if _photo_proj else None
                assigned_aids = sorted({
                    a["activity_id"] for a in assignments
                    if (_project_aids is None or a["activity_id"] in _project_aids)
                })
                filter_id = st.selectbox(
                    "Filter by Activity",
                    options=["— All —"] + assigned_aids,
                    key="photo_filter_id",
                )
            with f_col2:
                filter_text = st.text_input(
                    "Search comments",
                    placeholder="Type to search…",
                    key="photo_filter_text",
                ).strip().lower()

            # Apply filters
            if _photo_proj and _project_aids is not None:
                # Find all photo IDs assigned to any activity in this project
                _proj_pids = {
                    a["photo_id"] for a in assignments
                    if a["activity_id"] in _project_aids
                    and get_project_from_wbs(a.get("wbs_id","")) == _photo_proj
                }
                filtered = [p for p in photos if p["id"] in _proj_pids]
            else:
                filtered = list(photos)

            if filter_id != "— All —":
                _filt_entry = next(
                    (e for e in entries if e["activity_id"].upper() == filter_id.upper()), {}
                )
                _filt_proj   = get_project_from_wbs(_filt_entry.get("wbs_id",""))
                visible_pids = set(aid_to_pids.get((filter_id.upper(), _filt_proj), []))
                filtered = [p for p in filtered if p["id"] in visible_pids]

            if filter_text:
                filtered = [p for p in filtered
                            if filter_text in p.get("comment", "").lower()
                            or any(filter_text in (aid[0] if isinstance(aid, tuple) else aid).lower()
                                   for aid in photo_to_aids.get(p["id"], []))]

            filtered = sorted(filtered, key=lambda p: p["photo_date"], reverse=True)

            if not filtered:
                st.info("No photos match the current filter.")
            else:
                PHOTO_PAGE_SIZE = 18  # 6 rows of 3 — adjust as needed
                total_photo_pages = max(1, (len(filtered) + PHOTO_PAGE_SIZE - 1) // PHOTO_PAGE_SIZE)
                photo_page = st.number_input(
                    f"Page (1 – {total_photo_pages})",
                    min_value=1, max_value=total_photo_pages, value=1, step=1,
                    key="photo_gallery_page",
                ) - 1  # zero-based
                page_start = photo_page * PHOTO_PAGE_SIZE
                page_end   = page_start + PHOTO_PAGE_SIZE
                page_photos = filtered[page_start:page_end]

                st.caption(
                    f"Showing {page_start + 1}–{min(page_end, len(filtered))} "
                    f"of {len(filtered)} photo{'s' if len(filtered) != 1 else ''}"
                )
                st.divider()

                COLS = 3
                for row_start in range(0, len(page_photos), COLS):
                    cols = st.columns(COLS)
                    for col_idx, photo in enumerate(page_photos[row_start:row_start + COLS]):
                        with cols[col_idx]:
                            with st.container(border=True):
                                # Use thumbnail for fast loading
                                _thumb = photo.get("thumb", "")
                                _tb    = load_image_bytes(_thumb) if _thumb else load_image_bytes(photo.get("filename",""))
                                if _tb:
                                    st.image(_tb, width=350)
                                else:
                                    st.caption("Image missing")

                                try:
                                    dt_str = datetime.strptime(
                                        photo["photo_date"], "%Y-%m-%d"
                                    ).strftime("%d/%m/%Y")
                                except ValueError:
                                    dt_str = photo["photo_date"]

                                st.caption(f"📅 {dt_str}")
                                if photo.get("comment"):
                                    st.caption(photo["comment"])

                                # Assigned activities
                                aids = photo_to_aids.get(photo["id"], [])  # list of (aid, project) tuples
                                if aids:
                                    for aid_key in aids:
                                        act      = known_map.get(aid_key, {})
                                        aid_str  = aid_key[0] if isinstance(aid_key, tuple) else aid_key
                                        proj_str = aid_key[1] if isinstance(aid_key, tuple) else ""
                                        st.caption(
                                            f"📌 {aid_str}"
                                            + (f" [{proj_str}]" if proj_str and proj_str != "(Unassigned)" else "")
                                            + f" — {act.get('activity_name','')} ({act.get('activity_status','')})"
                                        )
                                else:
                                    st.caption("Not assigned")

                                st.caption(f"By {photo.get('uploaded_by','—')}  ·  {photo['uploaded_at']}")

                                if can_upload:
                                    with st.expander("✏️ Assign / Remove / Delete"):
                                        _assigned_aids = {t[0] for t in aids} if aids and isinstance(aids[0], tuple) else {a.upper() for a in aids}
                                        new_assign = st.multiselect(
                                            "Assign to:",
                                            options=[a for a in known_ids_ordered
                                                     if a.upper() not in _assigned_aids],
                                            format_func=lambda a: (
                                                known_map.get(
                                                    (a.upper(), get_project_from_wbs(
                                                        next((e.get("wbs_id","") for e in entries
                                                              if e["activity_id"].upper() == a.upper()), "")
                                                    )), {}
                                                ).get("activity_name", a)
                                            ),
                                            key=f"gallery_assign_{photo['id']}",
                                        )
                                        if st.button("＋ Assign", key=f"gallery_assign_btn_{photo['id']}",
                                                     disabled=not new_assign):
                                            assign_photo(photo["id"], new_assign,
                                                         st.session_state.display_name, entries)
                                            st.session_state["_assign_sig"] = None
                                            st.toast(f"Assigned to {len(new_assign)} activity/activities.")

                                        if aids:
                                            for aid_key in aids:
                                                aid_str  = aid_key[0] if isinstance(aid_key, tuple) else aid_key
                                                aid_wbs  = next((e.get("wbs_id","") for e in entries
                                                                 if e["activity_id"].upper() == aid_str), "")
                                                proj_str = aid_key[1] if isinstance(aid_key, tuple) and aid_key[1] not in ("","(Unassigned)") else ""
                                                btn_lbl  = f"Remove from {aid_str}" + (f" [{proj_str}]" if proj_str else "")
                                                if st.button(btn_lbl, key=f"unassign_{photo['id']}_{aid_str}_{proj_str}"):
                                                    unassign_photo(photo["id"], aid_str, aid_wbs)
                                                    st.session_state["_assign_sig"] = None
                                                    st.toast(f"Removed from {aid_str}.")

                                        if st.button("🗑 Delete permanently",
                                                     key=f"photo_del_{photo['id']}", type="primary"):
                                            delete_photo_file(photo["id"])
                                            load_image_bytes.clear()
                                            for _k in ("photo_list","photo_map","photo_assignments",
                                                       "photo_to_aids","aid_to_pids","_assign_sig"):
                                                st.session_state.pop(_k, None)
                                            st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# TAB: SETTINGS  (Admin only)
# ══════════════════════════════════════════════════════════════════════════════

if "settings" in tab_index:
    with tab_index["settings"]:
        st.subheader("⚙️ Settings")
        st.caption("Admin-only configuration options.")

        # ── Notifications inbox ───────────────────────────────────────────
        st.markdown("### 🔔 Notifications")
        _all_notifs_raw = load_notifications()
        _username       = st.session_state.get("username","")
        # Only show notifications addressed to this user
        _all_notifs = [
            n for n in _all_notifs_raw
            if not n.get("recipients") or _username in n.get("recipients", [])
        ]

        if not _all_notifs:
            st.info("No notifications yet.")
        else:
            _unread_ids = {n["id"] for n in _all_notifs if _username not in n.get("read_by",[])}
            st.caption(f"{len(_all_notifs)} total  ·  {len(_unread_ids)} unread")

            # Table header
            _th1, _th2, _th3, _th4, _th5 = st.columns([3, 2, 2, 1, 1])
            _th1.caption("**Title**")
            _th2.caption("**From**")
            _th3.caption("**Date**")
            _th4.caption("**Read**")
            _th5.caption("**Delete**")
            st.divider()

            import pandas as _pd
            for notif in reversed(_all_notifs):
                _is_unread  = notif["id"] in _unread_ids
                _notif_status = notif.get("status", "info")

                # ── Pending approval — full review UI ─────────────────────────
                if _notif_status == "pending" and has_permission("settings"):
                    with st.container(border=True):
                        _pa1, _pa2 = st.columns([5, 1])
                        with _pa1:
                            st.write(f"🟡 **{notif['title']}**")
                            st.caption(
                                f"From: {notif.get('created_by','—')}  ·  "
                                f"{notif.get('created_at','—')}  ·  "
                                f"Project: {notif.get('project','—')}"
                            )
                        with _pa2:
                            if st.button("🗑", key=f"notif_del_{notif['id']}",
                                         help="Delete"):
                                delete_notification(notif["id"])
                                st.rerun()

                        if notif.get("body"):
                            st.write(notif["body"])

                        # Editable activity table
                        _edits = notif.get("edits", {})
                        _approved_edits = {}
                        if notif.get("rows"):
                            _entries_for_approval = load_entries()
                            _rpt_dt_apr = get_report_date(notif.get("project",""))

                            # ── Before / After Gantt toggle ───────────────────
                            if st.toggle("📊 Show Before / After Gantt",
                                         key=f"apr_gantt_{notif['id']}", value=False):
                                _before_acts = []
                                _after_acts  = []
                                for _row in notif["rows"]:
                                    _aid  = _row.get("Activity ID","")
                                    _prop = _edits.get(_aid, {})
                                    _cur  = next((e for e in _entries_for_approval
                                                  if e.get("activity_id") == _aid), {})
                                    _before_acts.append(_cur)
                                    # After = current with proposed edits merged
                                    _after_acts.append({**_cur, **_prop})

                                _svg_before = build_gantt_svg(
                                    _before_acts, _rpt_dt_apr, title="Before"
                                )
                                _svg_after  = build_gantt_svg(
                                    _after_acts,  _rpt_dt_apr, title="After"
                                )
                                _gc1, _gc2 = st.columns(2)
                                with _gc1:
                                    if _svg_before:
                                        st.markdown(_svg_before, unsafe_allow_html=True)
                                    else:
                                        st.caption("No chart data available.")
                                with _gc2:
                                    if _svg_after:
                                        st.markdown(_svg_after, unsafe_allow_html=True)
                                    else:
                                        st.caption("No chart data available.")

                            st.write("**Proposed changes — edit cells before approving:**")

                            # Build table rows: one per changed activity
                            _tbl_rows = []
                            for _row in notif["rows"]:
                                _aid  = _row.get("Activity ID","")
                                _prop = _edits.get(_aid, {})
                                _tbl_rows.append({
                                    "Activity ID":       _aid,
                                    "Activity Name":     _row.get("Activity Name",""),
                                    "Status":            _prop.get("activity_status",""),
                                    "% Complete":        int(_prop.get("pct_complete") or 0),
                                    "Remaining (days)":  str(_prop.get("remaining_dur","") or ""),
                                    "Actual Start":      display_dt(_prop.get("actual_start","")),
                                    "Actual Finish":     display_dt(_prop.get("actual_finish","")),
                                })

                            _edited_df = st.data_editor(
                                _pd.DataFrame(_tbl_rows),
                                key=f"apr_table_{notif['id']}",
                                use_container_width=True,
                                hide_index=True,
                                column_config={
                                    "Activity ID":      st.column_config.TextColumn(disabled=True),
                                    "Activity Name":    st.column_config.TextColumn(disabled=True),
                                    "Status":           st.column_config.SelectboxColumn(
                                                            options=STATUS_OPTIONS
                                                        ),
                                    "% Complete":       st.column_config.NumberColumn(
                                                            min_value=0, max_value=100, step=5
                                                        ),
                                    "Remaining (days)": st.column_config.TextColumn(),
                                    "Actual Start":     st.column_config.TextColumn(disabled=True),
                                    "Actual Finish":    st.column_config.TextColumn(disabled=True),
                                },
                            )

                            # Rebuild approved_edits from the (possibly edited) table
                            for _, _erow in _edited_df.iterrows():
                                _aid  = _erow["Activity ID"]
                                _prop = _edits.get(_aid, {})
                                _approved_edits[_aid] = {
                                    **_prop,
                                    "activity_status": _erow["Status"],
                                    "pct_complete":    str(int(_erow["% Complete"])),
                                    "remaining_dur":   str(_erow["Remaining (days)"]),
                                }

                        # Approve / Reject buttons
                        st.divider()
                        _ab1, _ab2, _ab3 = st.columns([2, 2, 3])
                        with _ab1:
                            if st.button("✅ Approve & Apply",
                                         key=f"apr_approve_{notif['id']}",
                                         type="primary", use_container_width=True):
                                # Write approved edits to entries
                                _all_e = load_entries()
                                _idx_map = {
                                    (e.get("activity_id","").upper(),
                                     get_project_from_wbs(e.get("wbs_id",""))): i
                                    for i, e in enumerate(_all_e)
                                }
                                for _aid, _edata in _approved_edits.items():
                                    _proj = get_project_from_wbs(
                                        next((e.get("wbs_id","") for e in _all_e
                                              if e.get("activity_id") == _aid), "")
                                    )
                                    _idx = _idx_map.get((_aid.upper(), _proj))
                                    if _idx is not None:
                                        _all_e[_idx] = {**_all_e[_idx], **_edata}
                                save_entries(_all_e)
                                set_last_walk_date(notif.get("project",""), date.today())
                                update_notification(notif["id"],
                                    status="approved",
                                    read_by=list({*notif.get("read_by",[]), _username}),
                                )
                                st.success("Walk approved and changes applied.")
                                st.rerun()

                        with _ab2:
                            if st.button("❌ Reject",
                                         key=f"apr_reject_{notif['id']}",
                                         use_container_width=True):
                                st.session_state[f"rejecting_{notif['id']}"] = True

                        if st.session_state.get(f"rejecting_{notif['id']}"):
                            with _ab3:
                                _rej_note = st.text_input(
                                    "Rejection note",
                                    placeholder="Explain why the walk was rejected…",
                                    key=f"apr_rej_note_{notif['id']}",
                                )
                                if st.button("Send rejection",
                                             key=f"apr_rej_send_{notif['id']}",
                                             disabled=not _rej_note):
                                    # Notify the engineer
                                    _eng_username = next(
                                        (u for u, d in USERS.items()
                                         if d["name"] == notif.get("created_by")), None
                                    )
                                    create_notification(
                                        created_by  = _username,
                                        project     = notif.get("project",""),
                                        title       = f"Walk Rejected — {notif.get('project','')} — {notif.get('created_at','').split(' ')[0]}",
                                        body        = (
                                            "Your site walk was rejected by "
                                            + f"**{st.session_state.display_name}**.\n\n"
                                            + f"**Note:** {_rej_note}"
                                        ),
                                        recipients  = [_eng_username] if _eng_username else [],
                                        status      = "rejected",
                                    )
                                    update_notification(notif["id"],
                                        status="rejected",
                                        rejection_note=_rej_note,
                                        read_by=list({*notif.get("read_by",[]), _username}),
                                    )
                                    st.session_state.pop(f"rejecting_{notif['id']}", None)
                                    st.success("Walk rejected and engineer notified.")
                                    st.rerun()

                    st.divider()
                    continue  # skip standard row rendering for pending notifs

                # ── Standard notification row ─────────────────────────────────
                _status_icon = {"approved": "✅ ", "rejected": "❌ ", "info": ""}.get(
                    _notif_status, ""
                )
                _tc1, _tc2, _tc3, _tc4, _tc5 = st.columns([3, 2, 2, 1, 1])

                with _tc1:
                    prefix = "🔵 " if _is_unread else ""
                    st.write(f"{prefix}{_status_icon}{notif['title']}")
                    if notif.get("project"):
                        st.caption(notif["project"])
                    with st.expander("Details"):
                        if notif.get("body"):
                            st.write(notif["body"])
                        if notif.get("rejection_note"):
                            st.error(f"Rejection note: {notif['rejection_note']}")
                        if notif.get("rows"):
                            st.dataframe(
                                _pd.DataFrame(notif["rows"]),
                                use_container_width=True,
                                hide_index=True,
                            )
                        # Attachment download buttons
                        _att_list = notif.get("edits", {}).get("attachments", [])
                        if _att_list:
                            st.write("**Attachments:**")
                            for _att in _att_list:
                                _att_path = PHOTO_DIR / _att["filename"]
                                if _att_path.exists():
                                    _att_bytes = _att_path.read_bytes()
                                    _mime = (
                                        "application/pdf"
                                        if _att["filename"].endswith(".pdf")
                                        else "image/jpeg"
                                    )
                                    st.download_button(
                                        label=f"⬇️ {_att['original_name']} ({_att['size_kb']} KB)",
                                        data=_att_bytes,
                                        file_name=_att["original_name"],
                                        mime=_mime,
                                        key=f"att_dl_{notif['id']}_{_att['filename']}",
                                    )
                                else:
                                    st.caption(f"⚠️ {_att['original_name']} — file not found")

                _tc2.write(notif.get("created_by", "—"))
                _tc3.write(notif.get("created_at", "—"))

                with _tc4:
                    if _is_unread:
                        if st.button("✓", key=f"notif_read_{notif['id']}",
                                     help="Mark as read"):
                            mark_notification_read(notif["id"], _username)
                            st.rerun()
                    else:
                        st.write("✅")

                with _tc5:
                    if st.button("🗑", key=f"notif_del_{notif['id']}",
                                 help="Delete"):
                        delete_notification(notif["id"])
                        st.rerun()

                st.divider()

        st.divider()

        # ── Rename Project ────────────────────────────────────────────────

        _settings_entries = load_entries()
        _settings_projects = get_all_projects(_settings_entries)
        _named_projects = [p for p in _settings_projects if p != "(Unassigned)"]

        # ── Tab Order (developer only) ───────────────────────────────────
        if has_permission("tab_order"):
            st.markdown("### 🔀 Tab Order")
            st.write(
                "Reorder tabs using the arrows. The first tab in the list is "
                "shown first. Changes take effect on next page load."
            )
            _cur_order = load_tab_order()
            _all_perms = [perm for _, perm in TAB_DEFS]
            _display_order = (_cur_order + [p for p in _all_perms
                                            if p not in _cur_order]
                              if _cur_order else list(_all_perms))
            _perm_to_label = {perm: lbl for lbl, perm in TAB_DEFS}

            for _oi, _perm in enumerate(_display_order):
                _oc1, _oc2, _oc3, _oc4 = st.columns([5, 1, 1, 1])
                _oc1.write(f"{_oi+1}. {_perm_to_label.get(_perm, _perm)}")
                with _oc2:
                    if _oi > 0 and st.button("↑", key=f"tab_up_{_perm}",
                                              use_container_width=True):
                        _display_order.insert(_oi-1, _display_order.pop(_oi))
                        save_tab_order(_display_order)
                        st.rerun()
                with _oc3:
                    if _oi < len(_display_order)-1 and st.button(
                            "↓", key=f"tab_dn_{_perm}", use_container_width=True):
                        _display_order.insert(_oi+1, _display_order.pop(_oi))
                        save_tab_order(_display_order)
                        st.rerun()
                with _oc4:
                    if st.button("⊙", key=f"tab_first_{_perm}",
                                 use_container_width=True, help="Move to first"):
                        _display_order.insert(0, _display_order.pop(_oi))
                        save_tab_order(_display_order)
                        st.rerun()

            if st.button("↺  Reset to default order", key="tab_order_reset"):
                save_tab_order([])
                st.rerun()
            st.divider()

        # ── Tab Visibility ────────────────────────────────────────────────
        if has_permission("tab_visibility"):
          st.markdown("### Tab Visibility")
          st.write(
              "Control which tabs are visible per role and per user. "
              "Role settings apply to all users of that role. "
              "User overrides take precedence. "
              "The **Developer** role always sees all tabs."
          )
        st.caption("Tabs can only be hidden here — not granted beyond what the role's permissions allow.")

        _tv = load_tab_visibility()
        _tv_changed = False

        _TAB_LABELS = {
            "sitewalk": " Site Walk",
            "view":     "📋 View All Entries",
            "submit":   "📝 Submit / Update",
            "import":   "📤 Import from Excel",
            "export":   "📥 Export to Excel",
            "photos":   "📸 Photo Log",
            "settings": "⚙️ Settings",
        }
        _PERM_ORDER = ["view","submit","import","export","photos","settings","sitewalk"]

        # Role defaults
        with st.expander("Role defaults", expanded=False):
            st.caption("Hide tabs for an entire role. Developer always sees all.")
            for _role in ("viewer","engineer","admin"):
                st.write(f"**{ROLE_LABEL.get(_role,_role)}**")
                _role_perms = PERMISSIONS.get(_role, set())
                _role_vis   = _tv.get("roles",{}).get(_role,{})
                _r_tabs = [p for p in _PERM_ORDER if p in _role_perms]
                _rcols  = st.columns(min(len(_r_tabs), 4))
                for _ci, _perm in enumerate(_r_tabs):
                    with _rcols[_ci % 4]:
                        _cur = _role_vis.get(_perm, True)
                        _new = st.checkbox(
                            _TAB_LABELS.get(_perm, _perm), value=_cur,
                            key=f"tv_role_{_role}_{_perm}",
                        )
                        if _new != _cur:
                            _tv.setdefault("roles",{}).setdefault(_role,{})[_perm] = _new
                            _tv_changed = True

        # User overrides
        with st.expander("User overrides", expanded=False):
            st.caption("Override tab visibility for a specific user, overriding the role default.")
            _override_user = st.selectbox(
                "Select user",
                options=list(USERS.keys()),
                format_func=lambda u: (
                    f"{u}  —  {USERS[u]['name']} "
                    f"({ROLE_LABEL.get(USERS[u]['role'], USERS[u]['role'])})"
                ),
                key="tv_user_select",
            )
            _u_role  = USERS[_override_user]["role"]
            _u_perms = PERMISSIONS.get(_u_role, set())
            _u_vis   = _tv.get("users",{}).get(_override_user,{})
            st.write(f"**{USERS[_override_user]['name']}** — {ROLE_LABEL.get(_u_role,_u_role)}")

            _perm_list = [p for p in _PERM_ORDER if p in _u_perms]
            _h1, _h2, _h3 = st.columns([3,1,1])
            _h1.caption("Tab"); _h2.caption("Role default"); _h3.caption("Override")

            for _perm in _perm_list:
                _c1, _c2, _c3 = st.columns([3,1,1])
                _role_default = _tv.get("roles",{}).get(_u_role,{}).get(_perm, True)
                _has_override = _perm in _u_vis
                _c1.write(_TAB_LABELS.get(_perm, _perm))
                _c2.write("✅" if _role_default else "❌")
                _use_override = _c3.checkbox(
                    "Override", value=_has_override,
                    key=f"tv_uo_{_override_user}_{_perm}",
                    label_visibility="collapsed",
                )
                if _use_override:
                    _ov = st.checkbox(
                        f"Show tab",
                        value=_u_vis.get(_perm, _role_default),
                        key=f"tv_uv_{_override_user}_{_perm}",
                    )
                    if _ov != _u_vis.get(_perm) or not _has_override:
                        _tv.setdefault("users",{}).setdefault(_override_user,{})[_perm] = _ov
                        _tv_changed = True
                elif _has_override:
                    _tv.get("users",{}).get(_override_user,{}).pop(_perm, None)
                    _tv_changed = True

        if _tv_changed:
            save_tab_visibility(_tv)
            st.success("Tab visibility saved.")

        st.divider()

        # ── Report Dates overview ─────────────────────────────────────────
        st.markdown("### Report Dates")
        st.write(
            "Report dates are set per-project in the sidebar. "
            "This table shows the current state across all projects."
        )
        _settings_proj_settings = load_project_settings()
        _all_named = [p for p in get_all_projects(load_entries()) if p != "(Unassigned)"]
        if not _all_named:
            st.info("No named projects yet.")
        else:
            for _proj in _all_named:
                _rd = _settings_proj_settings.get(_proj, {}).get("report_date")
                _rd_str = datetime.strptime(_rd, "%Y-%m-%d").strftime("%d/%m/%Y") if _rd else "— not set —"
                st.write(f"- **{_proj}:** {_rd_str}")

        st.divider()
        # Project Access — admins and developers can manage projects they have access to
        if has_permission("manage_users"):
            st.markdown("### Project Access")
            st.write(
                "Control which users can access each project. "
                "If no users are selected, **all users** can access it. "
                "You can only manage projects you have access to."
            )

            _access_proj_settings = load_project_settings()
            _all_usernames = list(USERS.keys())
            _username_cur  = st.session_state.get("username", "")

            # Only show projects this user has access to
            _manageable_projs = [
                p for p in _all_named
                if user_can_access_project(_username_cur, p)
            ]

            if not _manageable_projs:
                st.info("No projects available to manage.")
            else:
                for _proj in _manageable_projs:
                    with st.expander(f"📁  {_proj}", expanded=False):
                        _current_allowed = _access_proj_settings.get(_proj, {}).get("allowed_users", [])

                        # Show auto-assigned users clearly
                        if _current_allowed:
                            st.caption(
                                f"Currently restricted to: "
                                f"{', '.join(USERS[u]['name'] if u in USERS else u for u in _current_allowed)}"
                            )
                        else:
                            st.caption("Currently open to all users.")

                        _new_allowed = st.multiselect(
                            "Allowed users (empty = all users)",
                            options=_all_usernames,
                            default=[u for u in _current_allowed if u in _all_usernames],
                            format_func=lambda u: (
                                f"{u}  —  {USERS[u]['name']} · "
                                f"{ROLE_LABEL.get(USERS[u]['role'], USERS[u]['role'])}"
                                + (" (auto-assigned)" if u == _username_cur
                                   and u in _current_allowed else "")
                            ),
                            key=f"access_{_proj}",
                        )
                        # Validate at least one manage_users user is in the selection
                        _has_manager = any(
                            has_permission.__func__ if hasattr(has_permission, '__func__') else
                            (lambda u: "manage_users" in PERMISSIONS.get(
                                USERS.get(u, {}).get("role",""), set()))(u)
                            for u in _new_allowed
                        ) if _new_allowed else True  # empty = all users, so managers included

                        _has_manager = (not _new_allowed) or any(
                            "manage_users" in PERMISSIONS.get(
                                USERS.get(u, {}).get("role",""), set()
                            )
                            for u in _new_allowed
                        )

                        if _new_allowed and not _has_manager:
                            st.error(
                                "⚠️ The selection must include at least one user with "
                                "project management permissions (Admin or Developer). "
                                "Please add one before saving."
                            )
                            _disabled=True
                        else:
                            _disabled=False
                        if st.button("💾  Save access", key=f"access_save_{_proj}",
                                     disabled=(_disabled)):
                            set_allowed_users(_proj, _new_allowed)
                            if _new_allowed:
                                st.success(
                                    f"Access restricted to: "
                                    f"{', '.join(USERS[u]['name'] if u in USERS else u for u in _new_allowed)}"
                                )
                            else:
                                st.success("Access open to all users.")

        st.divider()
        st.markdown("### Rename Project")
        st.write(
            "Renames a project by rewriting the WBS prefix on every activity "
            "that belongs to it. The new name must exactly match the project "
            "name in Primavera P6 if you intend to export and re-import."
        )

        if not _named_projects:
            st.info("No named projects found. Projects are identified by the "
                    "non-numeric prefix of WBS codes (e.g. 'ProjectA' in 'ProjectA.1.2.3').")
        else:
            rn_col1, rn_col2 = st.columns(2)
            with rn_col1:
                old_proj = st.selectbox(
                    "Project to rename",
                    options=_named_projects,
                    key="settings_rename_old",
                )
            with rn_col2:
                new_proj = st.text_input(
                    "New project name",
                    placeholder="e.g. ProjectB",
                    key="settings_rename_new",
                ).strip()

            # Count affected entries
            _affected = [e for e in _settings_entries
                         if get_project_from_wbs(e.get("wbs_id","")) == old_proj]

            if new_proj:
                if new_proj == old_proj:
                    st.warning("New name is the same as the current name.")
                elif new_proj in _named_projects:
                    st.error(
                        f"A project named **{new_proj}** already exists. "
                        "Renaming into an existing project would merge them — "
                        "if that is intentional, confirm below."
                    )
                    if st.button(
                        f"⚠️  Merge {old_proj} into {new_proj} ({len(_affected)} activities)",
                        type="primary",
                        key="settings_rename_confirm_merge",
                    ):
                        _settings_entries, n = rename_project(_settings_entries, old_proj, new_proj)
                        save_entries(_settings_entries)
                        # Update sidebar project selection if it was pointing at old name
                        if st.session_state.get("selected_project") == old_proj:
                            st.session_state["selected_project"] = new_proj
                        st.success(f"Merged **{old_proj}** into **{new_proj}** — {n} activities updated.")
                        st.rerun()
                else:
                    st.info(
                        f"**{len(_affected)}** {'activity' if len(_affected) == 1 else 'activities'} "
                        f"will be renamed from **{old_proj}** to **{new_proj}**."
                    )
                    # Preview first 5
                    if _affected:
                        with st.expander("Preview affected activities"):
                            for e in _affected[:10]:
                                old_wbs = e.get("wbs_id","")
                                numeric = strip_wbs_prefix(old_wbs)
                                new_wbs = f"{new_proj}.{numeric}" if numeric else new_proj
                                st.caption(
                                    f"`{e['activity_id']}`  {e['activity_name']}  "
                                    f"{old_wbs} → {new_wbs}"
                                )
                            if len(_affected) > 10:
                                st.caption(f"… and {len(_affected) - 10} more")

                    if st.button(
                        f"✅  Rename {old_proj} → {new_proj}",
                        type="primary",
                        key="settings_rename_confirm",
                    ):
                        _settings_entries, n = rename_project(_settings_entries, old_proj, new_proj)
                        save_entries(_settings_entries)
                        # Update sidebar selection if it was pointing at the old name
                        if st.session_state.get("selected_project") == old_proj:
                            st.session_state["selected_project"] = new_proj
                        st.success(
                            f"Renamed **{old_proj}** → **{new_proj}** — "
                            f"{n} {'activity' if n == 1 else 'activities'} updated."
                        )
                        st.rerun()

        st.divider()
        st.markdown("### Delete Project")
        st.write(
            "Permanently deletes all activities belonging to a project. "
            "Photo assignments for those activities are also removed. "
            "The image files themselves are kept."
        )

        if not _named_projects:
            st.info("No named projects to delete.")
        else:
            del_proj = st.selectbox(
                "Project to delete",
                options=_named_projects,
                key="settings_delete_proj",
            )

            _del_affected = [e for e in _settings_entries
                             if get_project_from_wbs(e.get("wbs_id","")) == del_proj]
            st.warning(
                f"⚠️ This will permanently delete **{len(_del_affected)}** "
                f"{'activity' if len(_del_affected) == 1 else 'activities'} "
                f"and all their associated data (comments, photo assignments). "
                f"**This cannot be undone.**"
            )

            if _del_affected:
                with st.expander("Show activities that will be deleted"):
                    for e in _del_affected:
                        st.caption(
                            f"`{e['activity_id']}`  {e['activity_name']}  "
                            f"WBS: {e.get('wbs_id','')}"
                        )

            # Two-step confirmation — user must type the project name
            _del_confirm_text = st.text_input(
                f'Type **{del_proj}** to confirm deletion',
                placeholder=del_proj,
                key="settings_delete_confirm_text",
            ).strip()

            if st.button(
                f"🗑  Delete project {del_proj}",
                type="primary",
                key="settings_delete_confirm_btn",
                disabled=(_del_confirm_text != del_proj),
            ):
                # Remove all activities for this project
                _del_aids = {e["activity_id"].upper() for e in _del_affected}
                surviving = [e for e in _settings_entries
                             if get_project_from_wbs(e.get("wbs_id","")) != del_proj]
                save_entries(surviving)

                # Remove photo assignments for deleted activities
                assignments = load_assignments()
                surviving_asgn = [
                    a for a in assignments
                    if not (a["activity_id"].upper() in _del_aids
                            and get_project_from_wbs(a.get("wbs_id","")) == del_proj)
                ]
                if len(surviving_asgn) != len(assignments):
                    save_assignments(surviving_asgn)

                # Remove report date setting for this project
                _del_proj_settings = load_project_settings()
                if del_proj in _del_proj_settings:
                    del _del_proj_settings[del_proj]
                    save_project_settings(_del_proj_settings)

                # Reset sidebar project selection if it was pointing at deleted project
                if st.session_state.get("selected_project") == del_proj:
                    st.session_state["selected_project"] = ""

                # Invalidate photo assignment caches
                for _k in ("photo_assignments","photo_to_aids","aid_to_pids","_assign_sig"):
                    st.session_state.pop(_k, None)

                st.success(
                    f"Project **{del_proj}** deleted — "
                    f"{len(_del_affected)} activities removed."
                )
                st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# SITE WALK — GANTT EDIT DIALOG
# ══════════════════════════════════════════════════════════════════════════════

@st.dialog("Edit Activity", width="large")
def gantt_edit_dialog(act: dict, rpt_date, today, sw_edits: dict, sw_activities: list):
    """Modal edit form launched when a bar is clicked on the Gantt chart."""
    aid    = act["activity_id"]
    merged = {**act, **sw_edits.get(aid, {})}

    st.subheader(f"`{aid}`  {merged.get('activity_name','')}")
    st.caption(f"WBS: {merged.get('wbs_id','—')}")
    st.divider()

    new_status = st.selectbox(
        "Status", STATUS_OPTIONS,
        index=STATUS_OPTIONS.index(merged.get("activity_status","Not Started"))
              if merged.get("activity_status") in STATUS_OPTIONS else 0,
        key=f"dlg_status_{aid}",
    )

    dlg_start_dt = dlg_finish_dt = None
    dlg_pct = 0
    dlg_rem = ""

    if new_status == "Not Started":
        st.info("% Complete set to 0 automatically.", icon="ℹ️")

    elif new_status == "In Progress":
        dlg_start_dt = datetime_inputs(
            "Actual Start *", key=f"dlg_start_{aid}", required=True,
            default_dt=iso_to_dt(merged.get("actual_start","")),
        )
        c_p, c_r = st.columns(2)
        with c_r:
            dlg_rem = st.text_input(
                "Remaining Duration (days) *",
                value=str(merged.get("remaining_dur","") or ""),
                placeholder="e.g. 5",
                key=f"dlg_rem_{aid}",
            ).strip()
        _dlg_suggested = None
        if rpt_date and dlg_rem and dlg_start_dt:
            _ef = expected_finish_date(rpt_date, dlg_rem)
            if _ef:
                _dlg_suggested = calc_duration_pct(dt_to_iso(dlg_start_dt), _ef, rpt_date)
        with c_p:
            _dlg_pct_def = _dlg_suggested if _dlg_suggested is not None else                            int(merged.get("pct_complete") or 0)
            dlg_pct = st.number_input(
                "Duration % Complete *", min_value=0, max_value=99, step=5,
                value=_dlg_pct_def, key=f"dlg_pct_{aid}",
            )
            if _dlg_suggested is not None:
                st.caption(f"💡 Calculated: **{_dlg_suggested}%**")

    elif new_status == "Completed":
        dlg_start_dt = datetime_inputs(
            "Actual Start *", key=f"dlg_start_{aid}", required=True,
            default_dt=iso_to_dt(merged.get("actual_start","")),
        )
        dlg_finish_dt = datetime_inputs(
            "Actual Finish *", key=f"dlg_finish_{aid}", required=True,
            default_dt=iso_to_dt(merged.get("actual_finish","")),
        )
        dlg_pct = 100
        dlg_rem = "0"
        st.info("% Complete set to 100, Remaining to 0.", icon="✅")

    # Comments
    st.divider()
    _dlg_existing_cmts = merged.get("_comments",[])
    if _dlg_existing_cmts:
        with st.expander(f"💬 {len(_dlg_existing_cmts)} existing comments"):
            for c in _dlg_existing_cmts:
                st.write(f"**{c['at']}** — {c['by']}")
                st.write(c["text"])
    dlg_new_cmt = st.text_area(
        "Add comment", placeholder="Enter progress notes...",
        height=80, key=f"dlg_cmt_{aid}",
    ).strip()

    st.divider()
    col_save, col_cancel = st.columns(2)
    with col_save:
        if st.button("💾 Save", type="primary", key=f"dlg_save_{aid}",
                     use_container_width=True):
            errs = []
            if new_status in ("In Progress","Completed") and not dlg_start_dt:
                errs.append("Actual Start is required.")
            if new_status == "Completed" and not dlg_finish_dt:
                errs.append("Actual Finish is required.")
            if new_status == "In Progress" and not dlg_rem:
                errs.append("Remaining Duration is required.")
            if dlg_start_dt and dlg_finish_dt and dlg_finish_dt < dlg_start_dt:
                errs.append("Actual Finish cannot be before Actual Start.")
            if errs:
                for e in errs: st.error(e)
            else:
                upd_cmts = list(merged.get("_comments",[]))
                if dlg_new_cmt:
                    upd_cmts.insert(0, {
                        "text": dlg_new_cmt,
                        "by":   st.session_state.display_name,
                        "at":   datetime.now().strftime("%d/%m/%Y %H:%M"),
                    })
                sw_edits[aid] = {
                    "activity_status": new_status,
                    "actual_start":    dt_to_iso(dlg_start_dt)  if dlg_start_dt  else "",
                    "actual_finish":   dt_to_iso(dlg_finish_dt) if dlg_finish_dt else "",
                    "pct_complete":    str(dlg_pct),
                    "remaining_dur":   dlg_rem,
                    "_comments":       upd_cmts,
                    "_submitted_at":   datetime.now().strftime("%d/%m/%Y %H:%M"),
                    "_submitted_by":   st.session_state.display_name,
                }
                st.session_state["sw_edits"] = sw_edits
                st.session_state["sw_gantt_clicked"] = None
                st.rerun()
    with col_cancel:
        if st.button("✖ Cancel", key=f"dlg_cancel_{aid}",
                     use_container_width=True):
            st.session_state["sw_gantt_clicked"] = None
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# TAB: SITE WALK
# ══════════════════════════════════════════════════════════════════════════════
# Shows all In Progress activities + Not Started activities whose predicted
# start is on or before today.  The user reviews each one, edits fields, and
# saves everything in a single commit at the end.

if "sitewalk" in tab_index:
    with tab_index["sitewalk"]:
        _sel_project = st.session_state.get("selected_project", "")
        _rpt_date_sw = get_report_date(_sel_project) if _sel_project else None
        _today       = date.today()

        st.subheader(" Site Walk")
        _last_walk_date = get_last_walk_date(_sel_project) if _sel_project else None
        st.caption(
            f"Project: **{_sel_project or '—'}**"
            + (f"  ·  Report date: **{_rpt_date_sw.strftime('%d/%m/%Y')}**"
               if _rpt_date_sw else "  ·  ⚠️ No report date set — set one in the sidebar.")
            + (f"  ·  Last walk: **{_last_walk_date.strftime('%d/%m/%Y')}**"
               if _last_walk_date else "  ·  Last walk: **—**")
        )

        # ── Start / Reset walk ────────────────────────────────────────────────
        walk_active = st.session_state.get("sw_active", False)

        if not walk_active:
            st.write(
                "Press **Start Walk** to load all activities that need reviewing — "
                "In Progress activities and any Not Started activities whose "
                "predicted start is on or before today."
            )
            if st.button("▶️  Start Walk", type="primary"):
                all_entries = load_entries()
                project_entries = filter_by_project(all_entries, _sel_project)

                # Collect qualifying activities — Task Dependent type only
                sw_activities = []
                for e in project_entries:
                    if e.get("task_type", "") != "Task Dependent":
                        continue
                    status = e.get("activity_status", "")
                    if status == "In Progress":
                        sw_activities.append(e.copy())
                    elif status == "Not Started":
                        pred = e.get("predicted_start", "")
                        if pred:
                            pd = iso_to_dt(pred)
                            if pd and pd.date() <= _today + timedelta(days=LOOKAHEADDAYS):
                                sw_activities.append(e.copy())


                if not sw_activities:
                    st.info(
                        "No activities require review — no In Progress activities "
                        "and no Not Started activities with a predicted start on or before today."
                    )
                else:
                    # Initialise walk state
                    st.session_state["sw_active"]     = True
                    st.session_state["sw_activities"] = sw_activities
                    st.session_state["sw_edits"]      = {}   # activity_id → edited fields
                    st.rerun()

        else:
            sw_activities = st.session_state.get("sw_activities", [])
            sw_edits      = st.session_state.get("sw_edits", {})

            # ── Sort + search controls ────────────────────────────────────────
            ctl_col1, ctl_col2, ctl_col3 = st.columns([2, 2, 1])
            with ctl_col1:
                sw_search = st.text_input(
                    "Search",
                    placeholder="Activity ID or name…",
                    key="sw_search",
                ).strip().lower()
            with ctl_col2:
                sw_sort = st.selectbox(
                    "Sort by",
                    ["WBS Code", "Activity ID"],
                    key="sw_sort",
                )
            with ctl_col3:
                sw_asc = st.radio(
                    "Order", ["↑ Asc", "↓ Desc"],
                    key="sw_order", horizontal=True,
                ) == "↑ Asc"

            # Filter
            visible_acts = [
                a for a in sw_activities
                if not sw_search
                or sw_search in a.get("activity_id","").lower()
                or sw_search in a.get("activity_name","").lower()
            ]

            # Sort
            def _sw_wbs_key(e):
                segs = []
                for p in str(e.get("wbs_id","") or "").split("."):
                    try:    segs.append((0, int(p)))
                    except: segs.append((1, p.lower()))
                return segs or [(1,"")]

            if sw_sort == "WBS Code":
                visible_acts = sorted(visible_acts, key=_sw_wbs_key, reverse=not sw_asc)
            else:
                visible_acts = sorted(visible_acts,
                                      key=lambda e: e.get("activity_id","").upper(),
                                      reverse=not sw_asc)

            st.caption(
                f"Showing **{len(visible_acts)}** of **{len(sw_activities)}** activities"
                + (f" — {len(sw_edits)} edited" if sw_edits else "")
            )

            # ── Gantt chart toggle ────────────────────────────────────────────
            show_gantt = st.toggle("📊 Show Gantt Chart", value=True, key="sw_gantt_toggle")

            if show_gantt and visible_acts:
                # Collect bar data for each activity
                from datetime import timedelta as _td
                _gantt_bars = []
                for _ga in visible_acts:
                    _aid    = _ga["activity_id"]
                    _status = _ga.get("activity_status","")
                    _start  = None
                    _end    = None

                    if _status == "Completed":
                        _start = iso_to_dt(_ga.get("actual_start",""))
                        _end   = iso_to_dt(_ga.get("actual_finish",""))
                        if _start: _start = _start.date()
                        if _end:   _end   = _end.date()

                    elif _status == "In Progress":
                        _rpt = _rpt_date_sw or _today
                        _start = _rpt
                        try:
                            _rem_days = int(float(_ga.get("remaining_dur","0") or 0))
                        except ValueError:
                            _rem_days = 0
                        _end = _add_working_days(_rpt, _rem_days)

                    elif _status == "Not Started":
                        _pred = iso_to_dt(_ga.get("predicted_start",""))
                        if _pred:
                            _start = _pred.date()
                            try:
                                _rem_days = int(float(_ga.get("remaining_dur","0") or 0))
                            except ValueError:
                                _rem_days = 0
                            _end = _add_working_days(_start, _rem_days)

                    if _start and _end and _end >= _start:
                        _gantt_bars.append({
                            "aid":    _aid,
                            "name":   _ga.get("activity_name",""),
                            "status": _status,
                            "start":  _start,
                            "end":    _end,
                        })

                if not _gantt_bars:
                    st.info("No activities have enough date data to display on the chart.")
                else:
                    # Chart dimensions
                    _all_starts = [b["start"] for b in _gantt_bars]
                    _all_ends   = [b["end"]   for b in _gantt_bars]
                    _chart_start = min(_all_starts)
                    _chart_end   = max(_all_ends)
                    _total_days  = max((_chart_end - _chart_start).days + 1, 1)

                    _BAR_H    = 36    # bar height px — taller for two-line labels
                    _ROW_GAP  = 10    # gap between rows px
                    _LABEL_W  = 220   # left label column — wide enough for full names
                    _CHART_W  = 820   # chart area width px
                    _HEADER_H = 52    # header row height — two label rows
                    _ROW_H    = _BAR_H + _ROW_GAP
                    _SVG_H    = _HEADER_H + len(_gantt_bars) * _ROW_H + 20
                    _SVG_W    = _LABEL_W + _CHART_W + 10

                    _COLOURS = {
                        "Completed":   "#16a34a",
                        "In Progress": "#d97706",
                        "Not Started": "#6b7280",
                    }

                    def _x(d):
                        return _LABEL_W + int(((d - _chart_start).days / _total_days) * _CHART_W)

                    # Build SVG
                    _svg_parts = [
                        f'<svg xmlns="http://www.w3.org/2000/svg" '
                        f'width="{_SVG_W}" height="{_SVG_H}" '
                        f'style="font-family:Arial,sans-serif;background:#f8f9fb">',
                        # Clip path keeps label text inside the label column
                        f'<defs><clipPath id="label-clip">'
                        f'<rect x="0" y="0" width="{_LABEL_W - 6}" height="{_SVG_H}"/>'
                        f'</clipPath></defs>',
                    ]

                    # Vertical separator between label column and chart
                    _svg_parts.append(
                        f'<line x1="{_LABEL_W}" y1="0" x2="{_LABEL_W}" y2="{_SVG_H}" '
                        f'stroke="#d1d5db" stroke-width="1"/>'
                    )

                    # Background grid lines — every week (Monday)
                    # Date labels — every 2 weeks to avoid overlap
                    # Minimum px between labels to prevent crowding
                    _MIN_LABEL_GAP = 55
                    _week_num      = 0
                    _last_label_x  = -999
                    _cur = _chart_start
                    while _cur <= _chart_end:
                        if _cur.weekday() == 0:  # Monday
                            _gx = _x(_cur)
                            # Gridline every week
                            _svg_parts.append(
                                f'<line x1="{_gx}" y1="{_HEADER_H}" '
                                f'x2="{_gx}" y2="{_SVG_H}" '
                                f'stroke="#e2e8f0" stroke-width="1"/>'
                            )
                            # Date label only when far enough from last label
                            if _gx - _last_label_x >= _MIN_LABEL_GAP:
                                _lbl = _cur.strftime("%d %b")
                                _svg_parts.append(
                                    f'<text x="{_gx+2}" y="{_HEADER_H-4}" '
                                    f'font-size="10" fill="#94a3b8">{_lbl}</text>'
                                )
                                _last_label_x = _gx
                            _week_num += 1
                        _cur += _td(days=1)

                    # Bars
                    for _bi, _bar in enumerate(_gantt_bars):
                        _y    = _HEADER_H + _bi * _ROW_H + _ROW_GAP // 2
                        _x1   = _x(_bar["start"])
                        _x2   = _x(_bar["end"])
                        _bw   = max(_x2 - _x1, 4)
                        _col  = _COLOURS.get(_bar["status"], "#6b7280")
                        _is_staged = _bar["aid"] in sw_edits

                        # Row background
                        _bg = "#eef2f8" if _bi % 2 == 0 else "#ffffff"
                        _svg_parts.append(
                            f'<rect x="0" y="{_y}" width="{_SVG_W}" '
                            f'height="{_BAR_H}" fill="{_bg}"/>'
                        )

                        # Activity label — split into two lines if needed
                        _name = _bar["name"]
                        _max_chars = 28   # chars that fit in _LABEL_W at font-size 11
                        if len(_name) <= _max_chars:
                            # Single line — vertically centred
                            _svg_parts.append(
                                f'<text x="4" y="{_y + _BAR_H//2 + 4}" '
                                f'font-size="11" fill="#1C3557" font-weight="bold"'
                                f'clip-path="url(#label-clip)">{_name}</text>'
                            )
                        else:
                            # Two lines — split at last space before midpoint
                            _mid = _max_chars
                            _split = _name.rfind(" ", 0, _mid)
                            if _split == -1:
                                _split = _mid
                            _line1 = _name[:_split].strip()
                            _line2 = _name[_split:].strip()
                            if len(_line2) > _max_chars:
                                _line2 = _line2[:_max_chars-1] + "…"
                            _svg_parts.append(
                                f'<text x="4" y="{_y + _BAR_H//2 - 4}" '
                                f'font-size="11" fill="#1C3557" font-weight="bold"'
                                f'clip-path="url(#label-clip)">{_line1}</text>'
                            )
                            _svg_parts.append(
                                f'<text x="4" y="{_y + _BAR_H//2 + 10}" '
                                f'font-size="11" fill="#1C3557" font-weight="bold" '
                                f'clip-path="url(#label-clip)">{_line2}</text>'
                            )

                        # Bar
                        _stroke = "#1C3557" if _is_staged else "none"
                        _svg_parts.append(
                            f'<rect x="{_x1}" y="{_y+3}" width="{_bw}" '
                            f'height="{_BAR_H-6}" rx="3" '
                            f'fill="{_col}" stroke="{_stroke}" stroke-width="2" opacity="0.85"/>'
                        )

                        # Bar label — activity ID if bar wide enough
                        if _bw > 40:
                            _svg_parts.append(
                                f'<text x="{_x1+5}" y="{_y + _BAR_H//2 + 4}" '
                                f'font-size="10" fill="white" font-weight="bold"'
                                f'>{_bar["aid"]}</text>'
                            )

                        # Staged indicator
                        if _is_staged:
                            _svg_parts.append(
                                f'<text x="{_x2+4}" y="{_y + _BAR_H//2 + 4}" '
                                f'font-size="10" fill="#16a34a">✓</text>'
                            )

                    # Today + Data date lines drawn AFTER bars — appear on top
                    if _chart_start <= _today <= _chart_end:
                        _tx = _x(_today)
                        _svg_parts.append(
                            f'<line x1="{_tx}" y1="0" '
                            f'x2="{_tx}" y2="{_SVG_H}" '
                            f'stroke="#ef4444" stroke-width="2" stroke-dasharray="4,3"/>'
                        )
                        _svg_parts.append(
                            f'<text x="{_tx+3}" y="{_HEADER_H//2}" '
                            f'font-size="10" fill="#ef4444" font-weight="bold">Today</text>'
                        )

                    if _rpt_date_sw and _chart_start <= _rpt_date_sw <= _chart_end:
                        _rx = _x(_rpt_date_sw)
                        _svg_parts.append(
                            f'<line x1="{_rx}" y1="0" '
                            f'x2="{_rx}" y2="{_SVG_H}" '
                            f'stroke="#1d4ed8" stroke-width="2" stroke-dasharray="6,3"/>'
                        )
                        _svg_parts.append(
                            f'<text x="{_rx+3}" y="{_HEADER_H//2}" '
                            f'font-size="10" fill="#1d4ed8" font-weight="bold">Data date</text>'
                        )

                    # Legend
                    _lx = _LABEL_W + 4
                    _ly = _SVG_H - 14
                    for _ls, _lc in _COLOURS.items():
                        _svg_parts.append(
                            f'<rect x="{_lx}" y="{_ly-8}" width="12" height="10" '
                            f'rx="2" fill="{_lc}"/>'
                        )
                        _svg_parts.append(
                            f'<text x="{_lx+15}" y="{_ly}" font-size="10" '
                            f'fill="#374151">{_ls}</text>'
                        )
                        _lx += 100

                    _svg_parts.append("</svg>")

                    # Add onclick to each bar that scrolls to the activity card below
                    _svg_final = "".join(_svg_parts)
                    for _bar in _gantt_bars:
                        _safe_aid   = _bar["aid"].replace("'","")
                        _bar_colour = _COLOURS.get(_bar["status"], "#6b7280")
                        _bar_colour = _COLOURS.get(_bar["status"], "#6b7280")
                        _scroll_fn  = "scrollIntoView({behavior:'smooth',block:'center'})"
                        _el_id      = "'sw_act_" + _safe_aid + "'"
                        _scroll_js  = (
                            "onclick=\"window.parent.document.getElementById("
                            + _el_id + ")." + _scroll_fn + "\""
                        )
                        _old = f'fill="{_bar_colour}" stroke='
                        _new = f'fill="{_bar_colour}" style="cursor:pointer" {_scroll_js} stroke='
                        _svg_final = _svg_final.replace(_old, _new, 1)

                    st.caption("Click a bar to scroll to that activity below.")
                    import streamlit.components.v1 as _components
                    _components.html(
                        f'<div style="overflow-x:auto">{_svg_final}</div>',
                        height=_SVG_H + 30,
                        scrolling=False,
                    )

            st.divider()

            if not visible_acts:
                st.info("No activities match the search.")

            # ── Per-activity edit forms ───────────────────────────────────────
            for act in visible_acts:
                aid    = act["activity_id"]
                # Merge any already-saved edits back in for display
                merged = {**act, **sw_edits.get(aid, {})}

                status_colour = STATUS_COLOUR.get(merged.get("activity_status",""), "#6b7280")
                _is_staged = aid in sw_edits

                # Anchor so Gantt bar clicks can scrollIntoView here
                st.markdown(
                    f'<div id="sw_act_{aid}" style="scroll-margin-top:80px"></div>',
                    unsafe_allow_html=True,
                )
                with st.container(border=True):
                    # Header — show persistent staged indicator if reviewed
                    h_left, h_right = st.columns([4, 1])
                    with h_left:
                        st.write(f"**`{aid}`**  {merged.get('activity_name','')}")
                        st.caption(f"WBS: {merged.get('wbs_id','—')}")
                    with h_right:
                        if _is_staged:
                            st.success("✅ Staged")
                        else:
                            st.write(merged.get("activity_status",""))

                    # Edit form
                    new_status = st.selectbox(
                        "Status",
                        STATUS_OPTIONS,
                        index=STATUS_OPTIONS.index(merged.get("activity_status","Not Started"))
                              if merged.get("activity_status") in STATUS_OPTIONS else 0,
                        key=f"sw_status_{aid}",
                    )

                    sw_start_dt = sw_finish_dt = None
                    sw_pct = 0
                    sw_rem = 0

                        
                    if new_status == "In Progress":
                        sw_start_dt = datetime_inputs(
                            "Actual Start *", key=f"sw_start_{aid}", required=True,
                            default_dt=iso_to_dt(merged.get("actual_start","")),
                        )
                        
                        # ── Input mode selector ───────────────────────────────
                        _ip_mode_sw = st.radio(
                            "Update by",
                            ["% Complete", "Remaining Duration", "Expected Finish",
                             "Physical % Complete"],
                            horizontal=True,
                            key=f"sw_ip_mode_{aid}",
                        )

                        sw_pct = int(merged.get("pct_complete") or 0)
                        sw_rem = str(merged.get("remaining_dur","") or "")

                        _rpt = _rpt_date_sw or _today

                        if _ip_mode_sw == "% Complete":
                            sw_pct = st.number_input(
                                "Duration % Complete *",
                                min_value=0, max_value=99, step=5,
                                value=sw_pct,
                                key=f"sw_pct_{aid}",
                            )
                            if sw_start_dt and sw_pct > 0:
                                _elapsed  = working_days_between(sw_start_dt.date(), _rpt) + 1
                                _total    = int(_elapsed / (sw_pct / 100))
                                _rem_calc = max(0, _total - _elapsed)
                                sw_rem    = str(_rem_calc)
                                _ef_calc  = _add_working_days(_rpt, _rem_calc)
                                st.caption(
                                    f"💡 Remaining: **{_rem_calc} days**  ·  "
                                    f"Expected finish: **{_ef_calc.strftime('%d/%m/%Y')}**"
                                )

                        elif _ip_mode_sw == "Remaining Duration":
                            sw_rem = st.text_input(
                                "Remaining Duration (days) *",
                                value=sw_rem,
                                placeholder="e.g. 5",
                                key=f"sw_rem_{aid}",
                            ).strip()
                            if sw_rem and sw_start_dt:
                                _ef_calc  = expected_finish_date(_rpt, sw_rem)
                                _pct_calc = calc_duration_pct(
                                    dt_to_iso(sw_start_dt), _ef_calc, _rpt
                                ) if _ef_calc else None
                                sw_pct = _pct_calc if _pct_calc is not None else sw_pct
                                if _ef_calc:
                                    st.caption(
                                        f"💡 % Complete: **{sw_pct}%**  ·  "
                                        f"Expected finish: **{_ef_calc.strftime('%d/%m/%Y')}**"
                                    )

                        elif _ip_mode_sw == "Expected Finish":
                            _cur_ef = _add_working_days(
                                _rpt, int(float(sw_rem)) if sw_rem else 0
                            ) if sw_rem else _rpt
                            _ef_input = st.date_input(
                                "Expected Finish *",
                                value=_cur_ef,
                                format="DD/MM/YYYY",
                                key=f"sw_ef_{aid}",
                            )
                            _rem_calc = working_days_between(_rpt, _ef_input)
                            sw_rem    = str(_rem_calc)
                            _pct_calc = calc_duration_pct(
                                dt_to_iso(sw_start_dt), _ef_input, _rpt
                            ) if sw_start_dt else None
                            sw_pct = _pct_calc if _pct_calc is not None else sw_pct
                            st.caption(
                                f"💡 Remaining: **{_rem_calc} days**  ·  "
                                f"% Complete: **{sw_pct}%**"
                            )

                        else:  # Physical % Complete
                            st.caption(
                                "Enter the physical % complete directly. "
                                "Then select how to specify the remaining schedule."
                            )
                            sw_pct = st.number_input(
                                "Physical % Complete *",
                                min_value=0, max_value=99, step=1,
                                value=sw_pct,
                                key=f"sw_phys_pct_{aid}",
                            )
                            # Second selector — how to supply remaining schedule info
                            _phys_sub = st.selectbox(
                                "Also specify",
                                ["Remaining Duration", "Expected Finish",
                                 "Duration % Complete"],
                                key=f"sw_phys_sub_{aid}",
                            )
                            if _phys_sub == "Remaining Duration":
                                sw_rem = st.text_input(
                                    "Remaining Duration (days)",
                                    value=sw_rem,
                                    placeholder="e.g. 5",
                                    key=f"sw_phys_rem_{aid}",
                                ).strip()
                                if sw_rem:
                                    _ef_calc = expected_finish_date(_rpt, sw_rem)
                                    if _ef_calc:
                                        st.caption(
                                            f"💡 Expected finish: "
                                            f"**{_ef_calc.strftime('%d/%m/%Y')}**"
                                        )

                            elif _phys_sub == "Expected Finish":
                                _cur_ef = _add_working_days(
                                    _rpt, int(float(sw_rem)) if sw_rem else 0
                                ) if sw_rem else _rpt
                                _ef_input = st.date_input(
                                    "Expected Finish",
                                    value=_cur_ef,
                                    format="DD/MM/YYYY",
                                    key=f"sw_phys_ef_{aid}",
                                )
                                _rem_calc = working_days_between(_rpt, _ef_input)
                                sw_rem    = str(_rem_calc)
                                st.caption(
                                    f"💡 Remaining: **{_rem_calc} days**"
                                )

                            else:  # Duration % Complete
                                _dur_pct = st.number_input(
                                    "Duration % Complete",
                                    min_value=0, max_value=99, step=5,
                                    value=int(merged.get("pct_complete") or 0),
                                    key=f"sw_phys_dur_{aid}",
                                )
                                if sw_start_dt and _dur_pct > 0:
                                    _elapsed  = working_days_between(
                                        sw_start_dt.date(), _rpt) + 1
                                    _total    = int(_elapsed / (_dur_pct / 100))
                                    _rem_calc = max(0, _total - _elapsed)
                                    sw_rem    = str(_rem_calc)
                                    _ef_calc  = _add_working_days(_rpt, _rem_calc)
                                    st.caption(
                                        f"💡 Remaining: **{_rem_calc} days**  ·  "
                                        f"Expected finish: **{_ef_calc.strftime('%d/%m/%Y')}**"
                                    )

                    elif new_status == "Completed":
                        sw_start_dt = datetime_inputs(
                            "Actual Start *", key=f"sw_start_{aid}", required=True,
                            default_dt=iso_to_dt(merged.get("actual_start","")),
                        )
                        sw_finish_dt = datetime_inputs(
                            "Actual Finish *", key=f"sw_finish_{aid}", required=True,
                            default_dt=iso_to_dt(merged.get("actual_finish","")),
                        )
                        sw_pct = 100
                        sw_rem = "0"
                        st.info("% Complete set to 100, Remaining to 0.", icon="✅")

                    # Comments
                    with st.expander("💬  Comments"):
                        _sw_existing_cmts = merged.get("_comments", [])
                        if _sw_existing_cmts:
                            for c in _sw_existing_cmts:
                                st.write(f"**{c['at']}** — {c['by']}")
                                st.write(c["text"])
                                st.divider()
                        sw_new_cmt = st.text_area(
                            "Add comment",
                            placeholder="Enter progress notes...",
                            height=80,
                            key=f"sw_comment_{aid}",
                            label_visibility="collapsed",
                        ).strip()

                    # ── Photo capture ─────────────────────────────────────────
                    with st.expander("📷  Add Photo"):
                        st.caption(
                            "Photo will be saved to the Photo Log and automatically "
                            "assigned to this activity."
                        )
                        sw_photo_date = st.date_input(
                            "Photo date",
                            value=_today,
                            format="DD/MM/YYYY",
                            key=f"sw_photo_date_{aid}",
                        )
                        sw_photo_comment = st.text_input(
                            "Photo comment (optional)",
                            max_chars=100,
                            placeholder="e.g. North face formwork installed",
                            key=f"sw_photo_comment_{aid}",
                        ).strip()
                        sw_photo_file = st.file_uploader(
                            "Choose image",
                            type=["jpg","jpeg","png","webp"],
                            key=f"sw_photo_file_{aid}",
                        )
                        inject_camera_uploader(f"sw_photo_file_{aid}")
                        if sw_photo_file:
                            # Read once and cache — reused in upload block below
                            _sw_fb = sw_photo_file.read()
                            st.session_state[f"sw_photo_bytes_{aid}"] = _sw_fb
                            if _PILLOW:
                                _sw_prev = ImageOps.exif_transpose(
                                    Image.open(io.BytesIO(_sw_fb))
                                )
                                st.image(_sw_prev, width=300)
                            else:
                                st.image(io.BytesIO(_sw_fb), width=300)

                    # Mark as reviewed button — stages changes to sw_edits
                    if st.button(f"✅  Mark reviewed", key=f"sw_review_{aid}"):
                        # Validate
                        errs = []
                        if new_status in ("In Progress","Completed") and not sw_start_dt:
                            errs.append("Actual Start is required.")
                        if new_status == "Completed" and not sw_finish_dt:
                            errs.append("Actual Finish is required.")
                        if new_status == "In Progress" and not sw_rem:
                            errs.append("Remaining Duration is required.")
                        if sw_start_dt and sw_finish_dt and sw_finish_dt < sw_start_dt:
                            errs.append("Actual Finish cannot be before Actual Start.")

                        if errs:
                            for err in errs:
                                st.error(err)
                        else:
                            # Build updated comments
                            upd_cmts = list(merged.get("_comments", []))
                            if sw_new_cmt:
                                upd_cmts.insert(0, {
                                    "text": sw_new_cmt,
                                    "by":   st.session_state.display_name,
                                    "at":   datetime.now().strftime("%d/%m/%Y %H:%M"),
                                })

                            # Upload and assign photo if provided
                            if sw_photo_file:
                                # Use cached bytes from preview (file stream already consumed)
                                _sw_fb = st.session_state.get(f"sw_photo_bytes_{aid}", b"")
                                if _sw_fb:
                                    _sw_record = upload_photo(
                                        photo_date    = sw_photo_date,
                                        comment       = sw_photo_comment or f"Site walk — {aid}",
                                        file_bytes    = _sw_fb,
                                        original_name = sw_photo_file.name,
                                        uploaded_by   = st.session_state.display_name,
                                    )
                                    # Assign to this activity — pass wbs_id explicitly
                                    # so the assignment is stored with the correct project
                                    _sw_wbs = act.get("wbs_id", "")
                                    _sw_entries_with_wbs = [
                                        e for e in load_entries()
                                        if e.get("activity_id","").upper() == aid.upper()
                                        and e.get("wbs_id","") == _sw_wbs
                                    ]
                                    if not _sw_entries_with_wbs:
                                        _sw_entries_with_wbs = [act]
                                    assign_photo(
                                        _sw_record["id"], [aid],
                                        st.session_state.display_name,
                                        _sw_entries_with_wbs,
                                    )
                                    # Invalidate photo caches
                                    load_image_bytes.clear()
                                    build_photo_backup.clear()
                                    for _k in ("photo_list","photo_map","photo_assignments",
                                               "photo_to_aids","aid_to_pids","_assign_sig"):
                                        st.session_state.pop(_k, None)

                            sw_edits[aid] = {
                                "activity_status": new_status,
                                "actual_start":    dt_to_iso(sw_start_dt)  if sw_start_dt  else "",
                                "actual_finish":   dt_to_iso(sw_finish_dt) if sw_finish_dt else "",
                                "pct_complete":    str(sw_pct),
                                "remaining_dur":   sw_rem,
                                "_comments":       upd_cmts,
                                "_submitted_at":   datetime.now().strftime("%d/%m/%Y %H:%M"),
                                "_submitted_by":   st.session_state.display_name,
                            }
                            st.session_state["sw_edits"] = sw_edits
                            st.rerun()

            # ── Add activity outside filtered list ───────────────────────────
            st.divider()
            with st.expander("🔍  Add an activity outside the walk list"):
                st.caption(
                    "Search for any activity in this project by name or ID "
                    "and add it to the walk for review."
                )
                all_project_entries = filter_by_project(load_entries(), _sel_project)
                sw_current_ids      = {a["activity_id"] for a in sw_activities}

                add_search = st.text_input(
                    "Search",
                    placeholder="Activity ID or name…",
                    key="sw_add_search",
                ).strip().lower()

                if add_search:
                    candidates = [
                        e for e in all_project_entries
                        if e["activity_id"] not in sw_current_ids
                        and (add_search in e.get("activity_id","").lower()
                             or add_search in e.get("activity_name","").lower())
                    ]
                    if not candidates:
                        st.caption("No matching activities found outside the current walk list.")
                    else:
                        for cand in candidates[:20]:
                            c_info, c_btn = st.columns([5, 1])
                            with c_info:
                                st.write(
                                    f"**`{cand['activity_id']}`**  {cand['activity_name']}  "
                                    f"— WBS: {cand.get('wbs_id','—')}  "
                                    f"— {cand.get('activity_status','')}"
                                )
                            with c_btn:
                                if st.button("Add", key=f"sw_add_{cand['activity_id']}"):
                                    sw_activities.append(cand.copy())
                                    st.session_state["sw_activities"] = sw_activities
                                    st.toast(
                                        f"Added {cand['activity_id']} to walk."
                                    )
                                    st.rerun()
                        if len(candidates) > 20:
                            st.caption(f"… and {len(candidates)-20} more — refine your search.")

            # ── Commit + Cancel controls ──────────────────────────────────────
            st.divider()
            n_edited = len(sw_edits)
            n_total  = len(sw_activities)

            commit_col, cancel_col = st.columns([2, 1])

            with commit_col:
                st.info(
                    f"**{n_edited}** of **{n_total}** activities reviewed and staged."
                    + (" Press **Complete Walk** to save all changes."
                       if n_edited else " Mark activities as reviewed above.")
                )
                if st.button(
                    f"💾  Submit Walk for Approval ({n_edited} updates)",
                    type="primary",
                    disabled=(n_edited == 0),
                    key="sw_commit",
                ):
                    # Build rows table for the notification body
                    _notif_rows = []
                    for _eid, _edata in sw_edits.items():
                        _orig = next((e for e in sw_activities if e["activity_id"] == _eid), {})
                        _notif_rows.append({
                            "Activity ID":      _eid,
                            "Activity Name":    _orig.get("activity_name",""),
                            "Status":           _edata.get("activity_status","—"),
                            "% Complete":       _edata.get("pct_complete","—"),
                            "Remaining (days)": _edata.get("remaining_dur","—"),
                            "Actual Start":     display_dt(_edata.get("actual_start","")),
                            "Actual Finish":    display_dt(_edata.get("actual_finish","")),
                        })

                    # Store as pending approval — do NOT write to entries yet
                    create_notification(
                        created_by = st.session_state.display_name,
                        project    = _sel_project,
                        title      = f"Site Walk Approval — {_sel_project} — {_today.strftime('%d/%m/%Y')}",
                        body       = (
                            f"Walk submitted by **{st.session_state.display_name}**"
                            f" on {_today.strftime('%d/%m/%Y')}. Awaiting approval."
                        ),
                        rows       = _notif_rows,
                        status     = "pending",
                        edits      = dict(sw_edits),
                        recipients = get_project_admin_recipients(_sel_project),
                    )

                    for _k in ("sw_active", "sw_activities", "sw_edits"):
                        st.session_state.pop(_k, None)

                    st.success(
                        f"✅ Walk submitted — {n_edited} activities pending admin approval."
                    )
                    st.rerun()

            with cancel_col:
                if st.button("✖️  Cancel Walk", key="sw_cancel"):
                    for _k in ("sw_active", "sw_activities", "sw_edits"):
                        st.session_state.pop(_k, None)
                    st.rerun()
